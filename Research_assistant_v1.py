"""
Research Assistant (v1)

Major subsystems:
- Streamlit UI and styling
- LLM orchestration (Ollama/GLM), streaming, and prompt chains
- Wikipedia scouting and summarization
- Academic search + relevance scoring
- Project/session persistence with migration support
- PDF/RIS download tracking and annotation tools
"""

import streamlit as st
from streamlit_autorefresh import st_autorefresh
import requests
import json
import re
import time
import threading
import os
import shutil
from pathlib import Path
from datetime import datetime
import sqlite3
import hashlib
from semanticscholar import SemanticScholar
import wikipedia
import concurrent.futures
import html
import uuid
import markdown
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from streamlit_pdf_viewer import pdf_viewer
from habanero import Crossref
from difflib import SequenceMatcher
from urllib.parse import urlsplit
import xml.etree.ElementTree as ET

# -----------------------------------------------------------------------------
# Chromium automation state (reused across requests to avoid repeated startup).
# -----------------------------------------------------------------------------
_chrome_driver = None
_chrome_download_path = None
_chrome_binary_location = None
_chrome_profile_dir = None
_chrome_driver_path = None
_chrome_use_remote_debugging = False
_chrome_remote_debug_port = None
_chrome_reduce_automation = False
_download_manifest_lock = threading.RLock()
_FLASH_TTL_SECONDS = 5

# -----------------------------------------------------------------------------
# Flash messaging system (transient UI notifications with auto-expiration).
# -----------------------------------------------------------------------------
def _flash_store():
    return st.session_state.setdefault("_flash_messages", {})

def _normalize_flash_level(level: str) -> str:
    value = str(level or "").strip().lower()
    if value in {"success", "error", "warning", "info"}:
        return value
    return "info"

def _render_flash_message(entry: dict) -> None:
    msg = entry.get("message", "")
    level = _normalize_flash_level(entry.get("level"))
    if level == "success":
        st.success(msg)
    elif level == "error":
        st.error(msg)
    elif level == "warning":
        st.warning(msg)
    else:
        st.info(msg)

def _prune_flash_messages() -> None:
    store = st.session_state.get("_flash_messages")
    if not isinstance(store, dict):
        return
    now = time.time()
    expired = [k for k, v in store.items() if v.get("expires_at", 0) <= now]
    for key in expired:
        store.pop(key, None)

def flash(message, level="info", seconds=_FLASH_TTL_SECONDS, key=None, scope="main"):
    if not message:
        return
    level = _normalize_flash_level(level)
    scope = str(scope or "main")
    store = _flash_store()
    now = time.time()
    msg_key = key or f"{scope}:{level}:{message}"
    entry = store.get(msg_key)
    if not entry or entry.get("expires_at", 0) <= now:
        store[msg_key] = {
            "message": str(message),
            "level": level,
            "expires_at": now + max(1, int(seconds)),
            "scope": scope,
        }
        entry = store[msg_key]
    if entry.get("expires_at", 0) > now:
        _render_flash_message(entry)
    _prune_flash_messages()

def _render_flash_messages(scope="main") -> None:
    store = st.session_state.get("_flash_messages")
    if not isinstance(store, dict):
        return
    now = time.time()
    for entry in store.values():
        if entry.get("scope", "main") != scope:
            continue
        if entry.get("expires_at", 0) > now:
            _render_flash_message(entry)
    _prune_flash_messages()

def _flash_autorefresh() -> None:
    store = st.session_state.get("_flash_messages")
    if not isinstance(store, dict):
        return
    now = time.time()
    if any(entry.get("expires_at", 0) > now for entry in store.values()):
        st_autorefresh(interval=1000, limit=1, key="flash_refresh")

# -----------------------------------------------------------------------------
# PDF opening/download helpers (Chromium automation + tab reuse).
# -----------------------------------------------------------------------------
def open_pdf_in_chrome(pdf_url, download_path, paper_title=None, chromium_path=None, profile_dir=None,
                       chromedriver_path=None, use_remote_debugging=False, remote_debug_port=None,
                       reduce_automation=False):
    """Open a PDF URL in Chrome with custom download directory.
    
    Opens Chrome configured to download PDFs to the specified directory.
    Reuses the same browser window and opens new tabs for subsequent PDFs.
    
    Args:
        pdf_url: URL of the PDF to open
        download_path: Directory path where downloads should be saved
        paper_title: Optional title of the paper (for tracking)
    
    Returns:
        bool: True if successfully opened, False otherwise
    """
    global _chrome_driver, _chrome_download_path, _chrome_binary_location, _chrome_profile_dir, _chrome_driver_path
    global _chrome_use_remote_debugging, _chrome_remote_debug_port, _chrome_reduce_automation

    chromium_path = chromium_path or None
    profile_dir = profile_dir or None
    chromedriver_path = chromedriver_path or None
    use_remote_debugging = bool(use_remote_debugging)
    remote_debug_port = int(remote_debug_port) if remote_debug_port else None
    reduce_automation = bool(reduce_automation)
    
    absolute_path = os.path.abspath(download_path)
    os.makedirs(absolute_path, exist_ok=True)
    
    try:
        # Check if we have an existing driver and it's still alive
        if _chrome_driver is not None:
            try:
                # Test if the driver is still responding
                _chrome_driver.current_url
                
                # Rebuild driver if Chromium settings changed
                if (_chrome_binary_location != chromium_path or
                        _chrome_profile_dir != profile_dir or
                        _chrome_driver_path != chromedriver_path or
                        _chrome_use_remote_debugging != use_remote_debugging or
                        _chrome_remote_debug_port != remote_debug_port or
                        _chrome_reduce_automation != reduce_automation):
                    try:
                        _chrome_driver.quit()
                    except:
                        pass
                    _chrome_driver = None
                else:
                    # Update download path if needed
                    if _chrome_download_path != absolute_path:
                        try:
                            _chrome_driver.execute_cdp_cmd(
                                "Page.setDownloadBehavior",
                                {"behavior": "allow", "downloadPath": absolute_path}
                            )
                        except Exception:
                            pass
                        _chrome_download_path = absolute_path

                    # Open new tab and navigate
                    _chrome_driver.execute_script("window.open('');")
                    _chrome_driver.switch_to.window(_chrome_driver.window_handles[-1])
                    _chrome_driver.get(pdf_url)
                    return True
            except:
                # Driver is dead, reset it
                _chrome_driver = None
        
        # Create new Chrome instance
        chrome_options = Options()
        if use_remote_debugging and remote_debug_port:
            chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{remote_debug_port}")
        else:
            if chromium_path:
                chrome_options.binary_location = chromium_path
            if profile_dir:
                chrome_options.add_argument(f"--user-data-dir={profile_dir}")
        if reduce_automation:
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option("useAutomationExtension", False)
        prefs = {
            "download.default_directory": absolute_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option("detach", True)

        if chromedriver_path:
            service = Service(chromedriver_path)
        else:
            service = Service(ChromeDriverManager().install())
        _chrome_driver = webdriver.Chrome(service=service, options=chrome_options)
        _chrome_download_path = absolute_path
        _chrome_binary_location = chromium_path
        _chrome_profile_dir = profile_dir
        _chrome_driver_path = chromedriver_path
        _chrome_use_remote_debugging = use_remote_debugging
        _chrome_remote_debug_port = remote_debug_port
        _chrome_reduce_automation = reduce_automation
        try:
            _chrome_driver.execute_cdp_cmd(
                "Page.setDownloadBehavior",
                {"behavior": "allow", "downloadPath": absolute_path}
            )
        except Exception:
            pass
        _chrome_driver.get(pdf_url)
        return True
    except Exception as e:
        return False


def fetch_elsevier_pdf_url(doi, api_key):
    """
    Attempts to fetch PDF from Elsevier API using the provided API Key.
    Returns the Elsevier PDF URL if successful, None otherwise.
    
    Note: This function only verifies the PDF is accessible. Actual download
    happens during export.
    
    Args:
        doi: The DOI of the paper (with or without https://doi.org/ prefix)
        api_key: Elsevier API key
        
    Returns:
        str: The Elsevier API URL if PDF is accessible, None otherwise
    """
    if not doi or not api_key:
        return None
        
    headers = {
        "X-ELS-APIKey": api_key,
        "Accept": "application/pdf"
    }
    # Normalize DOI format for API requests.
    clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
    url = f"https://api.elsevier.com/content/article/doi/{clean_doi}"
    
    try:
        # Use a HEAD request to validate accessibility without downloading.
        response = requests.head(url, headers=headers, timeout=5)
        if response.status_code == 200:
            return url
        return None
    except Exception:
        return None


# -----------------------------------------------------------------------------
# Metadata enrichment via OpenAlex (concepts, venue, and impact stats).
# -----------------------------------------------------------------------------
def get_openalex_metadata(doi, email=None):
    """
    Fetch ALL metadata for a paper using OpenAlex API in a SINGLE call:
    - ML-tagged concepts
    - Venue/journal name
    - Impact metrics (2-year mean citedness)
    
    Args:
        doi: The DOI of the paper
        email: Optional email for polite pool access
        
    Returns:
        dict: {
            'concepts': ['Deep Learning', 'Neural Networks'],
            'venue': 'Nature',
            'venue_type': 'journal',
            'impact': 5.2  # 2-year mean citedness
        }
    """
    if not doi:
        return {'concepts': [], 'venue': None, 'venue_type': None, 'impact': 0}
    
    # Normalize DOI before calling OpenAlex.
    clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
    url = f"https://api.openalex.org/works/doi:{clean_doi}"
    
    params = {}
    if email:
        params['mailto'] = email
    
    try:
        response = requests.get(url, params=params, timeout=5)
        if response.status_code == 200:
            data = response.json()
            
            # Extract concepts (top 3 by score)
            concepts = data.get('concepts', [])
            sorted_concepts = sorted(concepts, key=lambda x: x.get('score', 0), reverse=True)
            concept_names = [c.get('display_name', '') for c in sorted_concepts[:3] if c.get('display_name')]
            
            # Extract venue/source info and impact.
            primary_location = data.get('primary_location', {}) or {}
            source = primary_location.get('source', {}) or {}
            venue_name = source.get('display_name')
            venue_type = source.get('type')  # e.g., "journal", "repository", "conference"
            
            # Extract impact from source summary stats
            summary_stats = source.get('summary_stats', {}) or {}
            impact = summary_stats.get('2yr_mean_citedness', 0) or 0
            
            return {
                'concepts': concept_names,
                'venue': venue_name,
                'venue_type': venue_type,
                'impact': impact
            }
        return {'concepts': [], 'venue': None, 'venue_type': None, 'impact': 0}
    except Exception:
        return {'concepts': [], 'venue': None, 'venue_type': None, 'impact': 0}


# =============================================================================
# UI CONFIGURATION & GLOBAL STYLES
# =============================================================================
st.set_page_config(page_title="Research Assistant", layout="wide", page_icon=None)

st.markdown("""
    <style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }

    .paper-card-success {
        background-color: #161b22;
        border-left: 4px solid #2ea043;
        padding: 16px;
        margin-bottom: 12px;
        border-radius: 6px;
    }
    .paper-card-failed {
        background-color: #161b22;
        border-left: 4px solid #484f58;
        padding: 16px;
        margin-bottom: 12px;
        border-radius: 6px;
        opacity: 0.8;
    }

    .paper-title {
        font-size: 16px;
        font-weight: 600;
        margin-bottom: 6px;
        color: #f0f6fc;
        line-height: 1.4;
    }
    .paper-title a { text-decoration: none; color: #f0f6fc; }
    .paper-title a:hover { color: #58a6ff; }

    .paper-authors {
        font-size: 13px;
        color: #8b949e;
        margin-bottom: 8px;
    }

    .paper-meta {
        font-size: 11px;
        color: #8b949e;
        margin-bottom: 10px;
        padding-top: 8px;
        border-top: 1px solid #21262d;
    }

    .paper-abstract, .paper-tldr {
        font-size: 13px;
        line-height: 1.5;
        color: #c9d1d9;
        margin-top: 8px;
    }

    .paper-tldr {
        background: #0d1117;
        padding: 10px;
        border-radius: 4px;
        border-left: 3px solid #58a6ff;
    }

    .badge-high { background: #8957e5; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-med { background: #1f6feb; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-low { background: #21262d; color: #8b949e; padding: 2px 6px; border-radius: 3px; font-size: 10px; }
    .paper-card-success, .paper-card-failed {
        transition: all 0.3s ease-in-out;
    }
    .paper-card-success:hover, .paper-card-failed:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.3);
    }

    .tag-relevant { background: #2ea043; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-downloaded { background: #238636; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-pending { background: #30363d; color: #8b949e; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 500; }
    .badge-excel { background: #1f6feb; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-citation { background: #8957e5; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-paper-ref { background: #f59e0b; color: #0d1117; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 700; }
    
    /* PDF Source Badges */
    .badge-source-openalex { background: #238636; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 600; }
    .badge-source-elsevier { background: #238636; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 600; }
    .badge-source-arxiv { background: #b45309; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 600; }
    .badge-source-paywalled { background: #da3633; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: 600; }
    
    /* Paper Quality Metrics Badges */
    .badge-influence { background: #58a6ff; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 600; }
    .badge-concept { background: #8957e5; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px; font-weight: 500; }

    .paper-also-in {
        margin-top: 8px;
        padding: 4px 8px;
        border-radius: 4px;
        border: 1px solid #30363d;
        background: #0d1117;
        font-size: 11px;
        color: #8b949e;
    }

    /* Popover sizing for notes */
    div[data-testid="stPopoverContent"] {
        width: min(18vw, 220px);
        min-width: 160px;
    }
    div[data-testid="stPopoverContent"] textarea {
        max-width: 100%;
    }
    .note-metadata {
        margin-top: 8px;
        padding: 6px 8px;
        border-radius: 4px;
        border: 1px solid #30363d;
        background: #0d1117;
        font-size: 12px;
        color: #8b949e;
        max-width: 100%;
        width: 100%;
        overflow-wrap: anywhere;
        word-break: break-word;
        white-space: normal;
    }
    
    /* Subtle fade-in for newly rendered elements */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .element-container {
        animation: fadeIn 0.5s ease-out;
    }
    
    /* Agent log styling */
    .agent-log-card {
        background-color: #161b22;
        border: 1px solid #30363d;
        border-radius: 8px;
        padding: 16px;
        margin-bottom: 12px;
    }
    .agent-log-header {
        font-size: 14px;
        font-weight: 600;
        color: #58a6ff;
        margin-bottom: 12px;
        padding-bottom: 8px;
        border-bottom: 1px solid #21262d;
    }
    .agent-log-content {
        font-size: 13px;
        color: #c9d1d9;
        line-height: 1.6;
    }
    .agent-log-section {
        background: #0d1117;
        padding: 12px;
        border-radius: 6px;
        margin-top: 8px;
        border-left: 3px solid #1f6feb;
    }
    .query-tag {
        display: inline-block;
        background: #21262d;
        color: #8b949e;
        padding: 4px 8px;
        border-radius: 4px;
        font-size: 11px;
        margin: 2px 4px 2px 0;
    }
    .step-badge {
        display: inline-block;
        background: #1f6feb;
        color: white;
        padding: 3px 8px;
        border-radius: 10px;
        font-size: 10px;
        font-weight: 600;
        margin-right: 8px;
    }
    .web-result-item {
        background: #0d1117;
        border-left: 2px solid #58a6ff;
        padding: 8px 12px;
        margin: 6px 0;
        border-radius: 4px;
    }
    .web-result-title {
        color: #58a6ff;
        font-size: 12px;
        font-weight: 500;
    }
    .web-result-snippet {
        color: #8b949e;
        font-size: 11px;
        margin-top: 4px;
    }
    
    /* Unified control panel */
    .control-panel {
        background-color: #161b22;
        border: 1px solid #30363d;
        border-radius: 8px;
        padding: 16px;
        margin-top: 20px;
    }
    .control-panel-header {
        font-size: 14px;
        font-weight: 600;
        color: #f0f6fc;
        margin-bottom: 12px;
        padding-bottom: 8px;
        border-bottom: 1px solid #21262d;
    }
    .control-row {
        margin-bottom: 10px;
    }
    .control-label {
        font-size: 12px;
        color: #8b949e;
        margin-bottom: 4px;
        display: block;
    }

    /* --- Button styling --- */
    
    /* Sidebar primary button (Delete Selected) */
    [data-testid="stSidebar"] button[kind="primary"] {
        background-color: #da3633;
        border-color: #da3633;
        color: white;
    }
    [data-testid="stSidebar"] button[kind="primary"]:hover {
        background-color: #b62324;
        border-color: #b62324;
    }

    /* Main column primary button (Start Research) */
    div[data-testid="column"]:nth-of-type(1) button[kind="primary"] {
        background-color: #2ea043;
        border-color: #2ea043;
        color: white;
    }
    div[data-testid="column"]:nth-of-type(1) button[kind="primary"]:hover {
        background-color: #2c974b;
        border-color: #2c974b;
    }

    /* Main column secondary button (Stop Agent) */
    div[data-testid="column"]:nth-of-type(1) button[kind="secondary"] {
        background-color: #da3633;
        border-color: #da3633;
        color: white;
    }
    div[data-testid="column"]:nth-of-type(1) button[kind="secondary"]:hover {
        background-color: #b62324;
        border-color: #b62324;
        color: white;
    }
    
    /* --- Markdown content styling --- */
    .agent-log-content table {
        width: 100%;
        border-collapse: collapse;
        margin: 10px 0;
        font-size: 13px;
        color: #c9d1d9;
    }
    .agent-log-content th {
        background-color: #21262d;
        color: #f0f6fc;
        font-weight: 600;
        text-align: left;
        padding: 8px 12px;
        border: 1px solid #30363d;
    }
    .agent-log-content td {
        padding: 8px 12px;
        border: 1px solid #30363d;
    }
    .agent-log-content tr:nth-child(even) {
        background-color: #0d1117;
    }
    .agent-log-content strong {
        color: #f0f6fc;
    }
    .agent-log-content ul, .agent-log-content ol {
        padding-left: 20px;
        margin: 8px 0;
    }
    .agent-log-content li {
        margin-bottom: 4px;
    }
    </style>
""", unsafe_allow_html=True)

def _format_help_label(key: str) -> str:
    label = key.replace("_", " ").strip()
    for prefix in ("writer ",):
        if label.startswith(prefix):
            label = label[len(prefix):].strip()
            break
    acronyms = {
        "ai": "AI",
        "api": "API",
        "llm": "LLM",
        "pdf": "PDF",
        "ris": "RIS",
    }
    words = []
    for word in label.split():
        words.append(acronyms.get(word.lower(), word.capitalize()))
    return " ".join(words) if words else "Help"


def _render_section_help(key: str, description: str):
    """Render a question-mark popover with a short section explanation."""
    label = _format_help_label(key)
    with st.popover(f"? {label}", help="About this section"):
        st.markdown(description)

# =============================================================================
# LLM PROVIDERS, CONNECTIONS, AND STREAMING HELPERS
# =============================================================================
OLLAMA_URL = "http://localhost:11434"
GLM_BASE_URL = "https://api.z.ai/api/paas/v4"
GLM_ALT_BASE_URL = "https://open.bigmodel.cn/api/paas/v4"
PRIMARY_MODEL = "gpt-oss:120b-cloud"
WRAP_NOTE_MODEL = "gpt-oss:20b-cloud"
OLLAMA_TIMEOUT = 120
OPENALEX_MAILTO = None  # Optional contact email for OpenAlex politeness.
MAX_WEB_CONTEXT_CHARS = 20000
WIKI_CHUNK_TOKEN_LIMIT = 40000
WIKI_CHUNK_CHAR_LIMIT = 160000
ROOT_DIR = Path(__file__).resolve().parent.parent
APP_ROOT = Path(__file__).resolve().parent
RESEARCH_ASSISTANT_DATA_ROOT = Path(
    os.environ.get("RESEARCH_ASSISTANT_DATA_ROOT", str(APP_ROOT / "research_assistant_data"))
).expanduser()
API_CONFIG_FILE = RESEARCH_ASSISTANT_DATA_ROOT / "api_config.json"
LEGACY_API_CONFIG_FILE = APP_ROOT / "api_config.txt"
PROJECTS_DIR = RESEARCH_ASSISTANT_DATA_ROOT / "projects"
PROJECT_STATE_FILE = RESEARCH_ASSISTANT_DATA_ROOT / "project_state.json"
PROJECT_STATE_TEMPLATE = {
    "projects": {},
    "doi_registry": {},
    "active_session": {"project": None},
}
WRITER_SESSIONS_KEY = "writer_sessions"
LEGACY_SESSION_DIR = Path(__file__).resolve().parent / "session_tracking"
LEGACY_EXPORTS_DIR = Path(__file__).resolve().parent / "exports"
LEGACY_PROJECTS_FILE = LEGACY_SESSION_DIR / "projects.json"
WRAP_NOTE_SYSTEM_MESSAGE = (
    "You are a careful writing assistant. Organize the user's note into a clear, ordered structure.Keep every point they make intact and use them directly in your final output.\n"
    "Preserve all factual content, names, and citations. Do not add new information.\n"
    "Return only the rewritten note text, with concise headings or numbered bullets."
)
WRAP_NOTE_TEMPERATURE = 0.2
WRAP_NOTE_REASONING = "medium"

def _normalize_base_url(base_url):
    return (base_url or OLLAMA_URL).rstrip("/")

def _normalize_provider(provider):
    if not provider:
        return "ollama"
    value = str(provider).strip().lower()
    if "glm" in value or "z.ai" in value or "bigmodel" in value:
        return "glm"
    if "ollama" in value:
        return "ollama"
    return "ollama"

def _ollama_headers(api_key):
    if not api_key:
        return {}
    return {"Authorization": f"Bearer {api_key}"}

def _glm_headers(api_key):
    if not api_key:
        return {}
    return {"Authorization": f"Bearer {api_key}"}

def _glm_raise_for_status(response, label="GLM request"):
    if response.status_code < 400:
        return
    error_detail = ""
    try:
        payload = response.json()
        error_detail = (
            payload.get("error", {}).get("message")
            or payload.get("message")
            or payload.get("msg")
            or ""
        )
    except Exception:
        error_detail = response.text[:500]
    raise ValueError(f"{label} error: {response.status_code} {error_detail}".strip())

def _glm_extract_text(value):
    """Normalize GLM content fields to text."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                text = item.get("text") or item.get("content") or ""
                if text:
                    parts.append(str(text))
            elif isinstance(item, str):
                parts.append(item)
        return "".join(parts)
    if isinstance(value, dict):
        text = value.get("text") or value.get("content") or ""
        return str(text) if text else ""
    return str(value)

def _glm_extract_content(payload):
    if not isinstance(payload, dict):
        return ""
    choices = payload.get("choices", [])
    if not choices:
        return ""
    choice = choices[0] or {}
    message = choice.get("message", {}) or {}
    content = _glm_extract_text(message.get("content", ""))
    if content:
        return content.strip()
    reasoning = _glm_extract_text(message.get("reasoning_content", ""))
    if reasoning:
        return reasoning.strip()
    text = _glm_extract_text(choice.get("text", ""))
    return text.strip()

def _ollama_generate_stream(
    prompt_text,
    model_name,
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7,
    reasoning_effort="high"
):
    """Yield streamed response chunks from Ollama."""
    base_url = _normalize_base_url(base_url)
    payload = {
        "model": model_name,
        "prompt": prompt_text,
        "stream": True,  # Enable streaming responses.
        "options": {
            "temperature": float(temperature),
            "num_predict": int(max_tokens),
        },
        'Reasoning_effort': reasoning_effort,
    }
    if system_message:
        payload["system"] = system_message

    with requests.post(
        f"{base_url}/api/generate",
        headers=_ollama_headers(api_key),
        json=payload,
        timeout=OLLAMA_TIMEOUT,
        stream=True
    ) as response:
        response.raise_for_status()
        for line in response.iter_lines():
            if line:
                try:
                    chunk = json.loads(line)
                    if "response" in chunk:
                        yield chunk["response"]
                except json.JSONDecodeError:
                    pass

def _glm_generate_stream(
    prompt_text,
    model_name,
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7
):
    """Yield streamed response chunks from GLM (OpenAI-compatible) chat completions."""
    base_url = (base_url or GLM_BASE_URL).rstrip("/")
    headers = {
        "Content-Type": "application/json",
        **_glm_headers(api_key)
    }
    messages = []
    if system_message:
        messages.append({"role": "system", "content": system_message})
    messages.append({"role": "user", "content": prompt_text})
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": float(temperature),
        "max_tokens": int(max_tokens),
        "stream": True
    }

    with requests.post(
        f"{base_url}/chat/completions",
        headers=headers,
        json=payload,
        timeout=OLLAMA_TIMEOUT,
        stream=True
    ) as response:
        _glm_raise_for_status(response, label="GLM stream request")
        content_seen = False
        reasoning_buffer = []
        for raw_line in response.iter_lines(decode_unicode=True):
            if not raw_line:
                continue
            line = raw_line.strip()
            if line.startswith("data:"):
                line = line[len("data:"):].strip()
            if line == "[DONE]":
                break
            try:
                chunk = json.loads(line)
            except json.JSONDecodeError:
                continue
            choices = chunk.get("choices", [])
            if not choices:
                continue
            delta = choices[0].get("delta", {}) or {}
            content = _glm_extract_text(delta.get("content") or delta.get("text") or "")
            if content:
                content_seen = True
                yield content
                continue
            reasoning = _glm_extract_text(delta.get("reasoning_content") or "")
            if reasoning:
                reasoning_buffer.append(reasoning)

            # Some GLM endpoints ignore streaming and return a full message payload.
            message = choices[0].get("message") or {}
            msg_content = _glm_extract_text(message.get("content") or message.get("text") or "")
            if msg_content:
                content_seen = True
                yield msg_content
                break
            msg_reasoning = _glm_extract_text(message.get("reasoning_content") or "")
            if msg_reasoning:
                reasoning_buffer.append(msg_reasoning)

        if not content_seen and reasoning_buffer:
            yield "".join(reasoning_buffer).strip()
        if not content_seen and not reasoning_buffer:
            fallback = _glm_generate(
                prompt_text,
                model_name,
                base_url=base_url,
                api_key=api_key,
                system_message=system_message,
                max_tokens=max_tokens,
                temperature=temperature
            )
            if fallback:
                yield fallback

def _ollama_generate(
    prompt_text,
    model_name,
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7,
    reasoning_effort="high"
):
    """Non-streaming wrapper for backward compatibility (or internal logic that needs full text)."""
    # Consume the stream to assemble a full response when needed.
    full_response = []
    for chunk in _ollama_generate_stream(
        prompt_text,
        model_name,
        base_url,
        api_key,
        system_message,
        max_tokens,
        temperature,
        reasoning_effort
    ):
        full_response.append(chunk)
    return "".join(full_response).strip()

def _glm_generate(
    prompt_text,
    model_name,
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7
):
    """Non-streaming wrapper for GLM chat completions."""
    base_url = (base_url or GLM_BASE_URL).rstrip("/")
    headers = {
        "Content-Type": "application/json",
        **_glm_headers(api_key)
    }
    messages = []
    if system_message:
        messages.append({"role": "system", "content": system_message})
    messages.append({"role": "user", "content": prompt_text})
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": float(temperature),
        "max_tokens": int(max_tokens),
        "stream": False
    }
    response = requests.post(
        f"{base_url}/chat/completions",
        headers=headers,
        json=payload,
        timeout=OLLAMA_TIMEOUT,
    )
    _glm_raise_for_status(response, label="GLM request")
    return _glm_extract_content(response.json())

def _llm_generate_stream(
    prompt_text,
    model_name,
    provider="ollama",
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7,
    reasoning_effort="high"
):
    provider = _normalize_provider(provider)
    if provider == "glm":
        return _glm_generate_stream(
            prompt_text,
            model_name,
            base_url=base_url,
            api_key=api_key,
            system_message=system_message,
            max_tokens=max_tokens,
            temperature=temperature
        )
    return _ollama_generate_stream(
        prompt_text,
        model_name,
        base_url=base_url,
        api_key=api_key,
        system_message=system_message,
        max_tokens=max_tokens,
        temperature=temperature,
        reasoning_effort=reasoning_effort
    )

def _llm_generate(
    prompt_text,
    model_name,
    provider="ollama",
    base_url=None,
    api_key=None,
    system_message=None,
    max_tokens=20000,
    temperature=0.7,
    reasoning_effort="high"
):
    provider = _normalize_provider(provider)
    if provider == "glm":
        return _glm_generate(
            prompt_text,
            model_name,
            base_url=base_url,
            api_key=api_key,
            system_message=system_message,
            max_tokens=max_tokens,
            temperature=temperature
        )
    return _ollama_generate(
        prompt_text,
        model_name,
        base_url=base_url,
        api_key=api_key,
        system_message=system_message,
        max_tokens=max_tokens,
        temperature=temperature,
        reasoning_effort=reasoning_effort
    )

def check_ollama_connection(base_url=None, preferred_model=None, api_key=None, debug=False):
    """Check connection to an Ollama server and return the preferred model name."""
    base_url = _normalize_base_url(base_url)
    try:
        if debug:
            flash("Connecting to Ollama...", level="info")

        response = requests.get(
            f"{base_url}/api/tags",
            headers=_ollama_headers(api_key),
            timeout=5,
        )
        response.raise_for_status()
        data = response.json()
        models = data.get("models", [])
        model_names = []
        for entry in models:
            if isinstance(entry, dict):
                model_names.append(entry.get("name") or entry.get("model") or entry.get("id"))
        model_names = [m for m in model_names if m]

        if debug:
            flash("Successfully connected to Ollama", level="success")
            if model_names:
                flash(f"Available models: {', '.join(model_names)}", level="info")

        pref_model = preferred_model or PRIMARY_MODEL
        if model_names and pref_model not in model_names:
            return True, pref_model, f"Model `{pref_model}` not found on server. Using it anyway."

        return True, pref_model, None
    except requests.RequestException as e:
        error_detail = str(e)
        if "connect" in error_detail.lower() or "refused" in error_detail.lower():
            return False, None, f"Cannot reach Ollama at {base_url}. Check the URL and credentials."
        return False, None, f"Ollama request error: {error_detail}"
    except Exception as e:
        error_detail = str(e)
        return False, None, f"Error: {error_detail}"

def check_glm_connection(base_url=None, preferred_model=None, api_key=None, debug=False):
    """Check connection to GLM (Z.ai/OpenBigModel) and return the preferred model name."""
    if not api_key:
        return False, None, "GLM API key is missing."
    base_url = (base_url or GLM_BASE_URL).rstrip("/")
    try:
        if debug:
            flash("Connecting to GLM...", level="info")
        payload = {
            "model": preferred_model or "glm-4.7",
            "messages": [{"role": "user", "content": "ping"}],
            "max_tokens": 1,
            "temperature": 0.0,
            "stream": False
        }
        response = requests.post(
            f"{base_url}/chat/completions",
            headers={"Content-Type": "application/json", **_glm_headers(api_key)},
            json=payload,
            timeout=10
        )
        if response.status_code == 200:
            if debug:
                flash("Successfully connected to GLM", level="success")
            model_name = preferred_model or "glm-4.7"
            return True, model_name, None
        error_detail = ""
        try:
            error_detail = response.json().get("error", {}).get("message", "")
        except Exception:
            error_detail = response.text[:200]
        return False, None, f"GLM request error: {response.status_code} {error_detail}".strip()
    except requests.RequestException as e:
        return False, None, f"GLM request error: {e}"
    except Exception as e:
        return False, None, f"Error: {e}"

def check_llm_connection(provider="ollama", base_url=None, preferred_model=None, api_key=None, debug=False):
    provider = _normalize_provider(provider)
    if provider == "glm":
        return check_glm_connection(
            base_url=base_url,
            preferred_model=preferred_model,
            api_key=api_key,
            debug=debug
        )
    return check_ollama_connection(
        base_url=base_url,
        preferred_model=preferred_model,
        api_key=api_key,
        debug=debug
    )

# -----------------------------------------------------------------------------
# Prompt templates for LLM-driven steps (planning, search, summarization).
# -----------------------------------------------------------------------------
THESIS_SYSTEM_MESSAGE = """You are a thesis writing assistant that turns a single topic request into a well-structured section plan with a logical flow.

USER INPUT TYPE
The user may provide:
- a topic only (e.g., "I want to write about agent-based modelling"), OR
- a topic plus constraints (audience, word limit, thesis chapter, study context, required concepts, excluded concepts).

CORE TASK
Given the user's topic, produce a structured decomposition into sub-sections (notions), order them for best narrative flow, and provide what to write under each.

STRICT RULES
1) Do not ask the user questions unless essential to proceed. If information is missing, make minimal, explicit assumptions and continue.
2) Do not invent empirical results, datasets, or citations. If citations are needed, insert placeholders like [REF] and list what kind of sources are required.
3) Keep the output directly usable in a thesis: clear headings, coherent progression, and no fluff.
4) Keep terminology consistent. Use ABM/LLM/Hybrid ABM-LLM exactly as written by the user if mentioned.
5) do not use tables unless user asks you to.
6) Do not use emojis in your response. Keep it strictly professional text.

DEFAULT STRUCTURE LOGIC (APPLY UNLESS USER OVERRIDES)
Start broad -> narrow:
- Definition and scope
- Historical evolution and related approaches
- Core components and mechanics
- Typical workflow (design -> calibration -> validation -> analysis)
- Strengths and limitations
- Relevance to the thesis problem and how it connects to the user's studies

OUTPUT FORMAT (ALWAYS)
Return four blocks in this exact order:

1) SECTION MAP (ORDERED)
A numbered list of sub-sections with:
- Heading title
- 1-line purpose
- Key transitions (why this comes next)

2) CONTENT BULLETS (PER SUB-SECTION)
For each heading:
- 4-10 bullets of what to say
- Definitions, key concepts, and must-include points
- Optional examples (clearly labeled as examples)

3) MINI-DRAFT (OPTIONAL BUT DEFAULT: ON)
Write 1 short paragraph per sub-section (3-6 sentences each), in academic thesis tone.
If the user asked for outline-only, skip this block.

4) OPEN ITEMS
- Assumptions made
- Missing details to confirm later
- Where citations are needed (use [REF] placeholders)

STYLE
- Academic, direct, precise.
- No motivational language, no filler.
- Prefer concrete nouns/verbs over vague phrasing.
- Keep paragraphs tight and logically connected.

WHEN USER PROVIDES ONLY A TOPIC
Assume:
- Target audience: thesis examiners / academic readers
- Depth: medium (overview + enough detail to support later methods/results)
- Voice: neutral academic
- Length: ~1-2 pages worth of structure (not a full chapter)

WEB CONTEXT (IF PROVIDED)
If Wikipedia context is provided below, treat it as up-to-date background knowledge.
Incorporate this knowledge into your planning where relevant - use it to inform definitions, recent developments, and key concepts.
Only use credible-looking content; ignore any irrelevant material.
"""

WIKI_MAPPER_PROMPT = """
You are a Wikipedia Search Optimizer.
The user wants to research a complex topic, but their query might be vague, slang, or a specific research question.

USER QUERY: "{user_topic}"

TASK:
Identify the **Core Encyclopedic Concepts** necessary to understand this topic.
Convert the user's query into 1-3 specific, standard Wikipedia page titles.

RULES:
1. Avoid "how-to" phrases (e.g. convert "how to build agents" -> "Agent-based model").
2. Resolve acronyms (e.g. convert "ABM" -> "Agent-based model").
3. Split hybrid topics (e.g. convert "ABM-LLM" -> ["Agent-based model", "Large language model"]).

OUTPUT JSON ONLY:
{{
    "wiki_pages": ["Page Title 1", "Page Title 2"]
}}
"""

# -----------------------------------------------------------------------------
# Wikipedia scouting + summarization pipeline.
# -----------------------------------------------------------------------------
def generate_wiki_pages(topic, model_name, provider="ollama", base_url=None, api_key=None):
    """Generate specific Wikipedia page titles from the topic using the LLM."""
    prompt = WIKI_MAPPER_PROMPT.format(user_topic=topic)
    raw = call_prompt_chain(
        prompt,
        "Wiki Page Mapper",
        model_name,
        provider=provider,
        base_url=base_url,
        api_key=api_key,
        show_debug=False,
        stream_output=False
    )
    try:
        # Use the robust JSON parser helper.
        data = parse_json_response(raw, "Wiki Page Mapper")
        return data.get("wiki_pages", [topic])
    except:
        return [topic]

def perform_deep_search(questions, max_results_per_query=2):
    """
    Searches Wikipedia for the topic and extracts content.
    """
    aggregated_context = ""
    seen_urls = set()

    for q in questions:
        # Optional debug logging for interactive sessions.
        try:
             # 1. Search for the best matching title.
            search_results = wikipedia.search(q)
            
            if not search_results:
                continue
            
            # Use the top results to limit noise.
            for title in search_results[:max_results_per_query]:
                try:
                    # 2. Fetch full page content.
                    # `auto_suggest=False` avoids incorrect auto-disambiguation.
                    page = wikipedia.page(title, auto_suggest=False)
                    
                    if page.url in seen_urls:
                        continue
                    seen_urls.add(page.url)

                    # 3. Format output to match the UI parsing regex.
                    # Format: === SOURCE: {title} === ... URL: {url} ... FULL CONTENT EXTRACT: ...
                    aggregated_context += f"=== SOURCE: {page.title} ===\n"
                    aggregated_context += f"Relevance: Found via query '{q}'\n"
                    aggregated_context += f"URL: {page.url}\n"
                    # Combine summary with full content for richer context.
                    full_text = f"{page.summary}\n\n{page.content}"
                    aggregated_context += f"FULL CONTENT EXTRACT:\n{full_text}\n"
                    aggregated_context += "-" * 40 + "\n\n"
                    
                except wikipedia.DisambiguationError as e:
                    # Fallback: select the first disambiguation option.
                    try:
                        first_option = e.options[0]
                        page = wikipedia.page(first_option, auto_suggest=False)
                        if page.url in seen_urls: continue
                        seen_urls.add(page.url)
                        
                        aggregated_context += f"=== SOURCE: {page.title} (Disambiguated) ===\n"
                        aggregated_context += f"Relevance: Found via query '{q}' (Ambiguous topic)\n"
                        aggregated_context += f"URL: {page.url}\n"
                         # Combine summary with full content for richer context.
                        full_text = f"{page.summary}\n\n{page.content}"
                        aggregated_context += f"FULL CONTENT EXTRACT:\n{full_text}\n"
                        aggregated_context += "-" * 40 + "\n\n"
                    except:
                        continue
                except Exception:
                    continue

        except Exception:
            pass

    if not aggregated_context:
        return "Scout failed to find relevant Wikipedia pages. Try different keywords."
        
    return aggregated_context

def _split_web_context_sources(context):
    if not context:
        return []
    text = str(context)
    if "=== SOURCE:" not in text:
        return [text]
    parts = re.split(r'(?==== SOURCE:)', text)
    return [part.strip() for part in parts if part.strip()]

def _chunk_text_by_chars(text, max_chars):
    text = str(text or "")
    if not text:
        return []
    paragraphs = re.split(r"\n\s*\n", text)
    chunks = []
    current = []
    current_len = 0
    for paragraph in paragraphs:
        para = paragraph.strip()
        if not para:
            continue
        para_len = len(para)
        if para_len > max_chars:
            if current:
                chunks.append("\n\n".join(current))
                current = []
                current_len = 0
            start = 0
            while start < para_len:
                end = start + max_chars
                chunks.append(para[start:end])
                start = end
            continue
        extra_len = para_len + (2 if current else 0)
        if current_len + extra_len > max_chars:
            if current:
                chunks.append("\n\n".join(current))
            current = [para]
            current_len = para_len
        else:
            current.append(para)
            current_len += extra_len
    if current:
        chunks.append("\n\n".join(current))
    return chunks or [text[:max_chars]]

def _parse_wiki_source_block(block):
    text = str(block or "")
    title_match = re.search(r'=== SOURCE: (.*?) ===', text)
    url_match = re.search(r'URL: (.*?)\n', text)
    title = title_match.group(1).strip() if title_match else "Unknown Source"
    url = url_match.group(1).strip() if url_match else ""
    content = ""
    if "FULL CONTENT EXTRACT:" in text:
        content = text.split("FULL CONTENT EXTRACT:", 1)[1].strip()
    else:
        content = text.strip()
    return {"title": title, "url": url, "content": content}

def _truncate_web_context(context, max_chars=MAX_WEB_CONTEXT_CHARS):
    text = str(context)
    if len(text) <= max_chars:
        return text
    blocks = _split_web_context_sources(text)
    if len(blocks) <= 1:
        return text[:max_chars] + "\n[TRUNCATED]"
    per_block = max(400, max_chars // len(blocks))
    trimmed_blocks = []
    for block in blocks:
        if len(block) > per_block:
            trimmed_blocks.append(block[:per_block] + "\n[TRUNCATED SOURCE]")
        else:
            trimmed_blocks.append(block)
    combined = "\n\n".join(trimmed_blocks)
    if len(combined) > max_chars:
        combined = combined[:max_chars] + "\n[TRUNCATED]"
    return combined

def summarize_web_context(context, model_name, provider="ollama", base_url=None, api_key=None, progress_callback=None):
    """Summarize the raw web context."""
    if not context or "Scout failed" in context or len(context) < 100:
        return ""

    def _legacy_single_pass_summary(raw_context):
        safe_context = _truncate_web_context(raw_context, MAX_WEB_CONTEXT_CHARS)
        prompt = SUMMARIZER_PROMPT.format(context=safe_context)
        return call_prompt_chain(
            prompt,
            "Research Summary",
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            show_debug=False,
            stream_output=True
        )

    def _normalize_key(text):
        return re.sub(r"\s+", " ", str(text or "").strip().lower())

    source_blocks = _split_web_context_sources(context)
    pages_to_process = []
    for block in source_blocks:
        meta = _parse_wiki_source_block(block)
        page_content = meta.get("content") or ""
        if len(page_content) < 100:
            continue
        pages_to_process.append(meta)

    for idx, meta in enumerate(pages_to_process, 1):
        meta["page_id"] = f"P{idx}"

    total_pages = len(pages_to_process)
    if progress_callback:
        progress_callback(0, total_pages)

    page_summaries = []

    for page_index, meta in enumerate(pages_to_process, 1):
        page_title = meta.get("title") or "Unknown Source"
        page_url = meta.get("url") or ""
        page_content = meta.get("content") or ""
        page_id = meta.get("page_id") or f"P{page_index}"

        chunks = _chunk_text_by_chars(page_content, WIKI_CHUNK_CHAR_LIMIT)
        if not chunks:
            if progress_callback:
                progress_callback(page_index, total_pages)
            continue

        term_map = {}
        sota_points = []
        methods = []
        applications = []
        limitations = []
        sota_seen = set()
        methods_seen = set()
        applications_seen = set()
        limitations_seen = set()

        def _merge_list(items, target, seen):
            for item in items:
                key = _normalize_key(item)
                if not key or key in seen:
                    continue
                target.append(str(item).strip())
                seen.add(key)

        for idx, chunk in enumerate(chunks, 1):
            prompt = WIKI_CHUNK_EXTRACT_PROMPT.format(
                title=page_title,
                url=page_url,
                chunk_index=idx,
                chunk_total=len(chunks),
                content=chunk
            )
            try:
                raw = call_prompt_chain(
                    prompt,
                    f"Wikipedia Chunk Extract ({page_title})",
                    model_name,
                    provider=provider,
                    base_url=base_url,
                    api_key=api_key,
                    show_debug=False,
                    response_language="json",
                    stream_output=False,
                    show_response=False
                )
                parsed = parse_json_response(raw, "Wiki Chunk Extract")
            except Exception:
                continue

            if isinstance(parsed, dict):
                for entry in parsed.get("terminology", []) or []:
                    if not isinstance(entry, dict):
                        continue
                    term = entry.get("term")
                    definition = entry.get("definition")
                    term_key = _normalize_key(term)
                    if term_key and definition:
                        term_map[term_key] = {
                            "term": str(term).strip(),
                            "definition": str(definition).strip()
                        }
                _merge_list(parsed.get("sota_points", []) or [], sota_points, sota_seen)
                _merge_list(parsed.get("methods_or_approaches", []) or [], methods, methods_seen)
                _merge_list(parsed.get("applications", []) or [], applications, applications_seen)
                _merge_list(parsed.get("limitations_or_gaps", []) or [], limitations, limitations_seen)

        has_structured = bool(term_map or sota_points or methods or applications or limitations)
        fallback_summary = None
        if not has_structured:
            try:
                fallback_prompt = WIKI_PAGE_FALLBACK_PROMPT.format(
                    title=page_title,
                    url=page_url,
                    content=page_content[:WIKI_CHUNK_CHAR_LIMIT]
                )
                raw_fallback = call_prompt_chain(
                    fallback_prompt,
                    f"Wikipedia Page Fallback ({page_title})",
                    model_name,
                    provider=provider,
                    base_url=base_url,
                    api_key=api_key,
                    show_debug=False,
                    response_language="json",
                    stream_output=False,
                    show_response=False
                )
                parsed_fallback = parse_json_response(raw_fallback, "Wiki Page Fallback")
            except Exception:
                parsed_fallback = None
            if isinstance(parsed_fallback, dict):
                for entry in parsed_fallback.get("terminology", []) or []:
                    if not isinstance(entry, dict):
                        continue
                    term = entry.get("term")
                    definition = entry.get("definition")
                    term_key = _normalize_key(term)
                    if term_key and definition:
                        term_map[term_key] = {
                            "term": str(term).strip(),
                            "definition": str(definition).strip()
                        }
                _merge_list(parsed_fallback.get("sota_points", []) or [], sota_points, sota_seen)
                _merge_list(parsed_fallback.get("methods_or_approaches", []) or [], methods, methods_seen)
                _merge_list(parsed_fallback.get("applications", []) or [], applications, applications_seen)
                _merge_list(parsed_fallback.get("limitations_or_gaps", []) or [], limitations, limitations_seen)
                fallback_summary = True

        page_summaries.append(
            {
                "page_id": page_id,
                "title": page_title,
                "url": page_url,
                "terminology": list(term_map.values()),
                "sota_points": sota_points,
                "methods_or_approaches": methods,
                "applications": applications,
                "limitations_or_gaps": limitations,
                "fallback_summary": bool(fallback_summary),
            }
        )
        if progress_callback:
            progress_callback(page_index, total_pages)

    if not page_summaries:
        return _legacy_single_pass_summary(context)

    final_payload = {"pages": page_summaries}
    final_prompt = WIKI_FINAL_SUMMARY_PROMPT.format(
        input_json=json.dumps(final_payload, ensure_ascii=False, indent=2)
    )
    final_summary = call_prompt_chain(
        final_prompt,
        "Research Summary",
        model_name,
        provider=provider,
        base_url=base_url,
        api_key=api_key,
        show_debug=False,
        response_language="text",
        stream_output=True
    )
    if not final_summary or not final_summary.strip():
        legacy = _legacy_single_pass_summary(context)
        if legacy:
            legacy += "\n\nNote: References could not be generated for this summary."
        return legacy
    return final_summary

QUERY_BUILDER_PROMPT = """You are a search query generator for academic APIs (Semantic Scholar).

TASK:
Read the Section Plan below and generate search queries for each section heading you identify.

RULES:
1. Identify all section headings from the plan.
2. For EACH section heading, generate exactly 4 search queries.
3. Keep queries concise (under 12 words).
4. Use simple keywords, no complex Boolean operators.

OUTPUT FORMAT (JSON ONLY, no extra text):
{{
  "section_queries": {{
    "<actual section heading 1>": ["query1", "query2", "query3", "query4"],
    "<actual section heading 2>": ["query1", "query2", "query3", "query4"]
  }}
}}

Section Plan:
{plan_text}

Output:
"""

RELEVANCE_PROMPT = """You are a relevance judge. Decide if each paper is relevant to the topic.

Return JSON only in this format:
{{ "results": [{{ "id": 1, "relevant": true, "score": 75 }}] }}

Rules:
- "relevant" must be true or false.
- "score" must be a number from 0 to 100 (higher is more relevant).
- Return exactly one result per input paper.

Input JSON:
{input_json}
Output:
"""

WIKI_CHUNK_EXTRACT_PROMPT = """You are a research extraction assistant.

TASK:
Extract structured research notes from a single Wikipedia content chunk.
Use ONLY the provided chunk content.

PAGE TITLE: {title}
PAGE URL: {url}
CHUNK: {chunk_index}/{chunk_total}

RULES:
1) Do not invent facts or terms not in the chunk.
2) Avoid duplicates within this chunk.
3) Keep definitions concise and precise.
4) Output JSON only in the exact schema below.

OUTPUT JSON:
{{
  "terminology": [{{"term": "...", "definition": "..."}}],
  "sota_points": ["..."],
  "methods_or_approaches": ["..."],
  "applications": ["..."],
  "limitations_or_gaps": ["..."]
}}

CHUNK CONTENT:
{content}
"""

WIKI_PAGE_FALLBACK_PROMPT = """You are a research extraction assistant.

TASK:
Summarize a single Wikipedia page into structured notes.
Use ONLY the provided page content.

PAGE TITLE: {title}
PAGE URL: {url}

RULES:
1) Do not invent facts or terms not in the page content.
2) Keep definitions concise and precise.
3) Output JSON only in the exact schema below.

OUTPUT JSON:
{{
  "terminology": [{{"term": "...", "definition": "..."}}],
  "sota_points": ["..."],
  "methods_or_approaches": ["..."],
  "applications": ["..."],
  "limitations_or_gaps": ["..."]
}}

PAGE CONTENT:
{content}
"""

WIKI_FINAL_SUMMARY_PROMPT = """You are a research synthesis assistant.

TASK:
Using the merged notes from multiple Wikipedia pages, produce the final summary.
Use ONLY the provided inputs. Do not add external knowledge.

REQUIRED OUTPUT FORMAT (Markdown only):
# Terminology Explanation
- Term: definition [P#]
- ...

# State of the Art
| Area/Theme | Summary | Evidence (Page) |
|---|---|---|
| ... | ... [P#] | P# |

- Additional key points (methods, applications, limitations) with [P#] on every bullet

# References
[P#] Page Title — URL

INPUT (JSON):
{input_json}
"""

SUMMARIZER_PROMPT = """You are a Research Assistant.
TASK: Summarize the following raw search results into a coherent validation of the user's topic.
FOCUS: Key definitions, state of the art, conflicting views, and practical applications.
OUTPUT: A concise markdown summary (approx 200-300 words).

RAW CONTENT:
{context}
"""

ABSTRACT_BATCH_ANALYSIS_PROMPT = """You are a research synthesis assistant.

TASK:
Analyze this batch of relevant papers for the user topic using ONLY the paper abstracts provided.

RULES:
1. Do not use external knowledge or invented facts.
2. If an abstract is unclear or limited, state that directly.
3. Base every point on the provided abstracts only.
4. Keep it concise and practical for writing support.
5. When citing a paper in your output, use its provided reference token exactly, like [P1].
6. Do not invent reference tokens.

USER TOPIC:
{topic}

PAPERS (JSON):
{input_json}

OUTPUT (markdown only):
### Batch Findings
- Key themes from this batch
- Important methods, claims, or trends
- Conflicts/limitations/gaps in this batch

### Batch Recommended Papers
- Up to 3 papers from this batch
- For each: include [P#] + title + one-line reason tied to its abstract
"""

ABSTRACT_FINAL_REPORT_PROMPT = """You are a research writing assistant.

TASK:
Create one final report for the user using ONLY:
1) The batch analyses below (which were derived only from abstracts), and
2) The provided paper catalog.

STRICT RULES:
1. Do not add facts not present in the inputs.
2. Clearly identify uncertainty when evidence is weak.
3. Keep recommendations tied to abstract evidence only.
4. Return markdown only.
5. Whenever you refer to a paper, cite it using [P#] from the provided catalog.
6. Do not invent or rename reference tokens.
7. Use paper citations throughout all report sections, not only in the final suggestions section.

USER TOPIC:
{topic}

BATCH ANALYSES:
{batch_analyses}

PAPER CATALOG (JSON):
{paper_catalog}

OUTPUT STRUCTURE:
# Abstract-Based Report
## Executive Summary
## Key Themes from Relevant Papers
## Areas of Agreement and Disagreement
## Evidence Gaps and Limitations
## Suggested Papers to Read Next
- 5-10 papers when available
- each bullet: title + short rationale based on abstract evidence
"""

THINK_BLOCK_PATTERN = re.compile(r'<think>(.*?)</think>', re.DOTALL | re.IGNORECASE)
AGENT_LOG_WRAPPER_RE = re.compile(
    r'^\s*<div[^>]*class=["\']agent-log-section["\'][^>]*>(?P<body>.*)</div>\s*$',
    re.DOTALL | re.IGNORECASE,
)
AGENT_LOG_HTML_RE = re.compile(
    r'</?(?:p|div|h[1-6]|ul|ol|li|table|thead|tbody|tr|th|td|blockquote|pre|code|br)\b',
    re.IGNORECASE,
)

def _normalize_agent_log_source(value: str) -> str:
    """Normalize model output before rendering it inside agent log cards."""
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    # Normalize common non-breaking characters that cause odd wrapping.
    text = (
        text.replace("\u00A0", " ")
        .replace("\u202F", " ")
        .replace("\u2011", "-")
    )
    match = AGENT_LOG_WRAPPER_RE.match(text)
    if match:
        text = (match.group("body") or "").strip()
    return text

def _render_agent_log_markdown_html(value: str) -> str:
    """Render markdown/plain-text safely for card HTML containers."""
    text = _normalize_agent_log_source(value)
    if not text:
        return ""
    # Check if the content is primarily HTML (starts with HTML structure)
    # Only skip markdown processing if text appears to be complete HTML output
    stripped = text.strip()
    is_primarily_html = (
        stripped.startswith(('<div', '<p', '<ul', '<ol', '<table', '<h1', '<h2', '<h3', '<h4', '<h5', '<h6'))
        and stripped.endswith(('</div>', '</p>', '</ul>', '</ol>', '</table>', '</h1>', '</h2>', '</h3>', '</h4>', '</h5>', '</h6>'))
    )
    if is_primarily_html:
        return text
    # Process markdown for all other content (including mixed markdown/HTML)
    # The 'extra' extension handles inline HTML gracefully
    return markdown.markdown(text, extensions=["extra", "nl2br", "sane_lists"])

# -----------------------------------------------------------------------------
# Agent timeline metadata (used for step status + progress rendering).
# -----------------------------------------------------------------------------
AGENT_TIMELINE_STEPS = [
    ("step1", "STEP 1", "Web research"),
    ("step1_5", "STEP 1.5", "Summarizing Wikipedia"),
    ("step2", "STEP 2", "Drafting section plan"),
    ("step3", "STEP 3", "Generating academic search queries"),
    ("search_and_analyze", "STEP 4", "Searching academic databases"),
]
AGENT_TIMELINE_MAP = {
    step_id: {"step_label": step_label, "label": label}
    for step_id, step_label, label in AGENT_TIMELINE_STEPS
}

def _set_agent_timeline_entry(step_id, status_text, detail, progress=None):
    timeline = st.session_state.setdefault("agent_timeline", {})
    meta = AGENT_TIMELINE_MAP.get(step_id, {"step_label": step_id, "label": ""})
    entry = timeline.get(step_id, {})
    entry["step_label"] = meta["step_label"]
    entry["label"] = meta["label"]
    entry["status_text"] = status_text
    entry["detail"] = detail
    entry["progress"] = progress
    timeline[step_id] = entry
    st.session_state["agent_timeline"] = timeline

def _render_agent_timeline(placeholder):
    timeline = st.session_state.get("agent_timeline", {})
    with placeholder.container():
        for step_id, step_label, _ in AGENT_TIMELINE_STEPS:
            entry = timeline.get(step_id)
            if not entry:
                continue
            status_text = str(entry.get("status_text") or "").strip()
            detail = str(entry.get("detail") or "").strip()
            if status_text and detail:
                line = f"{step_label} {status_text}: {detail}"
            elif detail:
                line = f"{step_label} {detail}"
            else:
                line = f"{step_label}"
            st.markdown(line)
            progress = entry.get("progress")
            if status_text.lower() == "running" and progress:
                if isinstance(progress, dict) and any(k in progress for k in ("primary", "analysis", "report")):
                    for key in ("primary", "analysis", "report"):
                        sub = progress.get(key)
                        if not isinstance(sub, dict):
                            continue
                        try:
                            total = int(sub.get("total") or 0)
                            current = int(sub.get("current") or 0)
                        except (TypeError, ValueError):
                            total = 0
                            current = 0
                        if total > 0:
                            ratio = min(1.0, max(0.0, current / total))
                            st.progress(ratio)
                            text = sub.get("text") or f"{current} / {total}"
                            st.markdown(text)
                else:
                    try:
                        total = int(progress.get("total") or 0)
                        current = int(progress.get("current") or 0)
                    except (TypeError, ValueError):
                        total = 0
                        current = 0
                    if total > 0:
                        ratio = min(1.0, max(0.0, current / total))
                        st.progress(ratio)
                        text = progress.get("text") or f"{current} / {total} pages done"
                        st.markdown(text)

def split_reasoning_and_answer(raw_text):
    # Split model output into <think> reasoning and final answer segments.
    if not raw_text:
        return "", ""
    # We look for <think> blocks (DOTALL required for multiline captures).
    reasoning_parts = []
    final_answer_parts = []

    # Simple state machine to remain robust with streamed/incomplete chunks.
    # Logic: inside <think>...</think> is reasoning; everything else is answer.
    # If <think> is unclosed, treat the remainder as reasoning.
    parts = re.split(r'(<think>|<\/think>)', raw_text, flags=re.IGNORECASE)
    
    in_think = False
    for part in parts:
        if part.lower() == "<think>":
            in_think = True
            continue
        elif part.lower() == "</think>":
            in_think = False
            continue
            
        if in_think:
            reasoning_parts.append(part)
        else:
            final_answer_parts.append(part)
            
    return "".join(reasoning_parts).strip(), "".join(final_answer_parts).strip()

def call_prompt_chain(
    prompt_text,
    label,
    model_name,
    provider="ollama",
    base_url=None,
    api_key=None,
    system_message=None,
    show_debug=False,
    response_language="json",
    stream_output=False,  # Enable streamed UI updates when True.
    status_container=None,  # Optional status container for progress updates.
    show_response=True  # Show the LLM response in a collapsible expander.
):
    """Send a prompt to the model using Ollama, optionally streaming updates."""
    
    if not stream_output:
        # Non-streaming path.
        content = _llm_generate(
            prompt_text,
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            system_message=system_message,
            max_tokens=20000,
            temperature=0.7,
        )
        reasoning, final_answer = split_reasoning_and_answer(content)
        answer_to_display = final_answer or content
        
        if show_debug:
            with st.expander(f"Debug: {label}", expanded=False):
                if reasoning:
                    st.markdown("**Reasoning**")
                    st.code(reasoning, language="text")
                st.markdown("**Response**")
                st.code(answer_to_display, language=response_language)
        return answer_to_display

    # Streaming mode: show a simple spinner and accumulate text chunks.
    full_text = ""
    
    with st.spinner(f"Agent is working on {label}..."):
        for chunk in _llm_generate_stream(
            prompt_text,
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            system_message=system_message,
            max_tokens=20000,
            temperature=0.7,
        ):
            full_text += chunk
    
    # Finalize output.
    reasoning, final_answer = split_reasoning_and_answer(full_text)
    answer_to_display = final_answer or full_text
    
    # Show the LLM response in a clean expander with scrollbar
    if show_response and answer_to_display:
        rendered_answer = _render_agent_log_markdown_html(answer_to_display)
        with st.expander(f"LLM Response: {label}", expanded=False):
            st.markdown(
                f'''<div class="agent-log-section" style="font-size: 12px; max-height: 400px; overflow-y: auto;">{rendered_answer}</div>''',
                unsafe_allow_html=True
            )
    
    if show_debug and reasoning:
        with st.expander(f"Debug: {label} Reasoning", expanded=False):
            st.code(reasoning, language="text")

    return answer_to_display

def parse_json_response(raw_text, label):
    """Normalize code fences and extract the JSON payload."""
    if not raw_text:
        raise ValueError(f"{label} response was empty.")

    _, cleaned_answer = split_reasoning_and_answer(raw_text)
    cleaned = (cleaned_answer or raw_text).strip()

    # Remove markdown code fences
    cleaned = re.sub(r'^```(?:json)?', '', cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r'```$', '', cleaned).strip()

    # Find and parse JSON
    for idx, ch in enumerate(cleaned):
        if ch in "{[":
            try:
                obj, _ = json.JSONDecoder().raw_decode(cleaned[idx:])
                return obj
            except ValueError:
                continue

    raise ValueError(f"{label} did not return valid JSON.")

# -----------------------------------------------------------------------------
# Legacy one-shot prompt chain (kept for backward compatibility).
# -----------------------------------------------------------------------------
def _execute_prompt_chain(topic, model_name, provider="ollama", base_url=None, api_key=None, show_debug=False, status_callback=None):
    """Generate a section plan, then derive search queries from it."""
    try:
        # Step 0: Map topic to Wikipedia pages.
        if status_callback: status_callback("Identifying relevant Wikipedia pages...")
        questions = generate_wiki_pages(topic, model_name, provider, base_url, api_key)
        
        # Step 1: Wikipedia retrieval.
        if status_callback: status_callback("Searching Wikipedia...")
        web_context = perform_deep_search(questions, max_results_per_query=1)
        
        # Step 1.5: Summarize retrieved context.
        web_summary = ""
        if web_context and "Scout failed" not in web_context:
            if status_callback: status_callback("Summarizing Search Results...")
            web_summary = summarize_web_context(web_context, model_name, provider, base_url, api_key)

        # Step 2: Create planner prompt with web context.
        if status_callback: status_callback("Drafting Research Plan...")
        planner_input = str(topic)
        
        # Prefer summary when available; fall back to raw context.
        if web_summary:
            planner_input += f"\n\n--- WIKIPEDIA KNOWLEDGE (USE IN PLANNING) ---\n{web_summary}"
        elif web_context:
            safe_context = _truncate_web_context(web_context, MAX_WEB_CONTEXT_CHARS)
            planner_input += f"\n\n--- WIKIPEDIA CONTEXT ---\n{safe_context}"
        
        section_plan = call_prompt_chain(
            planner_input,
            "Section Plan",
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            system_message=THESIS_SYSTEM_MESSAGE,
            show_debug=show_debug,
            response_language="text",
            stream_output=True,  # Enable streamed UI response.
        )
        
        # Step 3: Generate search queries.
        if status_callback: status_callback("Generating Search Queries...")
        # Pass the full plan to the query builder and let the LLM detect headings.
        query_prompt = QUERY_BUILDER_PROMPT.format(plan_text=section_plan)
        query_raw = call_prompt_chain(
            query_prompt,
            "Query Builder",
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            show_debug=show_debug,
            stream_output=True
        )
        query_bundle = parse_json_response(query_raw, "Query Builder")
        if not isinstance(query_bundle, dict):
            raise ValueError("Query builder response must be a JSON object.")

        # The LLM is expected to return a dict with section_queries.
        # Extract and trust that structure.
        section_queries = (
            query_bundle.get("section_queries")
            or query_bundle.get("queries")
            or {}
        )
        
        if not isinstance(section_queries, dict) or not section_queries:
            raise ValueError("Query builder response missing or invalid 'section_queries'.")

        return section_plan, section_queries, questions, web_context, web_summary
    except json.JSONDecodeError as exc:
        flash(f"JSON parsing failed: {exc}", level="error")
        raise
    except Exception as exc:
        flash(f"Error in prompt chain: {exc}", level="error")
        raise

def run_prompt_chain(topic, model_name=None, provider="ollama", base_url=None, api_key=None, show_debug=False, max_attempts=3, status_callback=None):
    """Run the prompt chain with retries for robustness."""
    attempts = max(1, int(max_attempts))
    model_to_use = model_name or PRIMARY_MODEL
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return _execute_prompt_chain(
                topic,
                model_to_use,
                provider=provider,
                base_url=base_url,
                api_key=api_key,
                show_debug=show_debug,
                status_callback=status_callback
            )
        except Exception as exc:
            last_error = exc
            if attempt < attempts:
                flash(f"Attempt {attempt}/{attempts} failed. Retrying...", level="warning")
            else:
                break
    raise last_error

def score_relevance(topic, papers, model_name, provider="ollama", base_url=None, api_key=None, show_debug=False, batch_size=5, progress_callback=None):
    """Score paper relevance to the topic using the LLM."""
    if not papers:
        return papers

    indexed = list(enumerate(papers))
    results = {}
    total = len(indexed)
    processed = 0

    for start in range(0, len(indexed), batch_size):
        batch = indexed[start:start + batch_size]
        payload = {
            "topic": topic,
            "papers": [
                {
                    "id": idx,
                    "title": p.get("title", ""),
                    "abstract": p.get("abstract", ""),
                    "tldr": p.get("tldr", ""),
                }
                for idx, p in batch
            ],
        }
        prompt = RELEVANCE_PROMPT.format(input_json=json.dumps(payload, ensure_ascii=False))
        raw = call_prompt_chain(
            prompt,
            "Prompt: Relevance",
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            show_debug=show_debug,
        )
        parsed = parse_json_response(raw, "Relevance")
        if not isinstance(parsed, dict):
            raise ValueError("Relevance response must be a JSON object.")
        entries = parsed.get("results", [])
        if not isinstance(entries, list):
            raise ValueError("Relevance response must include a results array.")

        for entry in entries:
            idx = entry.get("id")
            if idx is None:
                continue
            relevant_raw = entry.get("relevant")
            if isinstance(relevant_raw, bool):
                relevant = relevant_raw
            else:
                relevant = str(relevant_raw).strip().lower() in {"true", "yes", "1"}
            try:
                score_value = float(entry.get("score", 0))
            except (TypeError, ValueError):
                score_value = 0.0
            results[int(idx)] = (relevant, score_value)
        processed += len(batch)
        if progress_callback:
            progress_callback(processed, total)

    for idx, paper in indexed:
        relevant, score_value = results.get(idx, (False, 0.0))
        paper["is_relevant"] = relevant
        paper["relevance_score"] = score_value

    return papers

def _link_abstract_report_refs(report_markdown, references):
    """Convert [P#] references in report text to clickable in-page anchors."""
    if not report_markdown:
        return ""
    if not isinstance(references, list):
        return report_markdown

    ref_to_anchor = {}
    for entry in references:
        if not isinstance(entry, dict):
            continue
        ref = str(entry.get("ref") or "").strip().upper()
        anchor_id = str(entry.get("anchor_id") or "").strip()
        if not ref or not anchor_id:
            continue
        ref_to_anchor[ref] = anchor_id

    if not ref_to_anchor:
        return report_markdown

    def _replace(match):
        ref = f"P{match.group(1)}"
        anchor_id = ref_to_anchor.get(ref)
        if not anchor_id:
            return match.group(0)
        return f'<a href="#{html.escape(anchor_id)}">{html.escape(ref)}</a>'

    # Link any paper token reference form (e.g., P1, [P1], [P1, P2], (P3)).
    return re.sub(r"\bP\s*(\d+)\b", _replace, report_markdown, flags=re.IGNORECASE)

def _run_agent_step1(
    topic,
    model_name,
    provider="ollama",
    base_url=None,
    api_key=None,
    summary_start_callback=None,
    summary_progress_callback=None,
    summary_complete_callback=None
):
    """Step 1 (+1.5): map topic, retrieve web context, summarize."""
    questions = generate_wiki_pages(topic, model_name, provider, base_url, api_key)
    web_context = perform_deep_search(questions, max_results_per_query=1)
    web_summary = ""
    if web_context and "Scout failed" not in web_context:
        if summary_start_callback:
            summary_start_callback()
        web_summary = summarize_web_context(
            web_context,
            model_name,
            provider,
            base_url,
            api_key,
            progress_callback=summary_progress_callback
        )
        if summary_complete_callback:
            summary_complete_callback()
    return {
        "web_questions": questions,
        "web_context": web_context,
        "web_summary": web_summary,
    }

def _run_agent_step2(topic, web_context, web_summary, model_name, provider="ollama", base_url=None, api_key=None, show_debug=False):
    """Step 2: build section plan from topic + optional web context."""
    planner_input = str(topic)
    if web_summary:
        planner_input += f"\n\n--- WIKIPEDIA KNOWLEDGE (USE IN PLANNING) ---\n{web_summary}"
    elif web_context:
        safe_context = _truncate_web_context(web_context, MAX_WEB_CONTEXT_CHARS)
        planner_input += f"\n\n--- WIKIPEDIA CONTEXT ---\n{safe_context}"

    section_plan = call_prompt_chain(
        planner_input,
        "Section Plan",
        model_name,
        provider=provider,
        base_url=base_url,
        api_key=api_key,
        system_message=THESIS_SYSTEM_MESSAGE,
        show_debug=show_debug,
        response_language="text",
        stream_output=True,
    )
    return section_plan

def _run_agent_step3(section_plan, model_name, provider="ollama", base_url=None, api_key=None, show_debug=False):
    """Step 3: generate academic search queries."""
    query_prompt = QUERY_BUILDER_PROMPT.format(plan_text=section_plan)
    query_raw = call_prompt_chain(
        query_prompt,
        "Query Builder",
        model_name,
        provider=provider,
        base_url=base_url,
        api_key=api_key,
        show_debug=show_debug,
        stream_output=True,
    )
    query_bundle = parse_json_response(query_raw, "Query Builder")
    if not isinstance(query_bundle, dict):
        raise ValueError("Query builder response must be a JSON object.")

    section_queries = (
        query_bundle.get("section_queries")
        or query_bundle.get("queries")
        or {}
    )
    if not isinstance(section_queries, dict) or not section_queries:
        raise ValueError("Query builder response missing or invalid 'section_queries'.")
    return section_queries

def generate_abstract_report(topic, papers, model_name, provider="ollama", base_url=None, api_key=None, show_debug=False, batch_size=5, progress_callback=None):
    """Generate a report from relevant paper abstracts in batches.

    Returns:
        dict: {
            "report_markdown": str,
            "references": [{"ref": "P1", "paper_id": "...", "title": "...", "anchor_id": "..."}]
        }
    """
    if not papers:
        return {"report_markdown": "", "references": []}

    relevant_papers = []
    for paper in papers:
        if not paper.get("is_relevant", False):
            continue
        abstract_text = str(paper.get("abstract", "") or "").strip()
        if not abstract_text or abstract_text == "No abstract.":
            continue
        relevant_papers.append(paper)

    if not relevant_papers:
        return {"report_markdown": "", "references": []}

    try:
        references = []
        indexed_relevant = []
        for idx, paper in enumerate(relevant_papers, 1):
            ref_token = f"P{idx}"
            paper_id = str(paper.get("paper_id") or "")
            anchor_id = f"paper-ref-{ref_token.lower()}"
            references.append(
                {
                    "ref": ref_token,
                    "paper_id": paper_id,
                    "title": str(paper.get("title", "")),
                    "anchor_id": anchor_id,
                }
            )
            indexed_relevant.append(
                {
                    "ref": ref_token,
                    "paper_id": paper_id,
                    "title": paper.get("title", ""),
                    "year": paper.get("year", ""),
                    "abstract": paper.get("abstract", ""),
                }
            )

        batch_analyses = []
        batch_size = max(1, int(batch_size))

        total_batches = (len(indexed_relevant) + batch_size - 1) // batch_size
        for start in range(0, len(indexed_relevant), batch_size):
            batch = indexed_relevant[start:start + batch_size]
            batch_number = (start // batch_size) + 1
            payload = {
                "topic": topic,
                "batch_number": batch_number,
                "papers": batch,
            }
            prompt = ABSTRACT_BATCH_ANALYSIS_PROMPT.format(
                topic=topic,
                input_json=json.dumps(payload, ensure_ascii=False, indent=2),
            )
            analysis = call_prompt_chain(
                prompt,
                f"Prompt: Abstract Batch Analysis {batch_number}",
                model_name,
                provider=provider,
                base_url=base_url,
                api_key=api_key,
                show_debug=show_debug,
                response_language="text",
                stream_output=False,
                show_response=False,
            )
            if analysis and analysis.strip():
                batch_analyses.append(f"## Batch {batch_number}\n{analysis.strip()}")
            if progress_callback:
                progress_callback(batch_number, total_batches, "batches")

        if not batch_analyses:
            return {"report_markdown": "", "references": references}

        paper_catalog = indexed_relevant
        if progress_callback:
            progress_callback(total_batches, total_batches, "synthesis")
        final_prompt = ABSTRACT_FINAL_REPORT_PROMPT.format(
            topic=topic,
            batch_analyses="\n\n".join(batch_analyses),
            paper_catalog=json.dumps(paper_catalog, ensure_ascii=False, indent=2),
        )
        final_report = call_prompt_chain(
            final_prompt,
            "Prompt: Abstract Synthesis Report",
            model_name,
            provider=provider,
            base_url=base_url,
            api_key=api_key,
            show_debug=show_debug,
            response_language="text",
            stream_output=False,
            show_response=False,
        )
        return {
            "report_markdown": (final_report or "").strip(),
            "references": references,
        }
    except Exception as exc:
        flash(f"Abstract-based report generation failed: {exc}", level="warning")
        return {"report_markdown": "", "references": []}

def strip_quotes_in_sections(sections):
    """Remove wrapping quotes from section headings returned by the LLM."""
    if not isinstance(sections, list):
        return sections
    cleaned = []
    for item in sections:
        if isinstance(item, str):
            cleaned.append(item.strip().strip('"').strip("'"))
        else:
            cleaned.append(item)
    return cleaned

# -----------------------------------------------------------------------------
# API configuration (file + environment + Streamlit secrets overrides).
# -----------------------------------------------------------------------------
_API_CONFIG_KEYS = (
    "openalex_email",
    "s2_api_key",
    "elsevier_api_key",
    "glm_api_key",
    "glm_base_url",
    "llm_provider",
    "chromium_path",
    "chromedriver_path",
    "chromium_profile_dir",
    "use_chromium_downloads",
    "use_remote_debugging",
    "remote_debug_port",
    "reduce_automation",
)

def _read_json_file(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _migrate_legacy_api_config() -> None:
    if API_CONFIG_FILE.exists():
        return
    if not LEGACY_API_CONFIG_FILE.exists():
        return
    payload = _read_json_file(LEGACY_API_CONFIG_FILE)
    if not isinstance(payload, dict) or not payload:
        return
    try:
        API_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        API_CONFIG_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except Exception:
        return

def _coerce_config_value(key: str, value):
    if key in {"use_chromium_downloads", "use_remote_debugging", "reduce_automation"}:
        return str(value).strip().lower() in {"1", "true", "yes", "on"}
    if key == "remote_debug_port":
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    return value

def _get_config_overrides():
    overrides = {}
    locked = set()
    for key in _API_CONFIG_KEYS:
        env_value = os.getenv(f"RESEARCH_ASSISTANT_{key.upper()}")
        if env_value not in (None, ""):
            overrides[key] = _coerce_config_value(key, env_value)
            locked.add(key)
        secret_value = None
        try:
            if key in st.secrets:
                secret_value = st.secrets.get(key)
        except Exception:
            secret_value = None
        if secret_value not in (None, ""):
            overrides[key] = _coerce_config_value(key, secret_value)
            locked.add(key)
    return overrides, locked

def load_api_config():
    """Load API configuration from local file."""
    _migrate_legacy_api_config()
    if not API_CONFIG_FILE.exists():
        file_config = {}
    else:
        file_config = _read_json_file(API_CONFIG_FILE)
    if not isinstance(file_config, dict):
        file_config = {}
    overrides, locked = _get_config_overrides()
    merged = {}
    merged.update(file_config)
    merged.update(overrides)
    st.session_state["_config_locked_keys"] = locked
    return merged

def save_api_config(config):
    """Save API configuration to local file."""
    try:
        API_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        locked = st.session_state.get("_config_locked_keys") or set()
        safe_config = {}
        for key, value in (config or {}).items():
            if key in locked:
                continue
            safe_config[key] = value
        API_CONFIG_FILE.write_text(json.dumps(safe_config, indent=2), encoding="utf-8")
        return True
    except Exception:
        return False


# =============================================================================
# PROJECT & SESSION PERSISTENCE (STATE + MIGRATION)
# =============================================================================
def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _get_projects_root() -> Path:
    return _ensure_dir(PROJECTS_DIR)


def _get_project_state_path() -> Path:
    return PROJECT_STATE_FILE


def _load_project_state() -> dict:
    path = _get_project_state_path()
    if not path.exists():
        return json.loads(json.dumps(PROJECT_STATE_TEMPLATE))
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    payload.setdefault("projects", {})
    payload.setdefault("doi_registry", {})
    payload.setdefault("active_session", {"project": None})
    if not isinstance(payload.get("projects"), dict):
        payload["projects"] = {}
    if not isinstance(payload.get("doi_registry"), dict):
        payload["doi_registry"] = {}
    if not isinstance(payload.get("active_session"), dict):
        payload["active_session"] = {"project": None}
    projects = payload.get("projects", {})
    if isinstance(projects, dict):
        for name, info in list(projects.items()):
            if not isinstance(info, dict):
                projects[name] = {WRITER_SESSIONS_KEY: []}
                continue
            if "pages" in info:
                info.pop("pages", None)
            if not isinstance(info.get(WRITER_SESSIONS_KEY), list):
                info[WRITER_SESSIONS_KEY] = []
        payload["projects"] = projects
    return payload


def _save_project_state(state: dict) -> None:
    path = _get_project_state_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def _prune_doi_registry_for_project(state: dict, project_name: str) -> bool:
    if not project_name:
        return False
    registry = state.get("doi_registry")
    if not isinstance(registry, dict):
        return False
    changed = False
    empty_keys = []
    for doi, entries in registry.items():
        if not isinstance(entries, list):
            continue
        filtered = [
            entry for entry in entries
            if not (isinstance(entry, dict) and entry.get("project") == project_name)
        ]
        if len(filtered) != len(entries):
            changed = True
        if filtered:
            registry[doi] = filtered
        else:
            empty_keys.append(doi)
    for doi in empty_keys:
        registry.pop(doi, None)
    if empty_keys:
        changed = True
    if changed:
        state["doi_registry"] = registry
    return changed

def _validate_project_name(name: str):
    if not name:
        return None
    value = str(name).strip()
    if not value:
        return None
    if ".." in value:
        return None
    if "/" in value or "\\" in value:
        return None
    if os.path.sep in value:
        return None
    if os.path.altsep and os.path.altsep in value:
        return None
    return value

def _record_invalid_project_names(names):
    if not names:
        return
    try:
        existing = st.session_state.get("_invalid_project_names", [])
        if not isinstance(existing, list):
            existing = []
        st.session_state["_invalid_project_names"] = sorted(set(existing) | set(names))
    except Exception:
        return

def _consume_invalid_project_names():
    names = st.session_state.pop("_invalid_project_names", [])
    if not isinstance(names, list):
        return []
    return names


def _get_project_dir(project_name: str) -> Path:
    safe_name = _validate_project_name(project_name)
    if not safe_name:
        raise ValueError("Invalid project name")
    return _get_projects_root() / safe_name


def _ensure_project_layout(project_name: str) -> Path:
    project_dir = _get_project_dir(project_name)
    for folder in ("PDF", "RIS", "external RIS", "WORD", "tracking", "db", "sessions"):
        _ensure_dir(project_dir / folder)
    return project_dir


def _get_project_pdf_dir(project_name: str) -> Path:
    return _ensure_dir(_get_project_dir(project_name) / "PDF")


def _get_project_ris_dir(project_name: str) -> Path:
    return _ensure_dir(_get_project_dir(project_name) / "RIS")


def _get_project_external_ris_dir(project_name: str) -> Path:
    return _ensure_dir(_get_project_dir(project_name) / "external RIS")


def _get_project_sessions_dir(project_name: str) -> Path:
    return _ensure_dir(_get_project_dir(project_name) / "sessions")


def _get_session_dir(project_name: str, session_id: int) -> Path:
    return _get_project_sessions_dir(project_name) / f"session_{session_id}"


def _get_session_file(project_name: str, session_id: int) -> Path:
    return _get_session_dir(project_name, session_id) / "session.json"


def _get_session_export_dir(project_name: str, session_id: int) -> Path:
    return _ensure_dir(_get_session_dir(project_name, session_id) / "exports")


def _sanitize_topic(topic: str) -> str:
    safe = re.sub(r"[^\w\s-]", "", topic or "").strip().replace(" ", "_")
    return safe[:50] or "research"


def _coerce_session_id(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _session_ids_from_disk(project_name: str) -> list:
    sessions_dir = _get_project_sessions_dir(project_name)
    if not sessions_dir.exists():
        return []
    ids = []
    for entry in sessions_dir.iterdir():
        if entry.is_dir() and entry.name.startswith("session_"):
            session_id = _coerce_session_id(entry.name.replace("session_", ""))
            if session_id is not None:
                ids.append(session_id)
    return sorted(ids)


def _sync_project_state_with_disk(state: dict) -> tuple:
    changed = False
    invalid_names = []
    for name in list(state.get("projects", {}).keys()):
        if not _validate_project_name(name):
            invalid_names.append(name)
    if invalid_names:
        for name in invalid_names:
            state.get("projects", {}).pop(name, None)
        _record_invalid_project_names(invalid_names)
        changed = True
    projects_root = _get_projects_root()
    if projects_root.exists():
        for entry in projects_root.iterdir():
            if not entry.is_dir():
                continue
            if not _validate_project_name(entry.name):
                _record_invalid_project_names([entry.name])
                continue
            if entry.name not in state["projects"]:
                state["projects"][entry.name] = {WRITER_SESSIONS_KEY: []}
                changed = True

    for project_name in list(state["projects"].keys()):
        if not _validate_project_name(project_name):
            continue
        info = state["projects"].get(project_name)
        if not isinstance(info, dict):
            info = {WRITER_SESSIONS_KEY: []}
            state["projects"][project_name] = info
            changed = True
        if "pages" in info:
            info.pop("pages", None)
            changed = True
        if not isinstance(info.get(WRITER_SESSIONS_KEY), list):
            info[WRITER_SESSIONS_KEY] = []
            changed = True

        _ensure_project_layout(project_name)
        disk_ids = _session_ids_from_disk(project_name)
        if disk_ids:
            merged = sorted(set(info[WRITER_SESSIONS_KEY]) | set(disk_ids))
            if merged != info[WRITER_SESSIONS_KEY]:
                info[WRITER_SESSIONS_KEY] = merged
                changed = True

    return state, changed


def _load_legacy_projects_map() -> dict:
    if not LEGACY_PROJECTS_FILE.exists():
        return {}
    try:
        payload = json.loads(LEGACY_PROJECTS_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    cleaned = {}
    for name, ids in payload.items():
        if not isinstance(name, str) or not isinstance(ids, list):
            continue
        normalized = []
        for sid in ids:
            sid_int = _coerce_session_id(sid)
            if sid_int is not None:
                normalized.append(sid_int)
        if normalized:
            cleaned[name] = sorted(set(normalized))
    return cleaned


def _scan_legacy_session_ids() -> list:
    if not LEGACY_SESSION_DIR.exists():
        return []
    ids = []
    for entry in LEGACY_SESSION_DIR.glob("session_*.json"):
        session_id = _coerce_session_id(entry.stem.replace("session_", ""))
        if session_id is not None:
            ids.append(session_id)
    return sorted(set(ids))


def _copy_if_missing(src: Path, dest: Path) -> None:
    try:
        if dest.exists():
            return
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)
    except Exception:
        return


def _migrate_legacy_writer_storage() -> None:
    if not LEGACY_SESSION_DIR.exists():
        return

    legacy_projects = _load_legacy_projects_map()
    if not legacy_projects:
        legacy_ids = _scan_legacy_session_ids()
        if legacy_ids:
            legacy_projects = {"Thesis": legacy_ids}

    if not legacy_projects:
        return

    state = _load_project_state()
    changed = False

    for project_name, session_ids in legacy_projects.items():
        if not session_ids:
            continue
        _ensure_project_layout(project_name)
        project_entry = state["projects"].setdefault(project_name, {WRITER_SESSIONS_KEY: []})
        if not isinstance(project_entry, dict):
            project_entry = {WRITER_SESSIONS_KEY: []}
            state["projects"][project_name] = project_entry
            changed = True
        if "pages" in project_entry:
            project_entry.pop("pages", None)
            changed = True
        if not isinstance(project_entry.get(WRITER_SESSIONS_KEY), list):
            project_entry[WRITER_SESSIONS_KEY] = []
            changed = True

        for session_id in session_ids:
            legacy_file = LEGACY_SESSION_DIR / f"session_{session_id}.json"
            if not legacy_file.exists():
                continue
            try:
                session_data = json.loads(legacy_file.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(session_data, dict):
                continue
            session_data.setdefault("project", project_name)

            session_dir = _get_session_dir(project_name, session_id)
            _ensure_dir(session_dir)
            new_file = _get_session_file(project_name, session_id)
            if not new_file.exists():
                new_file.write_text(
                    json.dumps(session_data, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
            else:
                try:
                    existing = json.loads(new_file.read_text(encoding="utf-8"))
                except Exception:
                    existing = {}
                if isinstance(existing, dict) and not existing.get("project"):
                    existing["project"] = project_name
                    new_file.write_text(
                        json.dumps(existing, indent=2, ensure_ascii=False),
                        encoding="utf-8",
                    )

            if session_id not in project_entry[WRITER_SESSIONS_KEY]:
                project_entry[WRITER_SESSIONS_KEY].append(session_id)
                changed = True

            if LEGACY_EXPORTS_DIR.exists():
                sanitized_topic = _sanitize_topic(session_data.get("topic", "research"))
                legacy_export_dir = LEGACY_EXPORTS_DIR / f"session_{session_id}_{sanitized_topic}"
                if legacy_export_dir.exists():
                    project_pdf_dir = _get_project_pdf_dir(project_name)
                    for pdf in legacy_export_dir.glob("*.pdf"):
                        _copy_if_missing(pdf, project_pdf_dir / pdf.name)
                    session_export_dir = _get_session_export_dir(project_name, session_id)
                    for xlsx in legacy_export_dir.glob("*.xlsx"):
                        _copy_if_missing(xlsx, session_export_dir / xlsx.name)
                    legacy_ris_dir = legacy_export_dir / f"RIS_session_{session_id}_{sanitized_topic}"
                    if legacy_ris_dir.exists():
                        project_ris_dir = _get_project_ris_dir(project_name)
                        for ris in legacy_ris_dir.glob("*.ris"):
                            _copy_if_missing(ris, project_ris_dir / ris.name)

        project_entry[WRITER_SESSIONS_KEY] = sorted(set(project_entry[WRITER_SESSIONS_KEY]))

    if changed:
        _save_project_state(state)


def get_next_session_id(project_name: str) -> int:
    """Generate the next sequential session ID for a project."""
    existing = _session_ids_from_disk(project_name)
    return max(existing) + 1 if existing else 1


def save_session(project_name: str, session_id: int, data: dict) -> None:
    """Save session data under the project session directory."""
    safe_name = _validate_project_name(project_name)
    if not safe_name:
        flash("Invalid project name; session not saved.", level="warning")
        return
    project_name = safe_name
    _ensure_project_layout(project_name)
    session_dir = _get_session_dir(project_name, session_id)
    _ensure_dir(session_dir)
    data = data or {}
    data["project"] = project_name
    filepath = _get_session_file(project_name, session_id)
    tmp_path = filepath.with_suffix(".json.tmp")
    tmp_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    tmp_path.replace(filepath)
    papers = (data.get("found_papers") or []) + (data.get("missing_papers") or [])
    _update_doi_registry_for_session(project_name, session_id, papers)


def _ensure_manual_session(topic_hint=None, num_results=0):
    existing = st.session_state.get("active_session")
    if isinstance(existing, dict):
        return existing

    curr_proj = st.session_state.get("current_project") or "Thesis"
    safe_proj = _validate_project_name(curr_proj)
    if not safe_proj:
        flash("Invalid project name detected. Falling back to 'Thesis'.", level="warning")
        curr_proj = "Thesis"
    else:
        curr_proj = safe_proj
    projects_data = load_projects()
    if curr_proj not in projects_data:
        create_project(curr_proj, projects_data)
        projects_data = load_projects()

    session_id = get_next_session_id(curr_proj)
    session_topic = topic_hint or "Manual Search"
    session_data = {
        "session_id": session_id,
        "timestamp": datetime.now().isoformat(),
        "topic": session_topic,
        "num_results_per_query": num_results,
        "web_questions": [],
        "web_context": "",
        "web_summary": "",
        "section_plan": "",
        "section_queries": {},
        "found_papers": [],
        "missing_papers": [],
        "manual_papers": [],
        "abstract_report": "",
        "abstract_report_references": [],
        "project": curr_proj,
    }
    save_session(curr_proj, session_id, session_data)

    if curr_proj in projects_data:
        projects_data[curr_proj].append(session_id)
        projects_data[curr_proj] = sorted(list(set(projects_data[curr_proj])))
        save_projects(projects_data)

    st.session_state["active_session"] = session_data
    st.session_state["current_session_id"] = session_id
    return session_data


def load_session(project_name: str, session_id: int):
    """Load session data from a project session directory."""
    safe_name = _validate_project_name(project_name)
    if not safe_name:
        return None
    project_name = safe_name
    filepath = _get_session_file(project_name, session_id)
    if not filepath.exists():
        return None
    try:
        data = json.loads(filepath.read_text(encoding="utf-8"))
    except Exception:
        return None
    if isinstance(data, dict) and not data.get("project"):
        data["project"] = project_name
        filepath.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return data


def delete_session(project_name: str, session_id: int) -> None:
    """Delete a session folder and remove it from project state."""
    safe_name = _validate_project_name(project_name)
    if not safe_name:
        return
    project_name = safe_name
    session_dir = _get_session_dir(project_name, session_id)
    if session_dir.exists():
        shutil.rmtree(session_dir, ignore_errors=True)

    state = _load_project_state()
    project_entry = state.get("projects", {}).get(project_name)
    if isinstance(project_entry, dict):
        sessions = project_entry.get(WRITER_SESSIONS_KEY, [])
        if isinstance(sessions, list):
            project_entry[WRITER_SESSIONS_KEY] = [sid for sid in sessions if sid != session_id]
            _save_project_state(state)


def load_projects():
    """Load projects from the standalone project_state file."""
    state = _load_project_state()
    state, changed = _sync_project_state_with_disk(state)
    if changed:
        _save_project_state(state)
    projects = {}
    for name, info in state.get("projects", {}).items():
        if not _validate_project_name(name):
            _record_invalid_project_names([name])
            continue
        sessions = info.get(WRITER_SESSIONS_KEY, []) if isinstance(info, dict) else []
        normalized = []
        for sid in sessions:
            sid_int = _coerce_session_id(sid)
            if sid_int is not None:
                normalized.append(sid_int)
        projects[name] = sorted(set(normalized))
    return projects


def save_projects(data):
    """Persist writer session lists into the standalone project_state file."""
    state = _load_project_state()
    changed = False
    for name, session_ids in (data or {}).items():
        safe_name = _validate_project_name(name)
        if not safe_name:
            _record_invalid_project_names([name])
            continue
        project_entry = state.setdefault("projects", {}).setdefault(safe_name, {WRITER_SESSIONS_KEY: []})
        if not isinstance(project_entry, dict):
            project_entry = {WRITER_SESSIONS_KEY: []}
            state["projects"][safe_name] = project_entry
            changed = True
        normalized = []
        for sid in session_ids:
            sid_int = _coerce_session_id(sid)
            if sid_int is not None:
                normalized.append(sid_int)
        normalized = sorted(set(normalized))
        if project_entry.get(WRITER_SESSIONS_KEY) != normalized:
            project_entry[WRITER_SESSIONS_KEY] = normalized
            changed = True
    if changed:
        _save_project_state(state)
    return True


def init_projects():
    """Initialize projects from the standalone directory structure."""
    if os.getenv("RESEARCH_ASSISTANT_MIGRATE_LEGACY") == "1":
        _migrate_legacy_writer_storage()
    state = _load_project_state()
    state, changed = _sync_project_state_with_disk(state)
    if not state.get("projects"):
        state["projects"] = {"Thesis": {WRITER_SESSIONS_KEY: []}}
        _ensure_project_layout("Thesis")
        changed = True
    for project_name in state["projects"]:
        _ensure_project_layout(project_name)
    if changed:
        _save_project_state(state)
    return {
        name: info.get(WRITER_SESSIONS_KEY, [])
        for name, info in state.get("projects", {}).items()
        if isinstance(info, dict) and _validate_project_name(name)
    }


def create_project(name, projects_data):
    """Create a new project in the standalone directory tree."""
    safe_name = _validate_project_name(name)
    if not safe_name:
        flash("Invalid project name. Avoid path separators or '..'.", level="warning", scope="sidebar")
        return False
    state = _load_project_state()
    if safe_name in state.get("projects", {}):
        return False
    state.setdefault("projects", {})[safe_name] = {WRITER_SESSIONS_KEY: []}
    _save_project_state(state)
    _ensure_project_layout(safe_name)
    projects_data[safe_name] = []
    return True


def delete_project(name, projects_data):
    """Delete a project directory and remove it from standalone project state."""
    safe_name = _validate_project_name(name)
    if not safe_name:
        flash("Invalid project name. Avoid path separators or '..'.", level="warning", scope="sidebar")
        return False
    state = _load_project_state()
    if safe_name not in state.get("projects", {}):
        return False
    del state["projects"][safe_name]
    _prune_doi_registry_for_project(state, safe_name)
    active_session = state.get("active_session", {})
    if isinstance(active_session, dict) and active_session.get("project") == safe_name:
        state["active_session"] = {"project": None}
    _save_project_state(state)
    project_dir = _get_project_dir(safe_name)
    if project_dir.exists():
        shutil.rmtree(project_dir, ignore_errors=True)
    projects_data.pop(safe_name, None)
    return True


def get_project_sessions(project_name, projects_data):
    """Get list of session objects for a specific project."""
    safe_name = _validate_project_name(project_name)
    if not safe_name:
        return []
    if safe_name not in projects_data:
        return []
    session_ids = projects_data[safe_name]
    sessions = []
    valid_ids = []

    for sid in session_ids:
        sid_int = _coerce_session_id(sid)
        if sid_int is None:
            continue
        data = load_session(safe_name, sid_int)
        if data:
            valid_ids.append(sid_int)
            sessions.append({
                "id": sid_int,
                "topic": data.get("topic", "Unknown"),
                "timestamp": data.get("timestamp", ""),
                "paper_count": len(data.get("found_papers", [])),
            })

    if len(valid_ids) != len(session_ids):
        projects_data[safe_name] = valid_ids
        save_projects(projects_data)

    return sorted(sessions, key=lambda x: x["id"], reverse=True)


def list_sessions():
    """Return all writer sessions across projects."""
    sessions = []
    projects = load_projects()
    for project_name, session_ids in projects.items():
        for sid in session_ids:
            data = load_session(project_name, sid)
            if data:
                sessions.append({
                    "id": sid,
                    "project": project_name,
                    "topic": data.get("topic", "Unknown"),
                    "timestamp": data.get("timestamp", ""),
                    "paper_count": len(data.get("found_papers", [])),
                })
    return sessions


def find_session_project(session_id):
    """Resolve a project name for a known session id."""
    projects = load_projects()
    for project_name, session_ids in projects.items():
        if session_id in session_ids:
            return project_name
    return None

# =============================================================================
# PRESENTATION + DATA NORMALIZATION HELPERS
# =============================================================================

def get_quality_badge(impact):
    if not impact: return ""
    if impact > 5: return f'<span class="badge-high">HIGH IMPACT ({impact:.1f})</span>'
    if impact > 2: return f'<span class="badge-med">IMPACT {impact:.1f}</span>'
    return f'<span class="badge-low">IMPACT {impact:.1f}</span>'

def _format_card_text(text: str) -> str:
    if text is None:
        return ""
    value = str(text).strip()
    if not value:
        return ""
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"\n{2,}", "\n", value)
    escaped = html.escape(value)
    return escaped.replace("\n", "<br/>")


def _wrap_note_with_llm(note_text: str, model_name: str, provider: str, base_url: str, api_key: str) -> str:
    if not note_text:
        return ""
    model_to_use = model_name or WRAP_NOTE_MODEL
    raw = _llm_generate(
        note_text,
        model_to_use,
        provider=provider,
        base_url=base_url,
        api_key=api_key,
        system_message=WRAP_NOTE_SYSTEM_MESSAGE,
        max_tokens=1200,
        temperature=WRAP_NOTE_TEMPERATURE,
        reasoning_effort=WRAP_NOTE_REASONING
    )
    if not raw:
        return ""
    _, cleaned = split_reasoning_and_answer(raw)
    return cleaned or raw


def _parse_ris_metadata(ris_text: str) -> dict:
    if not ris_text:
        return {}
    metadata = {}
    authors = []
    for line in ris_text.splitlines():
        if " - " not in line:
            continue
        tag, value = line.split(" - ", 1)
        tag = tag.strip()
        value = value.strip()
        if not value:
            continue
        if tag in {"TI", "T1"} and not metadata.get("title"):
            metadata["title"] = value
        elif tag in {"AU", "A1"}:
            authors.append(value)
        elif tag in {"PY", "Y1"} and not metadata.get("year"):
            match = re.search(r"\d{4}", value)
            metadata["year"] = match.group(0) if match else value
        elif tag in {"JO", "JF", "JA", "T2"} and not metadata.get("journal"):
            metadata["journal"] = value
        elif tag == "DO" and not metadata.get("doi"):
            metadata["doi"] = value
    if authors:
        metadata["authors"] = authors
    if metadata.get("doi"):
        metadata["doi"] = _normalize_doi(metadata["doi"])
    return metadata


def _get_ris_metadata_for_doi(project_name: str, doi: str) -> dict:
    normalized = _normalize_doi(doi)
    if not normalized or not project_name:
        return {}
    cache = st.session_state.setdefault("ris_metadata_cache", {})
    cached = cache.get(normalized)
    if isinstance(cached, dict):
        return cached
    ris_dirs = [_get_project_ris_dir(project_name), _get_project_external_ris_dir(project_name)]
    for ris_dir in ris_dirs:
        ris_path = Path(ris_dir)
        if not ris_path.exists():
            continue
        for item in ris_path.rglob("*.ris"):
            if not item.is_file():
                continue
            try:
                ris_text = item.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            metadata = _parse_ris_metadata(ris_text)
            if _normalize_doi(metadata.get("doi")) == normalized:
                cache[normalized] = metadata
                return metadata
    cache[normalized] = {}
    return {}


def _note_key_for_paper(paper: dict) -> str:
    doi = _normalize_doi(paper.get("doi"))
    if doi:
        return f"doi:{doi}"
    paper_id = paper.get("paper_id")
    if paper_id:
        return f"id:{paper_id}"
    title = str(paper.get("title") or "").strip()
    if title:
        digest = hashlib.sha256(title.encode("utf-8")).hexdigest()[:12]
        return f"title:{digest}"
    return ""


def _get_active_session_notes() -> dict:
    session_data = st.session_state.get("active_session", {})
    notes = session_data.get("paper_notes", {})
    return notes if isinstance(notes, dict) else {}

def _get_active_session_highlights() -> dict:
    session_data = st.session_state.get("active_session", {})
    highlights = session_data.get("paper_highlights", {})
    return highlights if isinstance(highlights, dict) else {}


def _save_active_session_note(
    project_name: str,
    session_id: int,
    note_key: str,
    content: str,
    paper: dict,
    wrap_content: str = ""
) -> None:
    if not (project_name and session_id and note_key):
        return
    session_data = st.session_state.get("active_session") or load_session(project_name, session_id) or {}
    if not isinstance(session_data, dict):
        session_data = {}
    notes = session_data.get("paper_notes", {})
    if not isinstance(notes, dict):
        notes = {}
    trimmed = (content or "").strip()
    wrapped = (wrap_content or "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if trimmed:
        existing = notes.get(note_key, {}) if isinstance(notes.get(note_key), dict) else {}
        doi = _normalize_doi(paper.get("doi"))
        ris_metadata = existing.get("ris_metadata") if isinstance(existing.get("ris_metadata"), dict) else {}
        if doi:
            found_metadata = _get_ris_metadata_for_doi(project_name, doi)
            if found_metadata:
                ris_metadata = found_metadata
        notes[note_key] = {
            "content": trimmed,
            "wrap_up": wrapped,
            "doi": doi,
            "title": paper.get("title"),
            "paper_id": paper.get("paper_id"),
            "ris_metadata": ris_metadata or {},
            "created_at": existing.get("created_at", now),
            "updated_at": now,
        }
    else:
        notes.pop(note_key, None)

    session_data["paper_notes"] = notes
    st.session_state["active_session"] = session_data
    save_session(project_name, session_id, session_data)

def _save_active_session_highlights(
    project_name: str,
    session_id: int,
    note_key: str,
    highlights: list
) -> None:
    if not (project_name and session_id and note_key):
        return
    session_data = st.session_state.get("active_session") or load_session(project_name, session_id) or {}
    if not isinstance(session_data, dict):
        session_data = {}
    highlight_map = session_data.get("paper_highlights", {})
    if not isinstance(highlight_map, dict):
        highlight_map = {}
    if highlights:
        highlight_map[note_key] = highlights
    else:
        highlight_map.pop(note_key, None)
    session_data["paper_highlights"] = highlight_map
    st.session_state["active_session"] = session_data
    save_session(project_name, session_id, session_data)

def _render_pdf_viewer_with_highlights(
    pdf_path,
    paper: dict,
    paper_key: str,
    project_name: str,
    session_id: int,
    width: int = 1000,
    height: int = 600,
) -> None:
    note_key = _note_key_for_paper(paper)
    annotations = []
    if note_key:
        annotations = _get_active_session_highlights().get(note_key, [])
        if not isinstance(annotations, list):
            annotations = []
    pdf_viewer(
        str(pdf_path),
        width=width,
        height=height,
        render_text=True,
        annotations=annotations,
    )
    if not note_key:
        return
    with st.popover("Highlights"):
        st.caption("Coordinates are in PDF points, page numbers start at 1.")
        page = st.number_input("Page", min_value=1, value=1, key=f"hl_page_{paper_key}")
        x = st.number_input("X", min_value=0.0, value=0.0, step=1.0, key=f"hl_x_{paper_key}")
        y = st.number_input("Y", min_value=0.0, value=0.0, step=1.0, key=f"hl_y_{paper_key}")
        width_val = st.number_input("Width", min_value=1.0, value=200.0, step=1.0, key=f"hl_w_{paper_key}")
        height_val = st.number_input("Height", min_value=1.0, value=50.0, step=1.0, key=f"hl_h_{paper_key}")
        color = st.text_input(
            "Color",
            value="rgba(255, 235, 59, 0.35)",
            key=f"hl_color_{paper_key}",
        )
        label = st.text_input("Label", value="", key=f"hl_label_{paper_key}")
        if st.button("Add Highlight", key=f"hl_add_{paper_key}"):
            new_anno = {
                "page": int(page),
                "x": float(x),
                "y": float(y),
                "width": float(width_val),
                "height": float(height_val),
                "color": color.strip() or "rgba(255, 235, 59, 0.35)",
                "content": label.strip(),
            }
            updated = list(annotations)
            updated.append(new_anno)
            _save_active_session_highlights(project_name, session_id, note_key, updated)
            st.rerun()
        if annotations:
            if st.button("Clear Highlights", key=f"hl_clear_{paper_key}", type="secondary"):
                _save_active_session_highlights(project_name, session_id, note_key, [])
                st.rerun()


def _render_note_popover(
    widget_prefix: str,
    project_name: str,
    session_id: int,
    note_key: str,
    note_content: str,
    wrap_content: str,
    paper: dict,
    height: int = 120
) -> None:
    input_key = f"{widget_prefix}_note_input"
    pending_key = f"{input_key}_pending"
    wrap_key = f"{widget_prefix}_note_wrap_text"
    if pending_key in st.session_state:
        st.session_state[input_key] = st.session_state.pop(pending_key)
    if input_key not in st.session_state:
        st.session_state[input_key] = note_content or ""
    if wrap_key not in st.session_state:
        st.session_state[wrap_key] = wrap_content or ""
    with st.popover("Add note", use_container_width=False):
        st.text_area(
            "Note",
            key=input_key,
            height=height,
            placeholder="Write a note for this paper."
        )
        if note_key:
            notes = _get_active_session_notes()
            note_payload = notes.get(note_key, {}) if isinstance(notes, dict) else {}
            ris_metadata = note_payload.get("ris_metadata", {}) if isinstance(note_payload, dict) else {}
            if isinstance(ris_metadata, dict) and ris_metadata:
                meta_lines = []
                doi_value = ris_metadata.get("doi")
                year_value = ris_metadata.get("year")
                journal_value = ris_metadata.get("journal")
                authors_value = ris_metadata.get("authors")
                if doi_value:
                    meta_lines.append(f"DOI: {doi_value}")
                if year_value or journal_value:
                    year_label = year_value or "Unknown"
                    journal_label = journal_value or "Unknown"
                    meta_lines.append(f"Year: {year_label} | Journal: {journal_label}")
                if authors_value:
                    if isinstance(authors_value, list):
                        authors_label = "; ".join(authors_value)
                    else:
                        authors_label = str(authors_value)
                    meta_lines.append(f"Authors: {authors_label}")
                if meta_lines:
                    meta_html = "<br/>".join(html.escape(line) for line in meta_lines)
                    st.markdown(
                        f'<div class="note-metadata">'
                        f'<strong style="color: #c9d1d9;">Metadata (RIS)</strong><br/>{meta_html}</div>',
                        unsafe_allow_html=True
                    )
        if st.button("Save note", key=f"{widget_prefix}_note_save", type="primary", use_container_width=True):
            updated = st.session_state.get(input_key, "").strip()
            wrapped = st.session_state.get(wrap_key, "").strip()
            _save_active_session_note(
                project_name,
                session_id,
                note_key,
                updated,
                paper,
                wrap_content=wrapped
            )
            if updated:
                st.caption("Note saved.")
            else:
                st.caption("Note cleared.")
        if st.button("Wrap up note", key=f"{widget_prefix}_note_wrap", use_container_width=True):
            current = st.session_state.get(input_key, "").strip()
            if not current:
                flash("Add some text to wrap up.", level="warning")
            else:
                provider = st.session_state.get("llm_provider") or "ollama"
                if _normalize_provider(provider) == "glm":
                    base_url = st.session_state.get("glm_base_url") or GLM_BASE_URL
                    api_key = st.session_state.get("glm_api_key_runtime") or st.session_state.get("glm_api_key") or ""
                    model_name = st.session_state.get("active_model") or "glm-4.7"
                else:
                    base_url = st.session_state.get("ollama_url") or OLLAMA_URL
                    api_key = st.session_state.get("ollama_api_key") or ""
                    model_name = st.session_state.get("active_model") or PRIMARY_MODEL
                with st.spinner("Wrapping up note..."):
                    wrapped = _wrap_note_with_llm(current, model_name, provider, base_url, api_key)
                if wrapped:
                    st.session_state[wrap_key] = wrapped
                    st.caption("Wrap-up ready. Review and save.")
                else:
                    flash("Wrap up failed. Please try again.", level="warning")
        wrap_height = max(120, min(200, height))
        st.text_area(
            "Wrap-up",
            key=wrap_key,
            height=wrap_height,
            placeholder="Wrap-up result will appear here."
        )


def _normalize_doi(raw_value):
    if not raw_value:
        return None
    doi = str(raw_value).strip()
    if not doi or doi.lower() in {"unavailable", "n/a", "none"}:
        return None
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi, flags=re.IGNORECASE)
    doi = re.sub(r"^doi:\s*", "", doi, flags=re.IGNORECASE)
    doi = doi.strip()
    return doi.lower() if doi else None


def _format_session_label(page_name: str) -> str:
    if page_name.startswith("session_"):
        suffix = page_name.replace("session_", "", 1)
        return f"Session {suffix}"
    return page_name


def _update_doi_registry_for_session(project_name: str, session_id: int, papers: list) -> bool:
    if not project_name or not session_id or not papers:
        return False
    state = _load_project_state()
    registry = state.setdefault("doi_registry", {})
    if not isinstance(registry, dict):
        registry = {}
        state["doi_registry"] = registry
    changed = False
    session_key = f"session_{session_id}"

    for paper in papers:
        if not isinstance(paper, dict):
            continue
        doi = _normalize_doi(paper.get("doi"))
        if not doi:
            continue
        content_id = str(paper.get("paper_id") or paper.get("title") or doi)
        entry = {
            "project": project_name,
            "page": session_key,
            "content_id": content_id,
            "source_type": "writer_session",
        }
        bucket = registry.setdefault(doi, [])
        if not isinstance(bucket, list):
            bucket = []
        if not any(
            isinstance(existing, dict)
            and existing.get("project") == entry["project"]
            and existing.get("page") == entry["page"]
            and str(existing.get("content_id")) == entry["content_id"]
            for existing in bucket
        ):
            bucket.append(entry)
            registry[doi] = bucket
            changed = True
    if changed:
        state["doi_registry"] = registry
        _save_project_state(state)
    return changed


def _get_doi_cross_project_entries(doi: str, project_name: str, session_id: int) -> list:
    normalized = _normalize_doi(doi)
    if not normalized:
        return []
    state = _load_project_state()
    registry = state.get("doi_registry", {})
    if not isinstance(registry, dict):
        return []
    entries = registry.get(normalized, [])
    if not isinstance(entries, list):
        return []
    session_key = f"session_{session_id}" if session_id else None
    filtered = []
    seen = set()
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        proj = entry.get("project")
        page = entry.get("page") or entry.get("session")
        if not proj or not page:
            continue
        if proj == project_name and session_key and page == session_key:
            continue
        key = f"{proj}::{page}"
        if key in seen:
            continue
        seen.add(key)
        filtered.append({"project": proj, "page": page})
    return filtered


def _format_cross_project_badge(entries: list) -> str:
    if not entries:
        return ""
    labels = []
    for entry in entries:
        project = entry.get("project")
        page = entry.get("page")
        if not project or not page:
            continue
        labels.append(f"{project} / {_format_session_label(page)}")
    if not labels:
        return ""
    labels_html = " | ".join(html.escape(label) for label in labels)
    return f'<div class="paper-also-in">Also in: {labels_html}</div>'


def render_paper_card(
    p,
    is_success=True,
    downloaded=None,
    excel_exported=False,
    has_citation=False,
    also_in_html="",
    reference_tag=""
):
    """Render a clean paper card with essential information.
    
    Args:
        p: Paper data dictionary
        is_success: Whether the paper has an accessible PDF
        downloaded: None (don't show), True (show Downloaded), False (show Not Downloaded)
        excel_exported: If True, shows an "Excel" badge on the card
        has_citation: If True, shows a "Citation Added" badge on the card
    """
    impact = p.get('quality', {}).get('impact', 0) if p.get('quality') else 0
    # Prefer OpenAlex venue (from DOI lookup), fallback to ISSN-based quality name.
    venue = p.get('venue') or (p.get('quality', {}).get('name') if p.get('quality') else None) or "Unknown"
    badge_html = get_quality_badge(impact)
    relevance_tag = '<span class="tag-relevant">relevant</span>' if p.get("is_relevant") else ""
    relevance_html = f" | {relevance_tag}" if relevance_tag else ""
    
    # Influence ratio badge
    influence_badge = ""
    citations = p.get('citations', 0) or 0
    infl_citations = p.get('infl_citations', 0) or 0
    if citations > 0 and infl_citations > 0:
        influence_ratio = (infl_citations / citations) * 100
        influence_badge = f' | <span class="badge-influence">{influence_ratio:.1f}% Influential</span>'
    
    # Concept tags (ML-tagged topics from OpenAlex)
    concepts_html = ""
    concepts = p.get('concepts', [])
    if concepts:
        concept_badges = ' '.join([f'<span class="badge-concept">{c}</span>' for c in concepts[:3]])
        concepts_html = f' | {concept_badges}'
    
    # Download status badge
    download_badge = ""
    if downloaded is True:
        download_badge = ' | <span class="badge-downloaded">Downloaded</span>'
    elif downloaded is False:
        download_badge = ' | <span class="badge-pending">Not Downloaded</span>'
    
    # Excel export badge
    excel_badge = ""
    if excel_exported:
        excel_badge = ' | <span class="badge-excel">Excel</span>'
    
    # Citation badge
    citation_badge = ""
    if has_citation:
        citation_badge = ' | <span class="badge-citation">Citation Added</span>'
    
    reference_badge = ""
    if reference_tag:
        safe_ref = html.escape(str(reference_tag))
        reference_badge = f' | <span class="badge-paper-ref">{safe_ref}</span>'

    # PDF source badge.
    # Legacy fallback: infer OpenAlex when a PDF URL exists but source metadata is missing.
    pdf_source = p.get("pdf_source")
    if pdf_source == "OpenAlex":
        source_badge = '<span class="badge-source-openalex">[OpenAlex]</span>'
    elif pdf_source == "Elsevier":
        source_badge = '<span class="badge-source-elsevier">[Elsevier]</span>'
    elif pdf_source == "arXiv":
        source_badge = '<span class="badge-source-arxiv">[arXiv]</span>'
    elif is_success and p.get("pdf_url"):
        # Legacy fallback when source metadata is missing.
        source_badge = '<span class="badge-source-openalex">[OpenAlex]</span>'
    else:
        source_badge = '<span class="badge-source-paywalled">[Paywalled]</span>'

    css_class = "paper-card-success" if is_success else "paper-card-failed"
    title_text = html.escape(str(p.get("title", "Untitled")))
    link_html = f'<a href="{p["pdf_url"]}" target="_blank">{title_text}</a>' if is_success else title_text

    # Render TLDR (if available) and full abstract.
    content_blocks = []
    if p.get('tldr') and p['tldr'] != "No AI summary available.":
        content_blocks.append(f'<div class="paper-tldr">{_format_card_text(p["tldr"])}</div>')
    if p.get("abstract") and p["abstract"] != "No abstract.":
        content_blocks.append(f'<div class="paper-abstract">{_format_card_text(p["abstract"])}</div>')
    content_block = "\n".join(content_blocks)

    # Build DOI link HTML
    doi_html = ""
    if p.get('doi'):
        doi_url = f"https://doi.org/{p['doi']}"
        doi_html = f' | <a href="{doi_url}" target="_blank" style="color: #58a6ff; font-size: 11px;">DOI</a>'

    # Badge order: Year | Citations | Influence | Concepts | Relevance | Impact | Venue | DOI | Downloaded | Excel
    # Source badge moved to end for cleaner look
    year_display = html.escape(str(p.get("year", "N/A")))
    citations_display = html.escape(str(p.get("citations", 0)))
    venue_display = html.escape(str(venue))
    lines = [
        f'<div class="{css_class}">',
        f'<div class="paper-title">{link_html}</div>',
        f'<div class="paper-authors">{html.escape(str(p.get("authors", "")))}</div>',
        '<div class="paper-meta">',
        f'{year_display} | {citations_display} cites{influence_badge}{concepts_html}{relevance_html} | {badge_html} | {venue_display}{doi_html}{download_badge}{excel_badge}{citation_badge}{reference_badge} | {source_badge}',
        '</div>',
    ]
    if also_in_html:
        lines.append(also_in_html)
    if content_block:
        lines.append(content_block)
    lines.append('</div>')
    return "\n".join(lines)

# -----------------------------------------------------------------------------
# Agent log rendering (Step 1..4 summaries for the UI).
# -----------------------------------------------------------------------------
def _render_agent_log_cards(
    web_questions,
    web_context,
    web_summary,
    section_plan,
    section_queries,
    abstract_report="",
    abstract_report_meta=None
):
    """Render the agent log cards (Step 1..4)."""
    if web_questions:
        web_html = '<div class="agent-log-card"><div class="agent-log-header"><span class="step-badge">STEP 1</span>Web Research</div><div class="agent-log-content">'
        web_html += '<strong style="color: #f0f6fc;">Target Wikipedia Pages:</strong><br>'
        for i, q in enumerate(web_questions, 1):
            web_html += f'<div style="margin: 4px 0; padding-left: 12px; border-left: 2px solid #30363d;">{i}. {q}</div>'
        web_html += '<br>'
        if web_context:
            if "Scout failed" in web_context:
                web_html += f'<div style="color: #d2a8ff; margin: 8px 0; font-style: italic;">{html.escape(web_context)}</div>'
            else:
                web_html += '<strong style="color: #f0f6fc;">Retrieved Context:</strong><br>'
                sources = web_context.split("-" * 40)
                sources_found = False
                for source_block in sources:
                    if "=== SOURCE:" not in source_block:
                        continue
                    sources_found = True
                    title_match = re.search(r'=== SOURCE: (.*?) ===', source_block)
                    url_match = re.search(r'URL: (.*?)\n', source_block)
                    title = title_match.group(1) if title_match else "Unknown Source"
                    url = url_match.group(1).strip() if url_match else "#"
                    title_esc = html.escape(title)
                    url_esc = html.escape(url)
                    web_html += f'''<div class="web-result-item"><div class="web-result-title"><a href="{url_esc}" target="_blank" style="color: #58a6ff; text-decoration: none;">{title_esc}</a></div></div>'''
                if not sources_found:
                    web_html += '<div style="color: #8b949e; font-style: italic;">Wikipedia content retrieved but parsing failed.</div>'
        web_html += '</div></div>'
        st.markdown(web_html, unsafe_allow_html=True)

    if web_summary:
        summary_body = _render_agent_log_markdown_html(web_summary)
        summary_html = f'''
        <div class="agent-log-card">
            <div class="agent-log-header"><span class="step-badge">STEP 1.5</span>Research Summary</div>
            <div class="agent-log-content">
                <div class="agent-log-section">{summary_body}</div>
            </div>
        </div>
        '''
        st.markdown(summary_html, unsafe_allow_html=True)

    if section_plan:
        plan_body = _render_agent_log_markdown_html(section_plan)
        plan_html = f'''
        <div class="agent-log-card">
            <div class="agent-log-header"><span class="step-badge">STEP 2</span>Section Plan</div>
            <div class="agent-log-content">
                <div class="agent-log-section">{plan_body}</div>
            </div>
        </div>
        '''
        st.markdown(plan_html, unsafe_allow_html=True)

    if section_queries:
        queries_html = '<div class="agent-log-card"><div class="agent-log-header"><span class="step-badge">STEP 3</span>Academic Search Queries</div><div class="agent-log-content">'
        for section_name, queries in section_queries.items():
            queries_html += f'<strong style="color: #f0f6fc;">{section_name}</strong><br>'
            for q in queries:
                queries_html += f'<span class="query-tag">{q}</span>'
            queries_html += '<br><br>'
        queries_html += '</div></div>'
        st.markdown(queries_html, unsafe_allow_html=True)

    if abstract_report:
        meta_line = ""
        if isinstance(abstract_report_meta, dict):
            used_count = int(abstract_report_meta.get("used_for_report", 0) or 0)
            relevant_total = int(abstract_report_meta.get("relevant_total", 0) or 0)
            limit_val = int(abstract_report_meta.get("limit", 0) or 0)
            if limit_val > 0:
                meta_line = (
                    f'<div style="font-size: 12px; color: #8b949e; margin-bottom: 8px;">'
                    f'Report based on {used_count} of {relevant_total} relevant papers (limit: {limit_val}).'
                    f'</div>'
                )
        report_text = str(abstract_report or "")
        # The card already has a title; remove duplicated top H1 if present.
        report_text = re.sub(
            r'^\s*#\s*Abstract[\s\-‑–—]*Based[\s\-‑–—]*Report\s*\n+',
            '',
            report_text,
            flags=re.IGNORECASE,
        )
        # Also remove a duplicated leading Executive Summary heading if present.
        report_text = re.sub(
            r'^\s*##\s*Executive\s+Summary\s*\n+',
            '',
            report_text,
            flags=re.IGNORECASE,
        )
        report_body = _render_agent_log_markdown_html(report_text)
        report_html = f'''
        <div class="agent-log-card">
            <div class="agent-log-header"><span class="step-badge">STEP 4</span>Abstract-Based Report</div>
            <div class="agent-log-content">
                {meta_line}
                <div class="agent-log-section">{report_body}</div>
            </div>
        </div>
        '''
        st.markdown(report_html, unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# File naming + tracking helpers (PDF/RIS mappings and download manifests).
# -----------------------------------------------------------------------------
def get_paper_filename(paper, idx=0):
    """Generate a safe filename for a paper PDF."""
    title = paper.get("title", f"paper_{idx}")
    safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')[:80]
    return f"{safe_title}.pdf"

def get_ris_filename(paper, idx=0):
    """Generate a safe filename for a paper RIS citation file."""
    title = paper.get("title", f"paper_{idx}")
    safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')[:80]
    return f"{safe_title}.ris"


def _get_pdf_ris_map_path(project_name: str) -> Path:
    return _get_project_dir(project_name) / "tracking" / "pdf_ris_map.json"


def _default_pdf_ris_map(project_name: str) -> dict:
    return {
        "schema_version": 1,
        "project": project_name,
        "updated_at": "",
        "mappings": []
    }


def _load_pdf_ris_map(project_name: str) -> dict:
    path = _get_pdf_ris_map_path(project_name)
    if not path.exists():
        return _default_pdf_ris_map(project_name)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            payload.setdefault("schema_version", 1)
            payload.setdefault("project", project_name)
            payload.setdefault("updated_at", "")
            payload.setdefault("mappings", [])
            return payload
    except Exception:
        pass
    return _default_pdf_ris_map(project_name)


def _persist_pdf_ris_map(project_name: str, payload: dict) -> None:
    path = _get_pdf_ris_map_path(project_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _upsert_pdf_ris_mapping(project_name: str, entry: dict) -> None:
    payload = _load_pdf_ris_map(project_name)
    mappings = payload.get("mappings", [])
    serial = entry.get("serial")
    updated = False
    for idx, existing in enumerate(mappings):
        if serial and existing.get("serial") == serial:
            merged = existing.copy()
            merged.update(entry)
            mappings[idx] = merged
            updated = True
            break
    if not updated:
        mappings.append(entry)
    payload["mappings"] = mappings
    payload["updated_at"] = datetime.utcnow().isoformat()
    _persist_pdf_ris_map(project_name, payload)


def _ensure_pdf_ris_db(project_name: str) -> sqlite3.Connection:
    # SQLite index for fast PDF/RIS lookups by serial and PDF id.
    db_path = _get_project_dir(project_name) / "db" / "file_directory.db"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute(
        "CREATE TABLE IF NOT EXISTS pdf_ris_links ("
        "serial TEXT PRIMARY KEY, "
        "pdf_id INTEGER, "
        "pdf_filename TEXT, "
        "pdf_rel_path TEXT, "
        "ris_rel_path TEXT, "
        "metadata_hash TEXT, "
        "created_at TEXT, "
        "updated_at TEXT"
        ");"
    )
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_pdf_ris_links_pdf_id "
        "ON pdf_ris_links(pdf_id)"
    )
    conn.commit()
    return conn


def _generate_serial(conn) -> str:
    for _ in range(5):
        serial = f"SR-{uuid.uuid4().hex[:10].upper()}"
        row = conn.execute(
            "SELECT 1 FROM pdf_ris_links WHERE serial = ? LIMIT 1",
            (serial,)
        ).fetchone()
        if not row:
            return serial
    return f"SR-{uuid.uuid4().hex.upper()}"


def _upsert_pdf_ris_link(conn, link: dict) -> None:
    serial = link.get("serial")
    if not serial:
        return
    existing = conn.execute(
        "SELECT 1 FROM pdf_ris_links WHERE serial = ? LIMIT 1",
        (serial,)
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE pdf_ris_links SET "
            "pdf_id = ?, pdf_filename = ?, pdf_rel_path = ?, "
            "ris_rel_path = ?, metadata_hash = ?, updated_at = ? "
            "WHERE serial = ?",
            (
                link.get("pdf_id"),
                link.get("pdf_filename"),
                link.get("pdf_rel_path"),
                link.get("ris_rel_path"),
                link.get("metadata_hash"),
                link.get("updated_at"),
                serial,
            )
        )
    else:
        conn.execute(
            "INSERT INTO pdf_ris_links (serial, pdf_id, pdf_filename, pdf_rel_path, "
            "ris_rel_path, metadata_hash, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                serial,
                link.get("pdf_id"),
                link.get("pdf_filename"),
                link.get("pdf_rel_path"),
                link.get("ris_rel_path"),
                link.get("metadata_hash"),
                link.get("created_at"),
                link.get("updated_at"),
            )
        )
    conn.commit()


def _compute_sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _resolve_unique_path(directory: Path, filename: str) -> Path:
    path = directory / filename
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for idx in range(1, 1000):
        candidate = directory / f"{stem}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
    return directory / f"{stem}_{uuid.uuid4().hex[:6]}{suffix}"


def _write_external_ris(project_name: str, filename: str, ris_text: str) -> Path:
    external_dir = _get_project_external_ris_dir(project_name)
    target = _resolve_unique_path(external_dir, filename)
    target.write_text(ris_text, encoding="utf-8")
    return target

# -----------------------------------------------------------------------------
# Download manifest (tracks Chromium/manual downloads and RIS status).
# -----------------------------------------------------------------------------
def _get_download_manifest_path(project_name: str) -> Path:
    return _get_project_dir(project_name) / "tracking" / "downloads.json"


def _default_download_manifest(project_name: str) -> dict:
    return {
        "schema_version": 1,
        "project": project_name,
        "updated_at": "",
        "downloads": []
    }


def _load_download_manifest(project_name: str) -> dict:
    path = _get_download_manifest_path(project_name)
    with _download_manifest_lock:
        if not path.exists():
            return _default_download_manifest(project_name)
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            payload = {}
        if not isinstance(payload, dict):
            payload = {}
        payload.setdefault("schema_version", 1)
        payload.setdefault("project", project_name)
        payload.setdefault("updated_at", "")
        payload.setdefault("downloads", [])
        return payload


def _save_download_manifest(project_name: str, payload: dict) -> None:
    path = _get_download_manifest_path(project_name)
    with _download_manifest_lock:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _upsert_download_entry(project_name: str, entry: dict) -> None:
    payload = _load_download_manifest(project_name)
    downloads = payload.get("downloads", [])
    download_id = entry.get("download_id")
    updated = False
    if download_id:
        for idx, existing in enumerate(downloads):
            if existing.get("download_id") == download_id:
                merged = existing.copy()
                merged.update(entry)
                downloads[idx] = merged
                updated = True
                break
    if not updated:
        downloads.append(entry)
    payload["downloads"] = downloads
    payload["updated_at"] = datetime.utcnow().isoformat()
    _save_download_manifest(project_name, payload)


def _update_download_entry(project_name: str, download_id: str, **updates) -> None:
    if not download_id:
        return
    payload = _load_download_manifest(project_name)
    downloads = payload.get("downloads", [])
    updated = False
    for idx, existing in enumerate(downloads):
        if existing.get("download_id") == download_id:
            merged = existing.copy()
            merged.update(updates)
            downloads[idx] = merged
            updated = True
            break
    if not updated:
        updates["download_id"] = download_id
        downloads.append(updates)
    payload["downloads"] = downloads
    payload["updated_at"] = datetime.utcnow().isoformat()
    _save_download_manifest(project_name, payload)


def _is_pdf_header(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False


def _is_file_stable(path: Path, stable_seconds: float = 1.5) -> bool:
    try:
        size_before = path.stat().st_size
        if size_before <= 0:
            return False
        time.sleep(stable_seconds)
        size_after = path.stat().st_size
        return size_before == size_after and size_after > 0
    except Exception:
        return False


def _wait_for_download_complete(download_dir: Path, expected_filename: str, timeout_seconds: int = 120) -> Path:
    expected_path = download_dir / expected_filename
    start = time.time()
    while time.time() - start < timeout_seconds:
        # Direct match
        if expected_path.exists() and _is_file_stable(expected_path) and _is_pdf_header(expected_path):
            return expected_path

        # Look for completed files
        candidates = []
        for item in download_dir.iterdir():
            if not item.is_file():
                continue
            if item.name.endswith(".crdownload"):
                continue
            candidates.append(item)

        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            for candidate in candidates:
                if not _is_file_stable(candidate):
                    continue
                if not _is_pdf_header(candidate):
                    continue
                if candidate.name != expected_filename:
                    target = _resolve_unique_path(download_dir, expected_filename)
                    try:
                        candidate.rename(target)
                        return target
                    except Exception:
                        return candidate
                return candidate

        time.sleep(0.5)
    return None


def _fetch_pdf_ris_link(conn: sqlite3.Connection, serial: str) -> dict:
    if not serial:
        return {}
    row = conn.execute(
        "SELECT pdf_id, pdf_filename, pdf_rel_path, ris_rel_path, metadata_hash, created_at "
        "FROM pdf_ris_links WHERE serial = ? LIMIT 1",
        (serial,)
    ).fetchone()
    if not row:
        return {}
    return {
        "pdf_id": row[0],
        "pdf_filename": row[1],
        "pdf_rel_path": row[2],
        "ris_rel_path": row[3],
        "metadata_hash": row[4],
        "created_at": row[5],
    }


def _record_pdf_link(project_name: str, serial: str, pdf_filename: str, pdf_path: Path) -> None:
    project_root = _get_project_dir(project_name)
    conn = _ensure_pdf_ris_db(project_name)
    try:
        existing = _fetch_pdf_ris_link(conn, serial)
        now = datetime.utcnow().isoformat()
        entry = {
            "serial": serial,
            "pdf_id": existing.get("pdf_id"),
            "pdf_filename": pdf_filename or existing.get("pdf_filename"),
            "pdf_rel_path": pdf_path.relative_to(project_root).as_posix(),
            "ris_rel_path": existing.get("ris_rel_path"),
            "metadata_hash": existing.get("metadata_hash"),
            "created_at": existing.get("created_at") or now,
            "updated_at": now,
        }
        _upsert_pdf_ris_link(conn, entry)
        _upsert_pdf_ris_mapping(project_name, entry)
    finally:
        conn.close()


def _save_ris_for_serial(project_name: str, serial: str, ris_filename: str, ris_data: str, pdf_filename: str = None) -> Path:
    ris_root = _get_project_ris_dir(project_name)
    ris_target = ris_root / serial / ris_filename
    ris_target.parent.mkdir(parents=True, exist_ok=True)
    ris_bytes = ris_data.encode("utf-8")
    ris_target.write_text(ris_data, encoding="utf-8")

    project_root = _get_project_dir(project_name)
    conn = _ensure_pdf_ris_db(project_name)
    try:
        existing = _fetch_pdf_ris_link(conn, serial)
        now = datetime.utcnow().isoformat()
        entry = {
            "serial": serial,
            "pdf_id": existing.get("pdf_id"),
            "pdf_filename": pdf_filename or existing.get("pdf_filename"),
            "pdf_rel_path": existing.get("pdf_rel_path"),
            "ris_rel_path": ris_target.relative_to(project_root).as_posix(),
            "metadata_hash": f"sha256:{_compute_sha256(ris_bytes)}",
            "created_at": existing.get("created_at") or now,
            "updated_at": now,
        }
        _upsert_pdf_ris_link(conn, entry)
        _upsert_pdf_ris_mapping(project_name, entry)
    finally:
        conn.close()
    return ris_target


def _monitor_chromium_download(project_name: str, download_id: str, download_dir: Path,
                               expected_filename: str, doi: str = None,
                               ris_filename: str = None, wait_for_ris: bool = False) -> None:
    try:
        _update_download_entry(project_name, download_id, pdf_status="in_progress")
        completed_path = _wait_for_download_complete(download_dir, expected_filename)
        if not completed_path:
            _update_download_entry(
                project_name,
                download_id,
                pdf_status="failed",
                error="Download timed out waiting for PDF."
            )
            if wait_for_ris:
                _update_download_entry(
                    project_name,
                    download_id,
                    ris_status="skipped",
                    error="PDF not detected; RIS not fetched."
                )
            return
        _record_pdf_link(project_name, download_id, expected_filename, completed_path)
        _update_download_entry(
            project_name,
            download_id,
            pdf_status="complete",
            pdf_rel_path=completed_path.relative_to(_get_project_dir(project_name)).as_posix(),
            completed_at=datetime.utcnow().isoformat(),
        )
        if wait_for_ris:
            if not doi:
                _update_download_entry(project_name, download_id, ris_status="skipped")
                return
            if not ris_filename:
                ris_filename = Path(expected_filename).with_suffix(".ris").name
            ris_data, error_msg = fetch_ris_citation(doi)
            if ris_data:
                _save_ris_for_serial(
                    project_name,
                    download_id,
                    ris_filename,
                    ris_data,
                    pdf_filename=expected_filename,
                )
                _update_download_entry(project_name, download_id, ris_status="complete")
            else:
                _update_download_entry(
                    project_name,
                    download_id,
                    ris_status="failed",
                    error=error_msg or "RIS fetch failed.",
                )
    except Exception as exc:
        _update_download_entry(
            project_name,
            download_id,
            pdf_status="failed",
            error=str(exc)
        )


def _start_chromium_download(paper: dict, idx: int, project_name: str, pdf_url: str,
                             chromium_path: str = None, profile_dir: str = None,
                             chromedriver_path: str = None, use_remote_debugging: bool = False,
                             remote_debug_port: int = None, reduce_automation: bool = False,
                             wait_for_pdf_for_ris: bool = False) -> dict:
    result = {
        "started": False,
        "download_id": None,
        "error": None,
        "ris_saved": False,
    }
    if not pdf_url:
        result["error"] = "No PDF URL available."
        return result

    _ensure_project_layout(project_name)
    pdf_root = _get_project_pdf_dir(project_name)
    conn = _ensure_pdf_ris_db(project_name)
    try:
        serial = _generate_serial(conn)
    finally:
        conn.close()

    download_dir = pdf_root / serial
    download_dir.mkdir(parents=True, exist_ok=True)
    expected_filename = get_paper_filename(paper, idx)

    now = datetime.utcnow().isoformat()
    entry = {
        "download_id": serial,
        "paper_id": paper.get("paper_id"),
        "doi": paper.get("doi"),
        "source_url": pdf_url,
        "download_dir": download_dir.relative_to(_get_project_dir(project_name)).as_posix(),
        "expected_filename": expected_filename,
        "pdf_status": "pending",
        "ris_status": "pending_pdf" if (wait_for_pdf_for_ris and paper.get("doi")) else ("pending" if paper.get("doi") else "skipped"),
        "started_at": now,
        "updated_at": now,
    }
    _upsert_download_entry(project_name, entry)

    opened = open_pdf_in_chrome(
        pdf_url,
        str(download_dir),
        paper_title=paper.get("title"),
        chromium_path=chromium_path,
        profile_dir=profile_dir,
        chromedriver_path=chromedriver_path,
        use_remote_debugging=use_remote_debugging,
        remote_debug_port=remote_debug_port,
        reduce_automation=reduce_automation,
    )
    if not opened:
        _update_download_entry(
            project_name,
            serial,
            pdf_status="failed",
            error="Failed to open Chromium session.",
        )
        result["error"] = "Failed to open Chromium session."
        return result

    result["started"] = True
    result["download_id"] = serial

    ris_filename = get_ris_filename(paper, idx)
    monitor_thread = threading.Thread(
        target=_monitor_chromium_download,
        args=(project_name, serial, download_dir, expected_filename, paper.get("doi"), ris_filename, wait_for_pdf_for_ris),
        daemon=True
    )
    monitor_thread.start()

    if wait_for_pdf_for_ris:
        return result

    doi = paper.get("doi")
    if doi:
        ris_data, error_msg = fetch_ris_citation(doi)
        if ris_data:
            _save_ris_for_serial(project_name, serial, ris_filename, ris_data, pdf_filename=expected_filename)
            _update_download_entry(project_name, serial, ris_status="complete")
            result["ris_saved"] = True
        else:
            _update_download_entry(
                project_name,
                serial,
                ris_status="failed",
                error=error_msg or "RIS fetch failed.",
            )
    return result

def fetch_ris_citation(doi):
    """Fetch RIS citation data for a paper using its DOI.
    
    Args:
        doi: The DOI of the paper (with or without https://doi.org/ prefix)
        
    Returns:
        tuple: (str: RIS formatted citation data or None, str: error message or None)
    """
    if not doi:
        return None, "No DOI provided"
    
    # Clean DOI if necessary
    clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "").strip()
    
    # Method 1: Try Crossref/habanero
    try:
        cr = Crossref()
        ris_data = cr.content_negotiation(ids=clean_doi, format="ris")
        if ris_data:
            return ris_data, None
    except Exception as e:
        pass  # Try fallback method
    
    # Method 2: Direct DOI content negotiation via requests
    try:
        headers = {"Accept": "application/x-research-info-systems"}
        response = requests.get(
            f"https://doi.org/{clean_doi}",
            headers=headers,
            timeout=10,
            allow_redirects=True
        )
        if response.status_code == 200 and response.text.strip().startswith("TY"):
            return response.text, None
        else:
            return None, f"DOI returned status {response.status_code}"
    except Exception as e:
        return None, f"Request failed: {str(e)}"

def _download_and_couple_pdf_ris(paper, idx, project_name, pdf_url):
    result = {
        "pdf_downloaded": False,
        "ris_saved": False,
        "serial": None,
        "error": None,
    }
    if not pdf_url:
        result["error"] = "No PDF URL available."
        return result

    _ensure_project_layout(project_name)
    project_root = _get_project_dir(project_name)
    pdf_root = _get_project_pdf_dir(project_name)
    ris_root = _get_project_ris_dir(project_name)
    pdf_filename = get_paper_filename(paper, idx)
    ris_filename = get_ris_filename(paper, idx)

    conn = _ensure_pdf_ris_db(project_name)
    try:
        serial = _generate_serial(conn)
        result["serial"] = serial

        pdf_target = pdf_root / serial / pdf_filename
        pdf_target.parent.mkdir(parents=True, exist_ok=True)
        try:
            with requests.get(pdf_url, timeout=30, stream=True) as response:
                if response.status_code != 200:
                    result["error"] = f"PDF download failed: {response.status_code}"
                    return result
                content_type = (response.headers.get("Content-Type") or "").lower()
                try:
                    content_iter = response.iter_content(chunk_size=8192)
                    first_chunk = next(content_iter, b"")
                except Exception as exc:
                    result["error"] = f"PDF download failed: {exc}"
                    return result
                if "pdf" not in content_type and not first_chunk.startswith(b"%PDF-"):
                    result["error"] = "PDF download failed: response was not a PDF."
                    return result
                with open(pdf_target, "wb") as f:
                    if first_chunk:
                        f.write(first_chunk)
                    for chunk in content_iter:
                        if chunk:
                            f.write(chunk)
        except Exception as exc:
            result["error"] = f"PDF download failed: {exc}"
            return result
        result["pdf_downloaded"] = True

        doi = paper.get("doi")
        if not doi:
            result["error"] = "Missing DOI for RIS fetch."
            return result

        ris_data, error_msg = fetch_ris_citation(doi)
        if not ris_data:
            result["error"] = error_msg or "RIS fetch failed."
            return result

        ris_target = ris_root / serial / ris_filename
        ris_target.parent.mkdir(parents=True, exist_ok=True)
        ris_bytes = ris_data.encode("utf-8")
        with open(ris_target, "w", encoding="utf-8") as f:
            f.write(ris_data)
        result["ris_saved"] = True

        metadata_hash = _compute_sha256(ris_bytes)
        now = datetime.utcnow().isoformat()
        entry = {
            "serial": serial,
            "pdf_id": None,
            "pdf_filename": pdf_filename,
            "pdf_rel_path": pdf_target.relative_to(project_root).as_posix(),
            "ris_rel_path": ris_target.relative_to(project_root).as_posix(),
            "metadata_hash": f"sha256:{metadata_hash}",
            "created_at": now,
            "updated_at": now,
        }
        _upsert_pdf_ris_link(conn, entry)
        _upsert_pdf_ris_mapping(project_name, entry)
        return result
    finally:
        conn.close()


def _find_project_pdf_path(project_name, filename):
    pdf_root = _get_project_pdf_dir(project_name)
    for path in pdf_root.rglob(filename):
        if path.is_file():
            return path
    return None


def get_ris_files(ris_dirs):
    """Return a set of RIS filenames that exist in the RIS directories."""
    if not isinstance(ris_dirs, (list, tuple, set)):
        ris_dirs = [ris_dirs]
    files = set()
    for ris_dir in ris_dirs:
        path = Path(ris_dir)
        if not path.exists():
            continue
        for item in path.rglob("*.ris"):
            if item.is_file():
                files.add(item.name)
    return files


def get_downloaded_papers(folder):
    """Return a set of PDF filenames that already exist in the target folder."""
    path = Path(folder)
    if not path.exists():
        return set()
    return {p.name for p in path.rglob("*.pdf") if p.is_file()}

def get_completed_downloads(project_name: str, folder: Path) -> set:
    """Return PDF filenames that are complete (manifest + filesystem)."""
    folder = Path(folder)
    completed = set()
    manifest = _load_download_manifest(project_name)
    for entry in manifest.get("downloads", []):
        status = entry.get("pdf_status") or entry.get("status")
        if status != "complete":
            continue
        expected = entry.get("expected_filename")
        if not expected:
            continue
        pdf_path = _find_project_pdf_path(project_name, expected)
        if pdf_path and pdf_path.exists() and _is_pdf_header(pdf_path):
            completed.add(expected)

    existing = get_downloaded_papers(folder)
    return completed | existing


def _has_pending_downloads(project_name: str) -> bool:
    manifest = _load_download_manifest(project_name)
    for entry in manifest.get("downloads", []):
        status = entry.get("pdf_status")
        if status in ("pending", "in_progress"):
            return True
    return False


def _enable_autorefresh(interval_ms: int = 4000, key: str = "downloads_autorefresh") -> bool:
    if hasattr(st, "autorefresh"):
        st.autorefresh(interval=interval_ms, key=key)
        return True
    try:
        from streamlit_autorefresh import st_autorefresh
    except Exception:
        return False
    st_autorefresh(interval=interval_ms, key=key)
    return True


def _set_force_autorefresh() -> None:
    st.session_state["force_autorefresh"] = True

def generate_excel_export(papers, export_path, topic="Research"):
    """Generate an Excel file with paper metadata.
    
    Args:
        papers: List of paper dictionaries
        export_path: Full path to save the Excel file
        topic: Research topic for the header
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    
    # Headers
    headers = [
        "Paper Title", "DOI", "Abstract", "Author Names", "Citation Count",
        "Year", "PDF URL", "Relevant", "Section", "Query"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, paper in enumerate(papers, 2):
        ws.cell(row=row_idx, column=1, value=paper.get("title", ""))
        ws.cell(row=row_idx, column=2, value=paper.get("doi", ""))
        ws.cell(row=row_idx, column=3, value=paper.get("abstract", ""))
        ws.cell(row=row_idx, column=4, value=paper.get("authors", ""))
        ws.cell(row=row_idx, column=5, value=paper.get("citations", 0))
        ws.cell(row=row_idx, column=6, value=paper.get("year", ""))
        ws.cell(row=row_idx, column=7, value=paper.get("pdf_url", ""))
        ws.cell(row=row_idx, column=8, value="Yes" if paper.get("is_relevant") else "No")
        ws.cell(row=row_idx, column=9, value=paper.get("section", ""))
        ws.cell(row=row_idx, column=10, value=paper.get("query", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50  # Title
    ws.column_dimensions['B'].width = 25  # DOI
    ws.column_dimensions['C'].width = 80  # Abstract
    ws.column_dimensions['D'].width = 40  # Authors
    ws.column_dimensions['E'].width = 12  # Citations
    ws.column_dimensions['F'].width = 8   # Year
    ws.column_dimensions['G'].width = 50  # PDF URL
    ws.column_dimensions['H'].width = 10  # Relevant
    ws.column_dimensions['I'].width = 30  # Section
    ws.column_dimensions['J'].width = 30  # Query
    
    wb.save(export_path)
    return True

def get_journal_quality(issn, email=None):
    if not issn: return None
    try:
        # Use provided email or fall back to global default (if any)
        mailto = email or OPENALEX_MAILTO
        params = {"mailto": mailto}
        r = requests.get(
            f"https://api.openalex.org/sources/issn:{issn}",
            params=params,
            timeout=2.0,
        )
        if r.status_code != 200: return None
        data = r.json()
        stats = data.get('summary_stats', {})
        return {"impact": stats.get('2yr_mean_citedness', 0), "name": data.get('display_name', 'Unknown Venue')}
    except: return None

from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED

def search_api(query_groups, limit, progress_callback=None, total_queries=None, timeout_seconds=120, timer_placeholder=None, s2_api_key=None, openalex_email=None, elsevier_api_key=None):
    """Search Semantic Scholar queries section-by-section using concurrent workers.
    
    Args:
        timeout_seconds: Max time in seconds for the entire search. Default 120 (2 min).
                        If exceeded, returns partial results.
        timer_placeholder: Optional Streamlit placeholder to show remaining time.
        s2_api_key: API Key for Semantic Scholar.
        openalex_email: Email for OpenAlex politeness.
        elsevier_api_key: API Key for Elsevier (fallback PDF source).
    """
    found, missing = [], []
    if not query_groups:
        flash("No queries generated.", level="warning")
        return found, missing

    # Normalize inputs into a flat list of (section, query) tasks.
    tasks = []
    if isinstance(query_groups, list):
        # Defensive fallback: treat a list as a single unnamed section.
        query_groups = {"section 1": query_groups}

    if isinstance(query_groups, dict):
        for section_name, section_queries in query_groups.items():
            if not section_queries:
                continue
            for q in section_queries:
                tasks.append((section_name, q))
                
    total = len(tasks)
    
    per_query_limit = max(1, int(limit))
    # User-provided API key (if any).
    sch = SemanticScholar(api_key=s2_api_key) 
    # SemanticScholar client is lightweight enough to share across threads.
    
    # Worker executes a single query and returns local found/missing lists.
    def _worker(task_tuple):
        section_name, query_text = task_tuple

        local_found = []
        local_missing = []
        
        try:
            # Semantic Scholar library may enforce rate limits.
            results_gen = sch.search_paper(query_text, limit=per_query_limit)
            attempts = 0
            # The SDK returns an iterator that fetches pages; iterate carefully.
            for paper in results_gen:
                attempts += 1
                if attempts > per_query_limit: 
                    break
                    
                try:
                    # Extract core fields with defensive defaults.
                    authors = ["Unknown"]
                    try:
                        if paper.authors:
                            authors = []
                            for a in paper.authors[:3]:
                                if hasattr(a, 'name'):
                                    authors.append(a.name)
                                elif isinstance(a, dict) and 'name' in a:
                                    authors.append(a['name'])
                            if len(paper.authors) > 3:
                                authors.append("et al.")
                            if not authors:
                                authors = ["Unknown"]
                    except:
                        authors = ["Unknown"]

                    tldr = "No AI summary available."
                    try:
                        if paper.tldr:
                            if isinstance(paper.tldr, dict) and 'text' in paper.tldr:
                                tldr = paper.tldr['text']
                            elif hasattr(paper.tldr, 'text'):
                                tldr = paper.tldr.text
                    except:
                        pass

                    # Extract DOI from externalIds or direct attribute.
                    doi = None
                    try:
                        if hasattr(paper, 'externalIds') and paper.externalIds:
                            doi = paper.externalIds.get('DOI')
                        elif hasattr(paper, 'doi') and paper.doi:
                            doi = paper.doi
                    except:
                        pass

                    p_data = {
                        "paper_id": str(uuid.uuid4()),  # Unique ID for feedback tracking
                        "title": paper.title or "Untitled",
                        "year": paper.year or "N/A",
                        "citations": paper.citationCount or 0,
                        "infl_citations": paper.influentialCitationCount or 0,
                        "authors": ", ".join(authors),
                        "tldr": tldr,
                        "abstract": paper.abstract or "No abstract.",
                        "pdf_url": paper.openAccessPdf.get('url') if (paper.openAccessPdf and isinstance(paper.openAccessPdf, dict)) else None,
                        "doi": doi,
                        "quality": None,
                        "query": query_text,
                        "section": section_name,
                        "pdf_source": None,  # Will be set below
                        "concepts": [],  # Will be set below from OpenAlex
                        "venue": None,  # Will be set below from OpenAlex
                    }

                    # Single OpenAlex call for concepts, venue, and impact.
                    if doi:
                        openalex_data = get_openalex_metadata(doi, email=openalex_email)
                        p_data['concepts'] = openalex_data.get('concepts', [])
                        p_data['venue'] = openalex_data.get('venue')
                        # Store impact in quality dict for backward compatibility with render_paper_card
                        impact = openalex_data.get('impact', 0)
                        if impact:
                            p_data['quality'] = {'impact': impact, 'name': openalex_data.get('venue')}

                    # Determine PDF source with fallback logic.
                    if p_data['pdf_url']:
                        # OpenAlex/Semantic Scholar has open access PDF
                        p_data['pdf_source'] = "OpenAlex"
                        local_found.append(p_data)
                    elif doi and elsevier_api_key:
                        # Attempt Elsevier API fallback
                        elsevier_result = fetch_elsevier_pdf_url(doi, elsevier_api_key)
                        if elsevier_result:
                            p_data['pdf_url'] = elsevier_result
                            p_data['pdf_source'] = "Elsevier"
                            local_found.append(p_data)
                        else:
                            p_data['pdf_source'] = "None"
                            local_missing.append(p_data)
                    else:
                        p_data['pdf_source'] = "None"
                        local_missing.append(p_data)
                except Exception:
                    continue
        except Exception:
            # Avoid Streamlit UI calls from worker threads.
            pass
            
        return local_found, local_missing

    completed = 0
    timed_out = False
    start_time = time.time()
    
    # Use limited concurrency to reduce rate-limit risk while improving throughput.
    executor = ThreadPoolExecutor(max_workers=5)
    try:
        future_to_task = {executor.submit(_worker, t): t for t in tasks}
        pending = set(future_to_task.keys())

        while pending:
            # Check global timeout
            elapsed = time.time() - start_time
            if elapsed > timeout_seconds:
                timed_out = True
                break

            # Wait briefly for completed tasks to keep UI responsive.
            done, pending = wait(pending, timeout=1.0, return_when=FIRST_COMPLETED)
            
            # Update timer display
            remaining = max(0, timeout_seconds - (time.time() - start_time))
            if timer_placeholder:
                timer_placeholder.markdown(f'''<div style="background: #21262d; padding: 8px 12px; border-radius: 4px; font-size: 11px; color: #8b949e;">Partial retrieval timeout: {int(remaining)}s remaining | Papers found: {len(found)} | Queries done: {completed}/{total}</div>''', unsafe_allow_html=True)

            # Process completed tasks
            for future in done:
                completed += 1
                if progress_callback:
                    progress_callback(completed, total)
                
                try:
                    l_found, l_missing = future.result() 
                    found.extend(l_found)
                    missing.extend(l_missing)
                except Exception as e:
                    # Task level error or timeout
                    pass
    finally:
        # Ensure we don't wait for stuck tasks
        # Try cancel_futures=True (Python 3.9+)
        try:
            executor.shutdown(wait=False, cancel_futures=True)
        except TypeError:
            # Fallback for older python
            executor.shutdown(wait=False)
    
    if timer_placeholder:
        timer_placeholder.empty()
    
    if timed_out:
        flash(
            f"Search timed out after {timeout_seconds}s. Proceeding with {len(found)} papers retrieved.",
            level="warning"
        )

    return found, missing

# -----------------------------------------------------------------------------
# Identifier normalization + direct lookups (DOI / arXiv).
# -----------------------------------------------------------------------------
DOI_REGEX = re.compile(r"^10\.\d{4,9}/\S+$", re.IGNORECASE)
ARXIV_NEW_REGEX = re.compile(r"^(\d{4}\.\d{4,5})(v\d+)?$", re.IGNORECASE)
ARXIV_OLD_REGEX = re.compile(r"^([a-z-]+(\.[a-z]{2})?/\d{7})(v\d+)?$", re.IGNORECASE)

def _clean_doi_input(raw_value):
    if not raw_value:
        return None
    doi = str(raw_value).strip()
    if not doi:
        return None
    doi = re.sub(r"^https?://(dx\.)?doi\.org/", "", doi, flags=re.IGNORECASE)
    doi = re.sub(r"^doi:\s*", "", doi, flags=re.IGNORECASE)
    doi = doi.strip().rstrip(".,;")
    if not doi:
        return None
    if not DOI_REGEX.match(doi):
        return None
    return doi.lower()

def _clean_arxiv_input(raw_value):
    if not raw_value:
        return None
    value = str(raw_value).strip()
    if not value:
        return None
    value = re.sub(r"^https?://(www\.)?arxiv\.org/(abs|pdf)/", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\.pdf$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^arxiv:\s*", "", value, flags=re.IGNORECASE)
    value = value.strip()
    if not value:
        return None
    match = ARXIV_NEW_REGEX.match(value)
    if match:
        return match.group(1).lower()
    match = ARXIV_OLD_REGEX.match(value)
    if match:
        return match.group(1).lower()
    return None

def _extract_arxiv_id_from_doi(doi_norm):
    if not doi_norm:
        return None
    match = re.match(r"^10\.48550/arxiv\.(.+)$", doi_norm, flags=re.IGNORECASE)
    if not match:
        return None
    return _clean_arxiv_input(match.group(1))

def _paper_arxiv_matches(paper, arxiv_norm):
    if not arxiv_norm or not isinstance(paper, dict):
        return False
    paper_id = _clean_arxiv_input(paper.get("arxiv_id"))
    return paper_id == arxiv_norm

def _parse_manual_identifier(raw_value):
    cleaned_doi = _clean_doi_input(raw_value)
    if cleaned_doi:
        arxiv_from_doi = _extract_arxiv_id_from_doi(cleaned_doi)
        if arxiv_from_doi:
            return "arxiv", arxiv_from_doi
        return "doi", cleaned_doi
    cleaned_arxiv = _clean_arxiv_input(raw_value)
    if cleaned_arxiv:
        return "arxiv", cleaned_arxiv
    return None, None

def _paper_doi_matches(paper, doi_norm):
    if not doi_norm or not isinstance(paper, dict):
        return False
    paper_doi = _clean_doi_input(paper.get("doi"))
    return paper_doi == doi_norm

def fetch_paper_by_doi(doi, s2_api_key=None, openalex_email=None, elsevier_api_key=None):
    if not doi:
        return None, "No DOI provided."
    clean_doi = _clean_doi_input(doi)
    if not clean_doi:
        return None, "Invalid DOI."
    arxiv_id = _extract_arxiv_id_from_doi(clean_doi)
    if arxiv_id:
        return fetch_paper_by_arxiv_id(arxiv_id, openalex_email=openalex_email)

    sch = SemanticScholar(api_key=s2_api_key)
    fields = [
        "title",
        "year",
        "citationCount",
        "influentialCitationCount",
        "authors",
        "abstract",
        "tldr",
        "openAccessPdf",
        "externalIds",
        "venue",
        "publicationVenue",
        "isOpenAccess",
    ]
    try:
        paper = sch.get_paper(clean_doi, fields=fields)
    except Exception as exc:
        return None, f"Semantic Scholar lookup failed: {exc}"

    authors = ["Unknown"]
    try:
        if paper.authors:
            authors = []
            for a in paper.authors[:3]:
                if hasattr(a, "name"):
                    authors.append(a.name)
                elif isinstance(a, dict) and "name" in a:
                    authors.append(a["name"])
            if len(paper.authors) > 3:
                authors.append("et al.")
            if not authors:
                authors = ["Unknown"]
    except Exception:
        authors = ["Unknown"]

    tldr = "No AI summary available."
    try:
        if paper.tldr:
            if isinstance(paper.tldr, dict) and "text" in paper.tldr:
                tldr = paper.tldr["text"]
            elif hasattr(paper.tldr, "text"):
                tldr = paper.tldr.text
    except Exception:
        pass

    pdf_url = None
    open_access_pdf = getattr(paper, "openAccessPdf", None)
    if isinstance(open_access_pdf, dict):
        pdf_url = open_access_pdf.get("url")
    elif hasattr(open_access_pdf, "get"):
        pdf_url = open_access_pdf.get("url")
    elif hasattr(open_access_pdf, "url"):
        pdf_url = open_access_pdf.url

    doi_value = None
    try:
        if paper.externalIds:
            doi_value = paper.externalIds.get("DOI") or paper.externalIds.get("doi")
    except Exception:
        pass
    if not doi_value:
        doi_value = clean_doi

    venue_value = None
    try:
        if getattr(paper, "venue", None):
            venue_value = paper.venue
        elif getattr(paper, "publicationVenue", None):
            pub = paper.publicationVenue
            if isinstance(pub, dict):
                venue_value = pub.get("name") or pub.get("displayName")
            elif hasattr(pub, "name"):
                venue_value = pub.name
    except Exception:
        pass

    p_data = {
        "paper_id": str(uuid.uuid4()),
        "title": paper.title or "Untitled",
        "year": paper.year or "N/A",
        "citations": paper.citationCount or 0,
        "infl_citations": paper.influentialCitationCount or 0,
        "authors": ", ".join(authors),
        "tldr": tldr,
        "abstract": paper.abstract or "No abstract.",
        "pdf_url": pdf_url,
        "doi": doi_value,
        "quality": None,
        "query": f"doi:{clean_doi}",
        "section": "Manual DOI",
        "pdf_source": None,
        "concepts": [],
        "venue": venue_value,
    }

    if doi_value:
        openalex_data = get_openalex_metadata(doi_value, email=openalex_email)
        p_data["concepts"] = openalex_data.get("concepts", [])
        if openalex_data.get("venue"):
            p_data["venue"] = openalex_data.get("venue")
        impact = openalex_data.get("impact", 0)
        if impact:
            p_data["quality"] = {"impact": impact, "name": openalex_data.get("venue") or venue_value}

    if p_data["pdf_url"]:
        p_data["pdf_source"] = "OpenAlex"
    elif doi_value and elsevier_api_key:
        elsevier_result = fetch_elsevier_pdf_url(doi_value, elsevier_api_key)
        if elsevier_result:
            p_data["pdf_url"] = elsevier_result
            p_data["pdf_source"] = "Elsevier"
        else:
            p_data["pdf_source"] = "None"
    else:
        p_data["pdf_source"] = "None"

    return p_data, None

def fetch_paper_by_arxiv_id(arxiv_id, openalex_email=None):
    if not arxiv_id:
        return None, "No arXiv ID provided."
    clean_id = _clean_arxiv_input(arxiv_id)
    if not clean_id:
        return None, "Invalid arXiv ID."

    params = {"search_query": f"id:{clean_id}", "start": 0, "max_results": 1}
    headers = {}
    if openalex_email:
        headers["User-Agent"] = f"ResearchAssistant/1.0 ({openalex_email})"

    try:
        resp = requests.get(
            "http://export.arxiv.org/api/query",
            params=params,
            headers=headers,
            timeout=10.0,
        )
    except Exception as exc:
        return None, f"arXiv lookup failed: {exc}"

    if resp.status_code != 200:
        return None, f"arXiv lookup failed: HTTP {resp.status_code}"

    try:
        root = ET.fromstring(resp.text)
    except Exception as exc:
        return None, f"arXiv response parsing failed: {exc}"

    ns = {"atom": "http://www.w3.org/2005/Atom", "arxiv": "http://arxiv.org/schemas/atom"}
    entry = root.find("atom:entry", ns)
    if entry is None:
        return None, "No arXiv record found."

    def _entry_findtext(tag, default=""):
        value = entry.findtext(tag, default=None, namespaces=ns)
        if value is None:
            tag_fallback = tag.split(":", 1)[-1]
            value = entry.findtext(tag_fallback, default=None)
        return default if value is None else value

    title = (_entry_findtext("atom:title", default="") or "").strip()
    title = re.sub(r"\s+", " ", title) if title else "Untitled"
    summary = (_entry_findtext("atom:summary", default="") or "").strip()
    summary = re.sub(r"\s+", " ", summary) if summary else "No abstract."
    published = (_entry_findtext("atom:published", default="") or "").strip()
    year = published[:4] if published else "N/A"

    authors = []
    author_nodes = entry.findall("atom:author", ns)
    if not author_nodes:
        author_nodes = entry.findall("author")
    for author in author_nodes:
        name = (author.findtext("atom:name", default=None, namespaces=ns) or author.findtext("name", default="") or "").strip()
        if name:
            authors.append(name)
    if not authors:
        authors = ["Unknown"]

    doi_value = (_entry_findtext("arxiv:doi", default="") or "").strip()
    journal_ref = (_entry_findtext("arxiv:journal_ref", default="") or "").strip()

    pdf_url = None
    for link in entry.findall("atom:link", ns):
        if link.get("title") == "pdf" or link.get("type") == "application/pdf":
            pdf_url = link.get("href")
            break
    if not pdf_url:
        pdf_url = f"https://arxiv.org/pdf/{clean_id}.pdf"

    p_data = {
        "paper_id": str(uuid.uuid4()),
        "title": title,
        "year": year,
        "citations": 0,
        "infl_citations": 0,
        "authors": ", ".join(authors),
        "tldr": "No AI summary available.",
        "abstract": summary or "No abstract.",
        "pdf_url": pdf_url,
        "doi": doi_value or None,
        "quality": None,
        "query": f"arxiv:{clean_id}",
        "section": "Manual arXiv",
        "pdf_source": "arXiv",
        "concepts": [],
        "venue": journal_ref or "arXiv",
        "arxiv_id": clean_id,
    }

    if doi_value:
        openalex_data = get_openalex_metadata(doi_value, email=openalex_email)
        p_data["concepts"] = openalex_data.get("concepts", [])
        if openalex_data.get("venue"):
            p_data["venue"] = openalex_data.get("venue")
        impact = openalex_data.get("impact", 0)
        if impact:
            p_data["quality"] = {"impact": impact, "name": openalex_data.get("venue") or journal_ref or "arXiv"}

    return p_data, None

def _handle_manual_fetch(
    raw_input,
    session_data,
    manual_papers,
    found,
    missing,
    project_name,
    session_id_export,
    topic_label,
    primary_model,
    provider,
    base_url,
    api_key,
    show_debug,
    s2_api_key,
    openalex_email,
    elsevier_api_key,
):
    raw_input = (raw_input or "").strip()
    token_count = len(raw_input.split()) if raw_input else 0
    if not raw_input:
        flash("Enter a DOI or arXiv ID to fetch.", level="warning")
        return
    if token_count > 1 or "," in raw_input or ";" in raw_input:
        flash("Enter one DOI or arXiv ID at a time.", level="warning")
        return

    id_kind, cleaned = _parse_manual_identifier(raw_input)
    if not cleaned:
        flash("Enter a valid DOI or arXiv ID.", level="warning")
        return

    if id_kind == "doi":
        already_manual = any(_paper_doi_matches(p, cleaned) for p in manual_papers)
        already_session = any(_paper_doi_matches(p, cleaned) for p in (found + missing))
    else:
        already_manual = any(_paper_arxiv_matches(p, cleaned) for p in manual_papers)
        already_session = any(_paper_arxiv_matches(p, cleaned) for p in (found + missing))

    if already_manual:
        flash("That identifier is already in Manual Search Results.", level="info")
        return
    if already_session:
        flash("That identifier already exists in the session results.", level="info")
        return

    with st.spinner("Fetching identifier..."):
        if id_kind == "doi":
            paper, error = fetch_paper_by_doi(
                cleaned,
                s2_api_key=s2_api_key,
                openalex_email=openalex_email,
                elsevier_api_key=elsevier_api_key,
            )
        else:
            paper, error = fetch_paper_by_arxiv_id(
                cleaned,
                openalex_email=openalex_email,
            )
    if error:
        flash(error, level="error")
        return
    if not paper:
        flash("No paper data returned.", level="warning")
        return

    try:
        if topic_label:
            paper = score_relevance(
                topic_label,
                [paper],
                model_name=primary_model,
                provider=provider,
                base_url=base_url,
                api_key=api_key,
                show_debug=show_debug,
            )[0]
    except Exception:
        flash("Relevance scoring failed; added without relevance tag.", level="warning")

    manual_papers.append(paper)
    session_data["manual_papers"] = manual_papers
    st.session_state["active_session"] = session_data
    save_session(project_name, session_id_export, session_data)
    st.session_state["manual_doi_input_reset"] = True
    st.rerun()

def flatten_string_queries(data):
    """Flatten nested query collections down to a simple list of strings."""
    flattened = []
    if isinstance(data, str):
        return [data]
    if isinstance(data, (list, tuple)):
        for item in data:
            flattened.extend(flatten_string_queries(item))
    return [q for q in flattened if isinstance(q, str)]

def _normalize_doi(doi):
    if not doi:
        return None
    cleaned = str(doi).strip()
    cleaned = cleaned.replace("https://doi.org/", "").replace("http://doi.org/", "")
    cleaned = cleaned.strip().lower()
    return cleaned or None

def _normalize_title(title):
    if not title:
        return None
    text = str(title).lower()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None

def _normalize_author(author):
    if not author:
        return None
    text = str(author).lower()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None

def _extract_first_author(authors):
    if not authors:
        return None
    if isinstance(authors, str):
        first = authors.split(",")[0]
        return _normalize_author(first)
    if isinstance(authors, (list, tuple)) and authors:
        first = authors[0]
        if isinstance(first, dict):
            first = first.get("name") or first.get("full_name") or ""
        return _normalize_author(first)
    return _normalize_author(authors)

def _normalize_year(year):
    if not year:
        return None
    match = re.search(r"\b(19|20)\d{2}\b", str(year))
    return match.group(0) if match else None

def _normalize_pdf_url(url):
    if not url:
        return None
    try:
        parts = urlsplit(str(url).strip())
        if not parts.scheme or not parts.netloc:
            return None
        netloc = parts.netloc.lower()
        path = parts.path.rstrip("/")
        return f"{parts.scheme.lower()}://{netloc}{path}"
    except Exception:
        return None

def _collect_external_ids(paper):
    ids = []
    for key in ("externalIds", "external_ids"):
        ext = paper.get(key)
        if isinstance(ext, dict):
            for val in ext.values():
                if val:
                    ids.append(str(val).strip().lower())
    for key in ("s2_paper_id", "paperId", "openalex_id", "openalexId"):
        val = paper.get(key)
        if val:
            ids.append(str(val).strip().lower())
    return list(dict.fromkeys(ids))

def _has_real_text(text, placeholders):
    if not text:
        return False
    if not isinstance(text, str):
        text = str(text)
    cleaned = text.strip()
    if not cleaned:
        return False
    return cleaned not in placeholders

def _score_canonical(paper):
    has_doi = bool(_normalize_doi(paper.get("doi")))
    has_pdf = bool(paper.get("pdf_url"))
    has_abstract = _has_real_text(paper.get("abstract"), {"No abstract."})
    has_tldr = _has_real_text(paper.get("tldr"), {"No AI summary available."})
    citations = paper.get("citations") or 0
    try:
        citations = int(citations)
    except Exception:
        citations = 0
    abstract_len = len(paper.get("abstract") or "")
    title_len = len(paper.get("title") or "")
    return (
        int(has_doi),
        int(has_pdf),
        int(has_abstract),
        int(has_tldr),
        citations,
        abstract_len,
        title_len,
    )

def _pdf_source_rank(paper):
    if not paper.get("pdf_url"):
        return 0
    source = paper.get("pdf_source") or ""
    if source == "OpenAlex":
        return 3
    if source == "Elsevier":
        return 2
    return 1

def _merge_paper_group(group):
    best = max(group["papers"], key=_score_canonical)
    canonical = best.copy()

    all_queries = []
    all_sections = []
    dup_ids = []
    doi_norms = set()

    for p in group["papers"]:
        doi_norm = _normalize_doi(p.get("doi"))
        if doi_norm:
            doi_norms.add(doi_norm)
        q = p.get("query")
        if q:
            all_queries.append(q)
        s = p.get("section")
        if s:
            all_sections.append(s)
        pid = p.get("paper_id")
        if pid and pid != canonical.get("paper_id"):
            dup_ids.append(pid)

    if doi_norms:
        canonical["doi"] = sorted(doi_norms)[0]

    best_pdf = canonical.get("pdf_url")
    best_pdf_source = canonical.get("pdf_source")
    best_pdf_rank = _pdf_source_rank(canonical)
    for p in group["papers"]:
        rank = _pdf_source_rank(p)
        if rank > best_pdf_rank:
            best_pdf_rank = rank
            best_pdf = p.get("pdf_url")
            best_pdf_source = p.get("pdf_source")
    if best_pdf:
        canonical["pdf_url"] = best_pdf
    if best_pdf_source:
        canonical["pdf_source"] = best_pdf_source

    for p in group["papers"]:
        title = p.get("title")
        if title and len(str(title)) > len(str(canonical.get("title") or "")):
            canonical["title"] = title

    for p in group["papers"]:
        authors = p.get("authors")
        if authors and len(str(authors)) > len(str(canonical.get("authors") or "")):
            canonical["authors"] = authors

    year_norm = _normalize_year(canonical.get("year"))
    for p in group["papers"]:
        cand_year = _normalize_year(p.get("year"))
        if cand_year and (not year_norm or int(cand_year) < int(year_norm)):
            year_norm = cand_year
    if year_norm:
        canonical["year"] = year_norm

    def _as_int(val):
        try:
            return int(val)
        except Exception:
            return 0

    canonical["citations"] = max(
        _as_int(canonical.get("citations")),
        *[_as_int(p.get("citations")) for p in group["papers"]],
    )
    canonical["infl_citations"] = max(
        _as_int(canonical.get("infl_citations")),
        *[_as_int(p.get("infl_citations")) for p in group["papers"]],
    )

    placeholder_abstracts = {"No abstract."}
    if not _has_real_text(canonical.get("abstract"), placeholder_abstracts):
        for p in group["papers"]:
            if _has_real_text(p.get("abstract"), placeholder_abstracts):
                canonical["abstract"] = p.get("abstract")
                break

    placeholder_tldr = {"No AI summary available."}
    if not _has_real_text(canonical.get("tldr"), placeholder_tldr):
        for p in group["papers"]:
            if _has_real_text(p.get("tldr"), placeholder_tldr):
                canonical["tldr"] = p.get("tldr")
                break

    concepts = []
    seen = set()
    for p in group["papers"]:
        for c in p.get("concepts") or []:
            if c in seen:
                continue
            seen.add(c)
            concepts.append(c)
    if concepts:
        canonical["concepts"] = concepts[:3]

    if not canonical.get("venue"):
        for p in group["papers"]:
            if p.get("venue"):
                canonical["venue"] = p.get("venue")
                break

    if not (canonical.get("quality") or {}).get("impact"):
        for p in group["papers"]:
            quality = p.get("quality") or {}
            if quality.get("impact"):
                canonical["quality"] = quality
                break

    if all_queries:
        canonical["all_queries"] = list(dict.fromkeys(all_queries))
    if all_sections:
        canonical["all_sections"] = list(dict.fromkeys(all_sections))
    if dup_ids:
        canonical["duplicate_paper_ids"] = list(dict.fromkeys(dup_ids))

    return canonical

# -----------------------------------------------------------------------------
# Deduplication and sorting utilities.
# -----------------------------------------------------------------------------
def deduplicate_papers(papers, fuzzy_threshold=0.95):
    if not papers:
        return [], {"input": 0, "groups": 0, "output": 0, "fuzzy_warnings": []}

    groups = []
    doi_map = {}
    ext_map = {}
    tya_map = {}
    title_map = {}
    pdf_map = {}
    fuzzy_warnings = []

    def _doi_conflict(group, doi_norm):
        if not doi_norm:
            return False
        existing = group.get("doi_norms") or set()
        return bool(existing and doi_norm not in existing)

    def _find_fuzzy_title_group(title_norm, doi_norm):
        best_idx = None
        best_ratio = 0.0
        for idx, group in enumerate(groups):
            existing_title = group.get("title_norm")
            if not existing_title:
                continue
            if _doi_conflict(group, doi_norm):
                continue
            ratio = SequenceMatcher(None, title_norm, existing_title).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_idx = idx
        if best_ratio >= fuzzy_threshold:
            if best_ratio < fuzzy_threshold + 0.02:
                fuzzy_warnings.append(
                    f"title similarity {best_ratio:.2f} for '{title_norm[:80]}'"
                )
            return best_idx
        return None

    def _add_to_group(group_idx, paper, keys):
        group = groups[group_idx]
        group["papers"].append(paper)
        doi_norm = keys.get("doi")
        if doi_norm:
            group["doi_norms"].add(doi_norm)
        title_norm = keys.get("title")
        if title_norm and not group.get("title_norm"):
            group["title_norm"] = title_norm

        if doi_norm:
            doi_map.setdefault(doi_norm, group_idx)
        for ext in keys.get("ext_ids", []):
            ext_map.setdefault(ext, group_idx)
        tya = keys.get("tya")
        if tya:
            tya_map.setdefault(tya, group_idx)
        if title_norm:
            title_map.setdefault(title_norm, group_idx)
        pdf_key = keys.get("pdf")
        if pdf_key:
            pdf_map.setdefault(pdf_key, group_idx)

    for paper in papers:
        doi_norm = _normalize_doi(paper.get("doi"))
        title_norm = _normalize_title(paper.get("title"))
        first_author = _extract_first_author(paper.get("authors"))
        year_norm = _normalize_year(paper.get("year"))
        tya_key = None
        if title_norm and year_norm and first_author:
            tya_key = f"{title_norm}|{year_norm}|{first_author}"
        pdf_key = _normalize_pdf_url(paper.get("pdf_url"))
        ext_ids = _collect_external_ids(paper)

        keys = {
            "doi": doi_norm,
            "title": title_norm,
            "tya": tya_key,
            "pdf": pdf_key,
            "ext_ids": ext_ids,
        }

        group_idx = None
        if doi_norm and doi_norm in doi_map:
            group_idx = doi_map[doi_norm]
        if group_idx is None:
            for ext in ext_ids:
                if ext in ext_map:
                    group_idx = ext_map[ext]
                    break
        if group_idx is None and tya_key and tya_key in tya_map:
            group_idx = tya_map[tya_key]
        if group_idx is None and title_norm:
            if title_norm in title_map:
                candidate_idx = title_map[title_norm]
                if not _doi_conflict(groups[candidate_idx], doi_norm):
                    group_idx = candidate_idx
            if group_idx is None:
                group_idx = _find_fuzzy_title_group(title_norm, doi_norm)
        if group_idx is None and pdf_key and pdf_key in pdf_map:
            candidate_idx = pdf_map[pdf_key]
            if not _doi_conflict(groups[candidate_idx], doi_norm):
                group_idx = candidate_idx

        if group_idx is None:
            group_idx = len(groups)
            groups.append({
                "papers": [],
                "title_norm": title_norm,
                "doi_norms": set(),
            })
        _add_to_group(group_idx, paper, keys)

    merged = []
    for group in groups:
        merged.append(_merge_paper_group(group))

    stats = {
        "input": len(papers),
        "groups": len(groups),
        "output": len(merged),
        "fuzzy_warnings": fuzzy_warnings,
    }
    return merged, stats

def sort_papers_by_relevance(papers):
    """Sort papers by relevance flag first, then citation count descending."""
    return sorted(
        papers,
        key=lambda p: (
            not p.get("is_relevant", False),
            -(p.get("citations") or 0),
        ),
    )

def sort_papers_with_feedback(papers, feedback_dict):
    """Sort papers by user feedback first, then relevance and citations.
    
    Thumbs up (+1) brings paper up, thumbs down (-1) pushes it down.
    """
    def sort_key(p):
        paper_id = p.get('paper_id', p.get('title', ''))  # Use UUID if available, fallback to title
        vote = feedback_dict.get(paper_id, 0)  # 0 = no vote, 1 = up, -1 = down
        return (
            -vote,  # Higher votes first (up=1 -> -1, down=-1 -> 1)
            not p.get("is_relevant", False),  # Relevant first
            -(p.get("citations") or 0),  # Higher citations first
        )
    return sorted(papers, key=sort_key)

# =============================================================================
# MAIN UI LAYOUT + INTERACTION FLOW
# =============================================================================
_flash_autorefresh()
st.title("Research Assistant")
_render_flash_messages(scope="main")

# Sidebar configuration
with st.sidebar:
    _render_flash_messages(scope="sidebar")
    invalid_projects = _consume_invalid_project_names()
    for invalid_name in invalid_projects:
        flash(f"Skipped invalid project name '{invalid_name}'.", level="warning", scope="sidebar")
    st.header("Project & Sessions")
    _render_section_help(
        "writer_projects_sessions",
        "Pick or create a project, then manage saved research sessions. Sessions store the agent outputs for a topic."
    )
    
    # Project management
    projects_data = init_projects()
    project_names = sorted(projects_data.keys())

    selected_project = ""
    if not project_names:
        st.caption("No projects yet. Create one to start.")
        st.session_state["current_project"] = ""
    else:
        if st.session_state.get("current_project") not in project_names:
            if "Thesis" in project_names:
                st.session_state["current_project"] = "Thesis"
            else:
                st.session_state["current_project"] = project_names[0]

        selected_index = project_names.index(st.session_state["current_project"])
        selected_project = st.selectbox(
            "Current Project",
            project_names,
            index=selected_index,
            key="project_selector",
        )

        if selected_project != st.session_state["current_project"]:
            st.session_state["current_project"] = selected_project
            st.rerun()

    # Manage projects
    st.markdown("#### Manage Projects")
    _render_section_help(
        "writer_manage_projects",
        "Create or remove projects. Projects group sessions and exported files."
    )
    new_proj_name = st.text_input("New Project Name", placeholder="e.g. Chapter 1")
    if st.button("Create Project", use_container_width=True):
        if new_proj_name:
            safe_name = _validate_project_name(new_proj_name)
            if not safe_name:
                flash("Invalid project name. Avoid path separators or '..'.", level="warning", scope="sidebar")
            elif create_project(safe_name, projects_data):
                flash(f"Created '{safe_name}'", level="success", scope="sidebar")
                st.session_state["current_project"] = safe_name
                st.rerun()
            else:
                flash("Project already exists", level="error", scope="sidebar")
    
    st.markdown("---")
    if selected_project:
        if st.button(f"Delete '{selected_project}'", type="primary", use_container_width=True):
            if delete_project(selected_project, projects_data):
                flash(f"Deleted project '{selected_project}'", level="warning", scope="sidebar")
                st.session_state.pop("current_project", None)
                st.rerun()

    # Session list (filtered)
    st.subheader(f"Sessions in {selected_project}")
    _render_section_help(
        "writer_sessions_list",
        "Load recent sessions to review results, or bulk delete old sessions."
    )

    sessions = get_project_sessions(selected_project, projects_data) if selected_project else []
    if sessions:
        sessions_to_delete = []
        for s in sessions[:10]:  # Show the most recent 10
            col1, col2 = st.columns([1, 5])
            with col1:
                checkbox_label = f"Select session {s['id']} for deletion"
                if st.checkbox(checkbox_label, key=f"del_{s['id']}", label_visibility="collapsed"):
                    sessions_to_delete.append(s["id"])
            with col2:
                label = f"#{s['id']}: {s['topic'][:25]}..."
                if st.button(label, key=f"load_{s['id']}", use_container_width=True):
                    st.session_state["load_session_id"] = s["id"]
                    st.session_state["load_session_project"] = selected_project
                    st.rerun()
                    
        if sessions_to_delete:
            if st.button("Delete Selected", type="primary", use_container_width=True):
                current_active_id = st.session_state.get("current_session_id")
                needs_clear = False
                for sid in sessions_to_delete:
                    delete_session(selected_project, sid)
                    if str(sid) == str(current_active_id):
                        needs_clear = True
                
                # Persist updated session list for this project.
                if selected_project in projects_data:
                    current_ids = projects_data[selected_project]
                    new_ids = [x for x in current_ids if x not in sessions_to_delete]
                    projects_data[selected_project] = new_ids
                    save_projects(projects_data)

                if needs_clear:
                    st.session_state.pop("active_session", None)
                    st.session_state.pop("current_session_id", None)
                    st.session_state.pop("load_session_id", None)
                st.rerun()
    else:
        st.caption("No sessions in this project.")

    st.divider()

    # Settings and connections
    st.header("Settings & Connections")
    _render_section_help(
        "writer_settings_connections",
        "Configure API keys for data sources and select the reasoning model."
    )
    
    # API configuration
    api_config = load_api_config()
    _render_section_help(
        "writer_api_config",
        "OpenAlex and Semantic Scholar improve metadata. Elsevier enables direct PDF access when available."
    )
    
    # Initialize config defaults.
    if "openalex_email" not in st.session_state:
        st.session_state["openalex_email"] = api_config.get("openalex_email", "")
    if "s2_api_key" not in st.session_state:
        st.session_state["s2_api_key"] = api_config.get("s2_api_key", "")
    if "elsevier_api_key" not in st.session_state:
        st.session_state["elsevier_api_key"] = api_config.get("elsevier_api_key", "")
    if "glm_api_key" not in st.session_state:
        st.session_state["glm_api_key"] = api_config.get("glm_api_key", "")
    if "glm_base_url" not in st.session_state:
        st.session_state["glm_base_url"] = api_config.get("glm_base_url", GLM_BASE_URL)
    if "llm_provider" not in st.session_state:
        st.session_state["llm_provider"] = api_config.get("llm_provider", "ollama")
    if "chromium_path" not in st.session_state:
        st.session_state["chromium_path"] = api_config.get("chromium_path", "")
    if "chromedriver_path" not in st.session_state:
        st.session_state["chromedriver_path"] = api_config.get("chromedriver_path", "")
    if "chromium_profile_dir" not in st.session_state:
        st.session_state["chromium_profile_dir"] = api_config.get("chromium_profile_dir", "")
    if "use_chromium_downloads" not in st.session_state:
        st.session_state["use_chromium_downloads"] = bool(api_config.get("use_chromium_downloads", False))
    if "use_remote_debugging" not in st.session_state:
        st.session_state["use_remote_debugging"] = bool(api_config.get("use_remote_debugging", False))
    if "remote_debug_port" not in st.session_state:
        st.session_state["remote_debug_port"] = int(api_config.get("remote_debug_port", 9222))
    if "reduce_automation" not in st.session_state:
        st.session_state["reduce_automation"] = bool(api_config.get("reduce_automation", True))
    
    current_email = st.session_state["openalex_email"]
    current_key = st.session_state["s2_api_key"]
    current_elsevier_key = st.session_state["elsevier_api_key"]
    current_glm_key = st.session_state["glm_api_key"]

    email_placeholder = "Email already added" if current_email else "Enter OpenAlex Email"
    new_email_input = st.text_input("OpenAlex Email", value="", placeholder=email_placeholder)
    
    key_placeholder = "API key already added" if current_key else "Enter S2 API Key"
    new_key_input = st.text_input("Semantic Scholar API Key", value="", placeholder=key_placeholder, type="password")
    
    elsevier_placeholder = "API key already added" if current_elsevier_key else "Enter Elsevier API Key"
    new_elsevier_input = st.text_input("Elsevier API Key", value="", placeholder=elsevier_placeholder, type="password")

    glm_placeholder = "API key already added" if current_glm_key else "Enter GLM API Key"
    new_glm_input = st.text_input("GLM API Key", value="", placeholder=glm_placeholder, type="password")
    glm_api_key_runtime = new_glm_input.strip() if new_glm_input else current_glm_key
    st.session_state["glm_api_key_runtime"] = glm_api_key_runtime

    if st.button("Save API Config", use_container_width=True):
        final_email = new_email_input.strip() if new_email_input else current_email
        final_key = new_key_input.strip() if new_key_input else current_key
        final_elsevier = new_elsevier_input.strip() if new_elsevier_input else current_elsevier_key
        final_glm_key = new_glm_input.strip() if new_glm_input else current_glm_key
        current_chromium_path = st.session_state["chromium_path"]
        current_chromedriver_path = st.session_state["chromedriver_path"]
        current_chromium_profile = st.session_state["chromium_profile_dir"]
        current_use_chromium = st.session_state["use_chromium_downloads"]
        current_glm_base_url = st.session_state.get("glm_base_url", GLM_BASE_URL)
        current_llm_provider = st.session_state.get("llm_provider", "ollama")
        final_chromium_path = current_chromium_path
        final_chromedriver_path = current_chromedriver_path
        final_chromium_profile = current_chromium_profile
        new_config = {
            "openalex_email": final_email, 
            "s2_api_key": final_key,
            "elsevier_api_key": final_elsevier,
            "glm_api_key": final_glm_key,
            "glm_base_url": current_glm_base_url,
            "llm_provider": current_llm_provider,
            "chromium_path": final_chromium_path,
            "chromedriver_path": final_chromedriver_path,
            "chromium_profile_dir": final_chromium_profile,
            "use_chromium_downloads": bool(current_use_chromium)
        }
        if save_api_config(new_config):
            st.session_state["openalex_email"] = final_email
            st.session_state["s2_api_key"] = final_key
            st.session_state["elsevier_api_key"] = final_elsevier
            st.session_state["glm_api_key"] = final_glm_key
            st.session_state["glm_base_url"] = current_glm_base_url
            st.session_state["llm_provider"] = current_llm_provider
            st.session_state["chromium_path"] = final_chromium_path
            st.session_state["chromedriver_path"] = final_chromedriver_path
            st.session_state["chromium_profile_dir"] = final_chromium_profile
            st.session_state["use_chromium_downloads"] = bool(current_use_chromium)
            flash("Saved!", level="success", scope="sidebar")
            st.rerun()

    st.markdown("---")  # Spacer

    _render_section_help(
        "writer_model_selection",
        "Choose the LLM used for planning, search queries, and relevance scoring."
    )

    provider_labels = ["Ollama (local)", "Z.ai GLM"]
    provider_values = ["ollama", "glm"]
    current_provider = _normalize_provider(st.session_state.get("llm_provider", "ollama"))
    try:
        provider_idx = provider_values.index(current_provider)
    except ValueError:
        provider_idx = 0
    selected_provider_label = st.selectbox("LLM Provider", provider_labels, index=provider_idx)
    llm_provider = provider_values[provider_labels.index(selected_provider_label)]
    st.session_state["llm_provider"] = llm_provider

    if llm_provider == "ollama":
        ollama_url = st.text_input("Ollama URL", value=st.session_state.get("ollama_url", OLLAMA_URL))
        st.session_state["ollama_url"] = ollama_url
        model_options = ["gpt-oss:120b-cloud", "nemotron-3-nano:30b-cloud", "kimi-k2-thinking:cloud"]
    else:
        glm_base_url = st.text_input("GLM Base URL", value=st.session_state.get("glm_base_url", GLM_BASE_URL))
        st.session_state["glm_base_url"] = glm_base_url
        st.caption(f"Mainland China endpoint: {GLM_ALT_BASE_URL}")
        model_options = ["glm-4.7", "glm-4.7-flash"]
    
    # Model selection
    if "active_model" not in st.session_state:
        st.session_state["active_model"] = PRIMARY_MODEL

    if llm_provider == "glm" and not str(st.session_state["active_model"]).startswith("glm-"):
        st.session_state["active_model"] = model_options[0]
    
    # Resolve current model index.
    current_idx = 0
    if st.session_state["active_model"] in model_options:
        current_idx = model_options.index(st.session_state["active_model"])
        
    selected_preset = st.selectbox("Select Model", model_options, index=current_idx)
    
    if st.button("Update Model", use_container_width=True):
        st.session_state["active_model"] = selected_preset
        st.rerun()

    primary_model = st.text_input("Active Model Name", value=st.session_state["active_model"])

    st.markdown("**Connection Test**")
    _render_section_help(
        "writer_connection_test",
        "Verify the model endpoint and credentials before starting the agent."
    )
    if llm_provider == "ollama":
        ollama_api_key_input = st.text_input("Ollama Key", value="", type="password", placeholder="Optional")
        ollama_api_key = ollama_api_key_input.strip() if ollama_api_key_input else st.session_state.get("ollama_api_key", "")
        st.session_state["ollama_api_key"] = ollama_api_key
        llm_base_url = ollama_url
        llm_api_key = ollama_api_key
    else:
        llm_base_url = glm_base_url
        llm_api_key = glm_api_key_runtime
        if not llm_api_key:
            flash("GLM API key is empty. Add it in API Config above to enable requests.", level="warning", scope="sidebar")
    show_debug = st.checkbox("Show Debug Info", value=False)
    
    if st.button("Test Connection", use_container_width=True):
        is_connected, model, msg = check_llm_connection(
            provider=llm_provider,
            base_url=llm_base_url,
            preferred_model=primary_model,
            api_key=llm_api_key,
            debug=show_debug,
        )
        if is_connected:
            flash(f"Connected: {model}", level="success", scope="sidebar")
            if msg:
                flash(msg, level="warning", scope="sidebar")
        else:
            flash(f"{msg}", level="error", scope="sidebar")

    st.markdown(
        "<div style='font-size:10px; opacity:0.5;'>developed by redredblood32</div>",
        unsafe_allow_html=True
    )

# Primary inputs
c1, c2, c3 = st.columns([4, 1, 1])
with c1:
    _render_section_help(
        "writer_research_topic",
        "Enter the topic and start the research agent. The agent plans sections, searches sources, and ranks papers."
    )
    topic = st.text_input("Research Topic", placeholder="e.g. Generative Agents in Social Simulations")
    if st.session_state.get("agent_running", False):
        if st.button("Stop Agent", use_container_width=True, type="secondary"):
            st.session_state["agent_stopped"] = True
            st.session_state["agent_running"] = False
            st.session_state.pop("agent_flow_state", None)
            st.session_state.pop("agent_timeline", None)
            st.session_state["agent_flow_notice"] = "Agent run stopped by user."
            st.rerun()
        start = False
    else:
        start = st.button("Start Research", use_container_width=True, type="primary")

with c2:
    _render_section_help(
        "writer_search_params",
        "Controls retrieval volume, timeout, and how many relevant papers are used for the Step 4 report."
    )
    num_results = st.number_input("Papers per query", min_value=5, max_value=50, value=10)
    search_timeout = st.number_input("Timeout (sec)", min_value=30, max_value=600, value=120)
    papers_for_review = st.number_input("Papers for review", min_value=5, max_value=200, value=30, step=5)

with c3:
    _render_section_help(
        "writer_export_controls",
        "Export the current session results to one full Excel file."
    )
    # Check if an active session exists to enable export.
    has_session = st.session_state.get("active_session") is not None
    export_clicked = st.button("Export Full Excel", use_container_width=True, disabled=not has_session)

# Export workflow
if export_clicked and has_session:
    session_data = st.session_state.get("active_session")
    if session_data:
        found_papers = session_data.get("found_papers", [])
        missing_papers = session_data.get("missing_papers", [])
        manual_papers = session_data.get("manual_papers", [])
        if not isinstance(found_papers, list):
            found_papers = []
        if not isinstance(missing_papers, list):
            missing_papers = []
        if not isinstance(manual_papers, list):
            manual_papers = []
        session_topic = session_data.get("topic", "research")
        session_id_export = _coerce_session_id(session_data.get("session_id")) or 0
        project_name = session_data.get("project") or st.session_state.get("current_project") or "Thesis"
        _ensure_project_layout(project_name)
        papers_to_export = found_papers + missing_papers + manual_papers

        if not papers_to_export:
            flash("No papers available to export.", level="warning")
        else:
            sanitized_topic = _sanitize_topic(session_topic)
            export_dir = _get_session_export_dir(project_name, session_id_export)
            excel_filename = f"papers_{sanitized_topic}.xlsx"
            excel_path = export_dir / excel_filename
            with st.spinner("Generating full Excel file..."):
                try:
                    generate_excel_export(papers_to_export, str(excel_path), session_topic)
                except Exception as e:
                    flash(f"Excel export failed: {e}", level="error")
                    excel_path = None

            export_label = f"{project_name}/session_{session_id_export}"
            if excel_path:
                flash(
                    f"Export to '{export_label}': Excel file saved: {excel_filename}",
                    level="success"
                )

# Load saved session if requested.
if st.session_state.get("load_session_id"):
    session_id = st.session_state.pop("load_session_id")
    project_name = st.session_state.pop("load_session_project", None) or st.session_state.get("current_project")
    if not project_name:
        project_name = find_session_project(session_id)
    session_data = load_session(project_name, session_id) if project_name else None
    if session_data:
        st.session_state["active_session"] = session_data
        st.session_state["current_session_id"] = session_id
        if project_name:
            st.session_state["current_project"] = project_name
        st.rerun()

# Handle user stop request.
if st.session_state.get("agent_stopped", False):
    st.session_state.pop("agent_flow_state", None)
    st.session_state.pop("agent_timeline", None)
    flash("Agent stopped by user.", level="warning")
    st.session_state["agent_stopped"] = False

flow_notice = st.session_state.pop("agent_flow_notice", None)
if flow_notice:
    flash(flow_notice, level="warning")

# -----------------------------------------------------------------------------
# AGENT EXECUTION FLOW (STEP 1-4, WITH APPROVAL GATE)
# -----------------------------------------------------------------------------

# 1) Trigger start
if start and topic:
    st.session_state["agent_running"] = True
    st.session_state["agent_flow_state"] = {
        "phase": "step1",
        "topic": topic,
        "model_name": primary_model,
        "provider": llm_provider,
        "base_url": llm_base_url,
        "api_key_present": bool(llm_api_key),
        "num_results": int(num_results),
        "search_timeout": int(search_timeout),
        "papers_for_review": int(papers_for_review),
        "draft_outputs": {
            "web_questions": [],
            "web_context": "",
            "web_summary": "",
            "section_plan": "",
            "section_queries": {},
        },
        "last_message": "Initialized",
        "cancel_reason": "",
    }
    st.session_state["agent_timeline"] = {}
    st.rerun()

flow_state = st.session_state.get("agent_flow_state")
approval_phase_map = {
    "await_approval_search": ("4", "Search ready. Approve to start Step 4."),
}
phase_progression = {
    "await_approval_search": "search_and_analyze",
}

if st.session_state.get("agent_running", False) and isinstance(flow_state, dict):
    phase = flow_state.get("phase")
    if phase in approval_phase_map:
        step_num, step_title = approval_phase_map[phase]
        st.markdown("---")
        st.markdown(f"### Approval Checkpoint: Step {step_num}")
        st.info(step_title)
        col_approve, col_deny, _ = st.columns([2, 2, 8])
        with col_approve:
            if st.button("Approve & Start Search", key=f"approve_{phase}", use_container_width=True):
                flow_state["phase"] = phase_progression[phase]
                flow_state["last_message"] = f"Approved Step {step_num}"
                st.session_state["agent_flow_state"] = flow_state
                st.rerun()
        with col_deny:
            if st.button("Deny & Cancel", key=f"deny_{phase}", type="secondary", use_container_width=True):
                st.session_state["agent_running"] = False
                st.session_state.pop("agent_flow_state", None)
                st.session_state["agent_flow_notice"] = f"Run cancelled by user at Step {step_num}."
                st.rerun()

# 2) Run agent (state machine)
if st.session_state.get("agent_running", False):
    flow_state = st.session_state.get("agent_flow_state")
    if not isinstance(flow_state, dict):
        flow_state = {
            "phase": "step1",
            "topic": topic,
            "model_name": primary_model,
            "provider": llm_provider,
            "base_url": llm_base_url,
            "api_key_present": bool(llm_api_key),
            "num_results": int(num_results),
            "search_timeout": int(search_timeout),
            "papers_for_review": int(papers_for_review),
            "draft_outputs": {
                "web_questions": [],
                "web_context": "",
                "web_summary": "",
                "section_plan": "",
                "section_queries": {},
            },
            "last_message": "Initialized",
            "cancel_reason": "",
        }
        st.session_state["agent_flow_state"] = flow_state

    phase = flow_state.get("phase", "step1")
    compute_phases = {"step1", "step2", "step3", "search_and_analyze"}
    if phase not in compute_phases:
        st.info("Agent is paused and waiting for your approval.")
    else:
        run_topic = flow_state.get("topic") or topic
        run_model = flow_state.get("model_name") or primary_model
        run_provider = flow_state.get("provider") or llm_provider
        run_base_url = flow_state.get("base_url") or llm_base_url
        run_num_results = int(flow_state.get("num_results", num_results))
        run_timeout = int(flow_state.get("search_timeout", search_timeout))
        run_papers_for_review = int(flow_state.get("papers_for_review", papers_for_review))

        is_connected, model, msg = check_llm_connection(
            provider=run_provider,
            base_url=run_base_url,
            preferred_model=run_model,
            api_key=llm_api_key,
            debug=show_debug,
        )
        if not is_connected:
            st.session_state["agent_running"] = False
            st.session_state.pop("agent_flow_state", None)
            flash(f"{msg}", level="error")
            st.stop()
        if msg:
            flash(msg, level="warning")

        with st.status("Agent is researching|", expanded=True) as status:
            timeline_placeholder = st.empty()
            _render_agent_timeline(timeline_placeholder)

            def _timeline_update(step_id, status_text, detail, progress=None):
                _set_agent_timeline_entry(
                    step_id,
                    status_text=status_text,
                    detail=detail,
                    progress=progress
                )
                _render_agent_timeline(timeline_placeholder)

            try:
                draft = flow_state.setdefault("draft_outputs", {})

                if phase == "step1":
                    status.update(state="running")
                    _timeline_update("step1", "Running", "Web research (finding Wikipedia pages)")

                    def _summary_start():
                        _timeline_update("step1_5", "Running", "Summarizing Wikipedia")

                    def _summary_progress(current, total):
                        progress = {
                            "current": current,
                            "total": total,
                            "text": f"{current} / {total} pages done"
                        }
                        _timeline_update("step1_5", "Running", "Summarizing Wikipedia", progress=progress)

                    def _summary_complete():
                        _timeline_update("step1_5", "Finished", "Summarizing Wikipedia complete", progress=None)

                    step1 = _run_agent_step1(
                        run_topic,
                        model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        summary_start_callback=_summary_start,
                        summary_progress_callback=_summary_progress,
                        summary_complete_callback=_summary_complete,
                    )
                    draft["web_questions"] = step1.get("web_questions", [])
                    draft["web_context"] = step1.get("web_context", "")
                    draft["web_summary"] = step1.get("web_summary", "")
                    flow_state["phase"] = "step2"
                    flow_state["last_message"] = "Step 1 complete"
                    st.session_state["agent_flow_state"] = flow_state
                    _timeline_update("step1", "Finished", "Web research complete")
                    status.update(state="complete", expanded=False)
                    st.rerun()

                if phase == "step2":
                    status.update(state="running")
                    _timeline_update("step2", "Running", "Drafting section plan")
                    section_plan = _run_agent_step2(
                        run_topic,
                        draft.get("web_context", ""),
                        draft.get("web_summary", ""),
                        model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        show_debug=show_debug,
                    )
                    draft["section_plan"] = section_plan
                    flow_state["phase"] = "step3"
                    flow_state["last_message"] = "Step 2 complete"
                    st.session_state["agent_flow_state"] = flow_state
                    _timeline_update("step2", "Finished", "Section plan complete")
                    status.update(state="complete", expanded=False)
                    st.rerun()

                if phase == "step3":
                    status.update(state="running")
                    _timeline_update("step3", "Running", "Generating academic search queries")
                    section_plan = draft.get("section_plan", "")
                    if not section_plan:
                        raise ValueError("Step 2 output missing section plan.")
                    section_queries = _run_agent_step3(
                        section_plan,
                        model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        show_debug=show_debug,
                    )
                    draft["section_queries"] = section_queries
                    flow_state["phase"] = "await_approval_search"
                    flow_state["last_message"] = "Step 3 complete"
                    st.session_state["agent_flow_state"] = flow_state
                    _timeline_update("step3", "Finished", "Academic search queries complete")
                    status.update(state="complete", expanded=False)
                    st.rerun()

                if phase == "search_and_analyze":
                    status.update(state="running")
                    _timeline_update("search_and_analyze", "Running", "Searching academic databases")
                    section_queries = draft.get("section_queries", {})
                    if not isinstance(section_queries, dict) or not section_queries:
                        raise ValueError("Step 3 output missing section queries.")

                    query_groups = {}
                    for section_name, query_list in section_queries.items():
                        cleaned = flatten_string_queries(query_list)
                        if cleaned:
                            query_groups[section_name] = cleaned

                    total_queries = sum(len(qs) for qs in query_groups.values()) if query_groups else 0
                    progress_bar = st.progress(0.0)
                    timer_placeholder = st.empty()

                    def update_progress(completed, total):
                        if total <= 0:
                            return
                        fraction = min(1.0, completed / total)
                        progress_bar.progress(fraction)
                        _timeline_update(
                            "search_and_analyze",
                            "Running",
                            f"Searching academic databases ({completed}/{total})",
                            progress={
                                "primary": {
                                    "current": completed,
                                    "total": total,
                                    "text": f"Search progress: {completed} / {total} queries"
                                }
                            }
                        )

                    found, missing = search_api(
                        query_groups,
                        run_num_results,
                        progress_callback=update_progress,
                        total_queries=total_queries,
                        timeout_seconds=run_timeout,
                        timer_placeholder=timer_placeholder,
                        s2_api_key=st.session_state.get("s2_api_key"),
                        openalex_email=st.session_state.get("openalex_email"),
                        elsevier_api_key=st.session_state.get("elsevier_api_key")
                    )

                    progress_bar.empty()
                    timer_placeholder.empty()

                    _timeline_update("search_and_analyze", "Running", "Deduplicating papers")
                    deduped, dedup_stats = deduplicate_papers(found + missing)
                    found = [p for p in deduped if p.get("pdf_url")]
                    missing = [p for p in deduped if not p.get("pdf_url")]
                    print(
                        f"[Dedup] input={dedup_stats['input']} "
                        f"groups={dedup_stats['groups']} output={dedup_stats['output']} "
                        f"found={len(found)} missing={len(missing)}"
                    )
                    for msg in dedup_stats.get("fuzzy_warnings", []):
                        print(f"[Dedup] warning: {msg}")

                    total_analysis = len(found) + len(missing)
                    analysis_done = {"count": 0, "phase": "found"}
                    total_found = len(found)
                    total_missing = len(missing)

                    def _analysis_progress(processed, _total):
                        if analysis_done["phase"] == "found":
                            current = min(processed, total_found)
                            analysis_done["count"] = min(current, total_analysis)
                            if current >= total_found:
                                analysis_done["phase"] = "missing"
                        else:
                            current_missing = min(processed, total_missing)
                            analysis_done["count"] = min(total_found + current_missing, total_analysis)
                        progress = {
                            "primary": {
                                "current": total_queries,
                                "total": total_queries,
                                "text": f"Search progress: {total_queries} / {total_queries} queries"
                            },
                            "analysis": {
                                "current": analysis_done["count"],
                                "total": total_analysis,
                                "text": f"Analyzing relevance: {analysis_done['count']} / {total_analysis} papers"
                            }
                        }
                        _timeline_update("search_and_analyze", "Running", "Analyzing content relevance", progress=progress)

                    _timeline_update("search_and_analyze", "Running", "Analyzing content relevance", progress={
                        "primary": {
                            "current": total_queries,
                            "total": total_queries,
                            "text": f"Search progress: {total_queries} / {total_queries} queries"
                        },
                        "analysis": {
                            "current": 0,
                            "total": total_analysis,
                            "text": f"Analyzing relevance: 0 / {total_analysis} papers"
                        }
                    })
                    found = score_relevance(
                        run_topic,
                        found,
                        model_name=model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        show_debug=show_debug,
                        progress_callback=_analysis_progress,
                    )
                    missing = score_relevance(
                        run_topic,
                        missing,
                        model_name=model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        show_debug=show_debug,
                        progress_callback=_analysis_progress,
                    )

                    found = sort_papers_by_relevance(found)
                    missing = sort_papers_by_relevance(missing)
                    _timeline_update("search_and_analyze", "Running", "Generating abstract-based report")
                    relevant_candidates = [p for p in (found + missing) if p.get("is_relevant", False)]

                    def _report_sort_key(paper):
                        try:
                            relevance_val = float(paper.get("relevance_score", 0) or 0)
                        except (TypeError, ValueError):
                            relevance_val = 0.0
                        try:
                            citations_val = int(paper.get("citations", 0) or 0)
                        except (TypeError, ValueError):
                            citations_val = 0
                        title_val = str(paper.get("title", "") or "")
                        paper_id_val = str(paper.get("paper_id", "") or "")
                        return (-relevance_val, -citations_val, title_val, paper_id_val)

                    sorted_relevant_for_report = sorted(relevant_candidates, key=_report_sort_key)
                    report_source_papers = sorted_relevant_for_report[:max(1, run_papers_for_review)]
                    if not report_source_papers:
                        _timeline_update("search_and_analyze", "Running", "Generating abstract-based report", progress={
                            "primary": {
                                "current": total_queries,
                                "total": total_queries,
                                "text": f"Search progress: {total_queries} / {total_queries} queries"
                            },
                            "analysis": {
                                "current": total_analysis,
                                "total": total_analysis,
                                "text": f"Analyzing relevance: {total_analysis} / {total_analysis} papers"
                            },
                            "report": {
                                "current": 0,
                                "total": 1,
                                "text": "Report generation skipped (no relevant papers)"
                            }
                        })
                    def _report_progress(current, total, phase_label):
                        if total <= 0:
                            return
                        if phase_label == "synthesis":
                            text = "Report generation: final synthesis"
                            curr_val = total
                        else:
                            text = f"Report generation: batch {current} / {total}"
                            curr_val = current
                        progress = {
                            "primary": {
                                "current": total_queries,
                                "total": total_queries,
                                "text": f"Search progress: {total_queries} / {total_queries} queries"
                            },
                            "analysis": {
                                "current": total_analysis,
                                "total": total_analysis,
                                "text": f"Analyzing relevance: {total_analysis} / {total_analysis} papers"
                            },
                            "report": {
                                "current": curr_val,
                                "total": total,
                                "text": text
                            }
                        }
                        _timeline_update("search_and_analyze", "Running", "Generating abstract-based report", progress=progress)

                    abstract_report_bundle = generate_abstract_report(
                        run_topic,
                        report_source_papers,
                        model_name=model,
                        provider=run_provider,
                        base_url=run_base_url,
                        api_key=llm_api_key,
                        show_debug=show_debug,
                        batch_size=5,
                        progress_callback=_report_progress,
                    )
                    abstract_report = abstract_report_bundle.get("report_markdown", "")
                    abstract_report_references = abstract_report_bundle.get("references", [])
                    abstract_report_meta = {
                        "relevant_total": len(sorted_relevant_for_report),
                        "used_for_report": len(report_source_papers),
                        "limit": int(run_papers_for_review),
                    }

                    curr_proj = st.session_state.get("current_project") or "Thesis"
                    projects_data = load_projects()
                    if curr_proj not in projects_data:
                        create_project(curr_proj, projects_data)
                        projects_data = load_projects()

                    session_id = get_next_session_id(curr_proj)
                    session_data = {
                        "session_id": session_id,
                        "timestamp": datetime.now().isoformat(),
                        "topic": run_topic,
                        "num_results_per_query": run_num_results,
                        "papers_for_review": int(run_papers_for_review),
                        "web_questions": draft.get("web_questions", []),
                        "web_context": draft.get("web_context", ""),
                        "web_summary": draft.get("web_summary", ""),
                        "section_plan": draft.get("section_plan", ""),
                        "section_queries": section_queries,
                        "found_papers": found,
                        "missing_papers": missing,
                        "manual_papers": [],
                        "abstract_report": abstract_report,
                        "abstract_report_references": abstract_report_references,
                        "abstract_report_meta": abstract_report_meta,
                        "project": curr_proj,
                    }
                    save_session(curr_proj, session_id, session_data)

                    if curr_proj in projects_data:
                        projects_data[curr_proj].append(session_id)
                        projects_data[curr_proj] = sorted(list(set(projects_data[curr_proj])))
                        save_projects(projects_data)
                    else:
                        projects_data[curr_proj] = [session_id]
                        save_projects(projects_data)

                    st.session_state["active_session"] = session_data
                    st.session_state["current_session_id"] = session_id
                    st.session_state["agent_running"] = False
                    st.session_state.pop("agent_flow_state", None)

                    _timeline_update("search_and_analyze", "Finished", "Research complete")
                    status.update(state="complete", expanded=False)
                    st.rerun()

            except Exception as e:
                st.session_state["agent_running"] = False
                st.session_state.pop("agent_flow_state", None)
                error_step = phase if phase in {"step1", "step2", "step3", "search_and_analyze"} else "step1"
                _timeline_update(error_step, "Error", f"{e}")
                status.update(state="error")
                flash(f"Error: {e}", level="error")
                st.stop()

flow_state = st.session_state.get("agent_flow_state")
if st.session_state.get("agent_running", False) and isinstance(flow_state, dict):
    draft = flow_state.get("draft_outputs", {})
    if not isinstance(draft, dict):
        draft = {}
    has_draft = bool(
        draft.get("web_questions")
        or draft.get("web_context")
        or draft.get("web_summary")
        or draft.get("section_plan")
        or draft.get("section_queries")
    )
    if has_draft:
        st.markdown("---")
        st.markdown(f"### In-Progress Results: {flow_state.get('topic', 'Research')}")
        col_draft_head, col_draft_toggle, _ = st.columns([2, 3, 10])
        with col_draft_head:
            st.markdown("#### Agent Log")
        with col_draft_toggle:
            show_draft_steps = st.toggle("Show Steps", value=True, key="show_draft_steps")
        if show_draft_steps:
            _render_agent_log_cards(
                draft.get("web_questions", []),
                draft.get("web_context", ""),
                draft.get("web_summary", ""),
                draft.get("section_plan", ""),
                draft.get("section_queries", {}),
                abstract_report="",
            )

# -----------------------------------------------------------------------------
# MANUAL IDENTIFIER LOOKUP (WHEN NO ACTIVE SESSION)
# -----------------------------------------------------------------------------
if not st.session_state.get("active_session") and not st.session_state.get("agent_running", False):
    st.markdown("---")
    st.markdown("### Manual Search")
    _render_section_help(
        "writer_manual_search_results",
        "Paste a DOI or arXiv ID to fetch metadata and PDF availability. Results will appear once a session is created."
    )

    col_manual_input, col_manual_action, _ = st.columns([6, 2, 7])
    with col_manual_input:
        if st.session_state.pop("manual_doi_input_reset", False):
            st.session_state["manual_doi_input"] = ""
        manual_doi_value = st.text_input(
            "Paste DOI or arXiv ID",
            placeholder="10.0000/example.doi or 2301.01234",
            key="manual_doi_input"
        )
    with col_manual_action:
        fetch_manual = st.button("Fetch Identifier", use_container_width=True, key="manual_doi_fetch_empty")

    if fetch_manual:
        session_data = _ensure_manual_session(topic_hint=topic or "Manual Search", num_results=num_results)
        project_name = session_data.get("project") or st.session_state.get("current_project") or "Thesis"
        session_id_export = _coerce_session_id(session_data.get("session_id")) or 0
        manual_papers = session_data.get("manual_papers", [])
        if not isinstance(manual_papers, list):
            manual_papers = []
        found = session_data.get("found_papers", [])
        missing = session_data.get("missing_papers", [])
        _handle_manual_fetch(
            manual_doi_value,
            session_data=session_data,
            manual_papers=manual_papers,
            found=found,
            missing=missing,
            project_name=project_name,
            session_id_export=session_id_export,
            topic_label=session_data.get("topic", ""),
            primary_model=primary_model,
            provider=llm_provider,
            base_url=llm_base_url,
            api_key=llm_api_key,
            show_debug=show_debug,
            s2_api_key=st.session_state.get("s2_api_key"),
            openalex_email=st.session_state.get("openalex_email"),
            elsevier_api_key=st.session_state.get("elsevier_api_key"),
        )

# -----------------------------------------------------------------------------
# ACTIVE SESSION DISPLAY (RESULTS + CONTROLS)
# -----------------------------------------------------------------------------
if st.session_state.get("active_session"):
    data = st.session_state["active_session"]
    
    # Extract variables
    web_questions = data.get("web_questions", [])
    web_context = data.get("web_context", "")
    web_summary = data.get("web_summary", "")
    section_plan = data.get("section_plan", "")
    section_queries = data.get("section_queries", {})
    abstract_report = data.get("abstract_report", "")
    abstract_report_meta = data.get("abstract_report_meta", {})
    if not isinstance(abstract_report_meta, dict):
        abstract_report_meta = {}
    abstract_report_references = data.get("abstract_report_references", [])
    if not isinstance(abstract_report_references, list):
        abstract_report_references = []
    linked_abstract_report = _link_abstract_report_refs(abstract_report, abstract_report_references)
    report_refs_by_paper_id = {}
    for ref in abstract_report_references:
        if not isinstance(ref, dict):
            continue
        paper_id = str(ref.get("paper_id") or "").strip()
        if not paper_id:
            continue
        report_refs_by_paper_id[paper_id] = ref
    found = data.get("found_papers", [])
    missing = data.get("missing_papers", [])
    manual_papers = data.get("manual_papers", [])
    if not isinstance(manual_papers, list):
        manual_papers = []
    
    # Calculate export directory and check which papers are downloaded
    session_topic = data.get("topic", "research")
    session_id_export = _coerce_session_id(data.get("session_id")) or 0
    project_name = data.get("project") or st.session_state.get("current_project") or "Thesis"
    sanitized_topic = _sanitize_topic(session_topic)
    export_dir = _get_session_export_dir(project_name, session_id_export)
    pdf_dir = _get_project_pdf_dir(project_name)
    ris_dir = _get_project_ris_dir(project_name)
    external_ris_dir = _get_project_external_ris_dir(project_name)
    downloaded_files = get_completed_downloads(project_name, pdf_dir)
    ris_files = get_ris_files([ris_dir, external_ris_dir])
    _update_doi_registry_for_session(project_name, session_id_export, found + missing + manual_papers)

    has_pending = _has_pending_downloads(project_name)
    if st.session_state.get("force_autorefresh") or has_pending:
        _enable_autorefresh(interval_ms=4000, key=f"downloads_autorefresh_{project_name}")
        if not has_pending:
            st.session_state["force_autorefresh"] = False
    
    st.markdown("---")
    st.markdown(f"### Results: {data.get('topic', 'Research')}")
    _render_section_help(
        "writer_results_overview",
        "Session output includes the agent log, search queries, and ranked paper lists."
    )
    col_header, col_toggle, _ = st.columns([2, 3, 10])
    with col_header:
        st.markdown("#### Agent Log")
    with col_toggle:
        show_steps = st.toggle("Show Steps", value=True)
    
    if show_steps:
        _render_agent_log_cards(
            web_questions,
            web_context,
            web_summary,
            section_plan,
            section_queries,
            abstract_report=linked_abstract_report,
            abstract_report_meta=abstract_report_meta,
        )

    # Check if Excel file exists for this session
    excel_filename = f"papers_{sanitized_topic}.xlsx"
    excel_exists = (export_dir / excel_filename).exists()

    col_manual_head, col_manual_toggle, _ = st.columns([2, 3, 10])
    with col_manual_head:
        st.markdown(f"#### Manual Search Results ({len(manual_papers)})")
    _render_section_help(
        "writer_manual_search_results",
        "Paste a DOI or arXiv ID to fetch metadata and PDF availability. Results appear here with the same controls as other sections."
    )
    with col_manual_toggle:
        show_manual = st.toggle("Show", value=True, key="toggle_manual")

    col_manual_input, col_manual_action, _ = st.columns([6, 2, 7])
    with col_manual_input:
        if st.session_state.pop("manual_doi_input_reset", False):
            st.session_state["manual_doi_input"] = ""
        manual_doi_value = st.text_input(
            "Paste DOI or arXiv ID",
            placeholder="10.0000/example.doi or 2301.01234",
            key="manual_doi_input"
        )
    with col_manual_action:
        fetch_manual = st.button("Fetch Identifier", use_container_width=True, key="manual_doi_fetch")

    if fetch_manual:
        _handle_manual_fetch(
            manual_doi_value,
            session_data=st.session_state.get("active_session") or data,
            manual_papers=manual_papers,
            found=found,
            missing=missing,
            project_name=project_name,
            session_id_export=session_id_export,
            topic_label=data.get("topic", ""),
            primary_model=primary_model,
            provider=llm_provider,
            base_url=llm_base_url,
            api_key=llm_api_key,
            show_debug=show_debug,
            s2_api_key=st.session_state.get("s2_api_key"),
            openalex_email=st.session_state.get("openalex_email"),
            elsevier_api_key=st.session_state.get("elsevier_api_key"),
        )

    if show_manual:
        if manual_papers:
            feedback_key = "paper_feedback_manual"
            if feedback_key not in st.session_state:
                st.session_state[feedback_key] = data.get("paper_feedback_manual", {})

            sorted_manual = sort_papers_with_feedback(manual_papers, st.session_state[feedback_key])

            for idx, p in enumerate(sorted_manual, 1):
                paper_id = p.get("paper_id", p.get("title", f"manual_{idx}"))
                paper_key = p.get("paper_id") or f"manual_{idx}"
                paper_filename = get_paper_filename(p, idx)
                is_downloaded = paper_filename in downloaded_files
                ris_filename = get_ris_filename(p, idx)
                has_citation = ris_filename in ris_files
                also_in_entries = _get_doi_cross_project_entries(
                    p.get("doi"),
                    project_name,
                    session_id_export
                )
                also_in_html = _format_cross_project_badge(also_in_entries)
                note_key = _note_key_for_paper(p)
                note_payload = _get_active_session_notes().get(note_key, {}) if note_key else {}
                note_content = note_payload.get("content", "") if isinstance(note_payload, dict) else ""
                wrap_content = note_payload.get("wrap_up", "") if isinstance(note_payload, dict) else ""
                pdf_toggle_key = f"pdf_toggle_manual_{paper_key}"
                note_height = 600 if st.session_state.get(pdf_toggle_key) else 120
                is_open_access = bool(p.get("pdf_url"))

                with st.container():
                    col1, col_fb, col2 = st.columns([4, 0.5, 1])
                    with col1:
                        st.markdown(
                            render_paper_card(
                                p,
                                is_open_access,
                                downloaded=is_downloaded,
                                excel_exported=excel_exists,
                                has_citation=has_citation,
                                also_in_html=also_in_html
                            ),
                            unsafe_allow_html=True
                        )
                    with col_fb:
                        current_vote = st.session_state[feedback_key].get(paper_id)
                        fb_default = None if current_vote is None else (1 if current_vote == 1 else 0)
                        feedback = st.feedback("thumbs", key=f"fb_manual_{paper_key}", default=fb_default)
                        if feedback is not None:
                            new_vote = 1 if feedback == 1 else -1
                            if st.session_state[feedback_key].get(paper_id) != new_vote:
                                st.session_state[feedback_key][paper_id] = new_vote
                                if st.session_state.get("active_session"):
                                    st.session_state["active_session"]["paper_feedback_manual"] = st.session_state[feedback_key]
                                    project_name = st.session_state.get("current_project") or st.session_state.get("active_session", {}).get("project")
                                    if project_name:
                                        save_session(project_name, st.session_state.get("current_session_id"), st.session_state["active_session"])
                                st.rerun()
                    with col2:
                        if note_key:
                            _render_note_popover(
                                f"note_manual_{paper_key}",
                                project_name,
                                session_id_export,
                                note_key,
                                note_content,
                                wrap_content,
                                p,
                                height=note_height
                            )

                    if is_downloaded:
                        show_pdf = st.toggle("View Downloaded PDF", key=pdf_toggle_key, value=True)
                        if show_pdf:
                            pdf_path = _find_project_pdf_path(project_name, paper_filename)
                            if pdf_path and pdf_path.exists():
                                _render_pdf_viewer_with_highlights(
                                    pdf_path,
                                    p,
                                    paper_key,
                                    project_name,
                                    session_id_export,
                                )
                            else:
                                flash(f"PDF file not found: {paper_filename}", level="warning")
        else:
            flash("No manual DOI results found.", level="info")
    
    # Paper results
    col_found_head, col_found_toggle, _ = st.columns([2, 3, 10])
    with col_found_head:
        st.markdown(f"#### Open Access Papers ({len(found)})")
    _render_section_help(
        "writer_open_access_papers",
        "These papers have open-access PDFs and full metadata."
    )
    with col_found_toggle:
        show_found = st.toggle("Show", value=True, key="toggle_found")

    if found:
        if show_found:
            # Retrieve feedback from session state (synced with session JSON).
            feedback_key = "paper_feedback_found"
            if feedback_key not in st.session_state:
                st.session_state[feedback_key] = data.get("paper_feedback_found", {})
            
            # Sort using feedback + relevance.
            sorted_found = sort_papers_with_feedback(found, st.session_state[feedback_key])
            
            for idx, p in enumerate(sorted_found, 1):
                paper_id = p.get('paper_id', p.get('title', f'found_{idx}'))  # Use UUID if available
                # Use paper_id for widget keys (guaranteed unique)
                paper_key = p.get('paper_id') or f"f{idx}"
                paper_id_actual = str(p.get("paper_id") or "")
                report_ref_entry = report_refs_by_paper_id.get(paper_id_actual) if paper_id_actual else None
                reference_tag = report_ref_entry.get("ref", "") if isinstance(report_ref_entry, dict) else ""
                reference_anchor_id = report_ref_entry.get("anchor_id", "") if isinstance(report_ref_entry, dict) else ""
                paper_filename = get_paper_filename(p, idx)
                is_downloaded = paper_filename in downloaded_files
                ris_filename = get_ris_filename(p, idx)
                has_citation = ris_filename in ris_files
                also_in_entries = _get_doi_cross_project_entries(
                    p.get("doi"),
                    project_name,
                    session_id_export
                )
                also_in_html = _format_cross_project_badge(also_in_entries)
                note_key = _note_key_for_paper(p)
                note_payload = _get_active_session_notes().get(note_key, {}) if note_key else {}
                note_content = note_payload.get("content", "") if isinstance(note_payload, dict) else ""
                wrap_content = note_payload.get("wrap_up", "") if isinstance(note_payload, dict) else ""
                wrap_content = note_payload.get("wrap_up", "") if isinstance(note_payload, dict) else ""
                pdf_toggle_key = f"pdf_toggle_{paper_key}"
                note_height = 600 if st.session_state.get(pdf_toggle_key) else 120
                
                with st.container():
                    col1, col_fb, col2 = st.columns([4, 0.5, 1])
                    with col1:
                        if reference_anchor_id:
                            st.markdown(f'<a id="{html.escape(reference_anchor_id)}"></a>', unsafe_allow_html=True)
                        st.markdown(
                            render_paper_card(
                                p,
                                True,
                                downloaded=is_downloaded,
                                excel_exported=excel_exists,
                                has_citation=has_citation,
                                also_in_html=also_in_html,
                                reference_tag=reference_tag,
                            ),
                            unsafe_allow_html=True
                        )
                    with col_fb:
                        current_vote = st.session_state[feedback_key].get(paper_id)
                        # Map: None->None, 1->1 (up), -1->0 (down).
                        fb_default = None if current_vote is None else (1 if current_vote == 1 else 0)
                        feedback = st.feedback("thumbs", key=f"fb_found_{paper_key}", default=fb_default)
                        if feedback is not None:
                            new_vote = 1 if feedback == 1 else -1
                            if st.session_state[feedback_key].get(paper_id) != new_vote:
                                st.session_state[feedback_key][paper_id] = new_vote
                                # Persist feedback into the session file.
                                if st.session_state.get("active_session"):
                                    st.session_state["active_session"]["paper_feedback_found"] = st.session_state[feedback_key]
                                    project_name = st.session_state.get("current_project") or st.session_state.get("active_session", {}).get("project")
                                    if project_name:
                                        save_session(project_name, st.session_state.get("current_session_id"), st.session_state["active_session"])
                                st.rerun()
                    with col2:
                        if note_key:
                            _render_note_popover(
                                f"note_found_{paper_key}",
                                project_name,
                                session_id_export,
                                note_key,
                                note_content,
                                wrap_content,
                                p,
                                height=note_height
                            )
                    
                    # PDF viewer toggle (only for downloaded PDFs).
                    if is_downloaded:
                        show_pdf = st.toggle("View Downloaded PDF", key=f"pdf_toggle_{paper_key}", value=True)
                        if show_pdf:
                            pdf_path = _find_project_pdf_path(project_name, paper_filename)
                            if pdf_path and pdf_path.exists():
                                _render_pdf_viewer_with_highlights(
                                    pdf_path,
                                    p,
                                    paper_key,
                                    project_name,
                                    session_id_export,
                                )
                            else:
                                flash(f"PDF file not found: {paper_filename}", level="warning")
    else:
        flash("No open access papers found.", level="info")

    if missing:
        col_miss_head, col_miss_toggle, _ = st.columns([2, 3, 10])
        with col_miss_head:
            st.markdown(f"#### Paywalled Papers ({len(missing)})")
        _render_section_help(
            "writer_paywalled_papers",
            "These papers are not open access. Review metadata and take notes here."
        )
        with col_miss_toggle:
            show_missing = st.toggle("Show", value=True, key="toggle_missing")

        if show_missing:
            # Retrieve feedback from session state (synced with session JSON).
            feedback_key = "paper_feedback_missing"
            if feedback_key not in st.session_state:
                st.session_state[feedback_key] = data.get("paper_feedback_missing", {})
            
            # Sort using feedback + relevance.
            sorted_missing = sort_papers_with_feedback(missing, st.session_state[feedback_key])
            
            for idx, p in enumerate(sorted_missing, 1):
                paper_id = p.get('paper_id', p.get('title', f'pay_{idx}'))  # Use UUID if available
                # Use paper_id for widget keys (guaranteed unique)
                paper_key = p.get('paper_id') or f"m{idx}"
                paper_id_actual = str(p.get("paper_id") or "")
                report_ref_entry = report_refs_by_paper_id.get(paper_id_actual) if paper_id_actual else None
                reference_tag = report_ref_entry.get("ref", "") if isinstance(report_ref_entry, dict) else ""
                reference_anchor_id = report_ref_entry.get("anchor_id", "") if isinstance(report_ref_entry, dict) else ""
                ris_filename = get_ris_filename(p, idx)
                has_citation = ris_filename in ris_files
                paper_filename_pay = get_paper_filename(p, idx)
                is_downloaded_pay = paper_filename_pay in downloaded_files
                also_in_entries = _get_doi_cross_project_entries(
                    p.get("doi"),
                    project_name,
                    session_id_export
                )
                also_in_html = _format_cross_project_badge(also_in_entries)
                note_key = _note_key_for_paper(p)
                note_payload = _get_active_session_notes().get(note_key, {}) if note_key else {}
                note_content = note_payload.get("content", "") if isinstance(note_payload, dict) else ""
                wrap_content = note_payload.get("wrap_up", "") if isinstance(note_payload, dict) else ""
                pdf_toggle_key = f"pdf_toggle_pay_{paper_key}"
                note_height = 600 if st.session_state.get(pdf_toggle_key) else 120
                
                with st.container():
                    col1, col_fb, col2 = st.columns([4, 0.5, 1])
                    with col1:
                        if reference_anchor_id:
                            st.markdown(f'<a id="{html.escape(reference_anchor_id)}"></a>', unsafe_allow_html=True)
                        st.markdown(
                            render_paper_card(
                                p,
                                False,
                                downloaded=is_downloaded_pay,
                                excel_exported=excel_exists,
                                has_citation=has_citation,
                                also_in_html=also_in_html,
                                reference_tag=reference_tag,
                            ),
                            unsafe_allow_html=True,
                        )
                    with col_fb:
                        current_vote = st.session_state[feedback_key].get(paper_id)
                        fb_default = None if current_vote is None else (1 if current_vote == 1 else 0)
                        feedback = st.feedback("thumbs", key=f"fb_miss_{paper_key}", default=fb_default)
                        if feedback is not None:
                            new_vote = 1 if feedback == 1 else -1
                            if st.session_state[feedback_key].get(paper_id) != new_vote:
                                st.session_state[feedback_key][paper_id] = new_vote
                                # Persist feedback into the session file.
                                if st.session_state.get("active_session"):
                                    st.session_state["active_session"]["paper_feedback_missing"] = st.session_state[feedback_key]
                                    project_name = st.session_state.get("current_project") or st.session_state.get("active_session", {}).get("project")
                                    if project_name:
                                        save_session(project_name, st.session_state.get("current_session_id"), st.session_state["active_session"])
                                st.rerun()
                    with col2:
                        if note_key:
                            _render_note_popover(
                                f"note_missing_{paper_key}",
                                project_name,
                                session_id_export,
                                note_key,
                                note_content,
                                wrap_content,
                                p,
                                height=note_height
                            )
                    
                    # PDF viewer toggle (only for downloaded PDFs).
                    if is_downloaded_pay:
                        show_pdf = st.toggle("View Downloaded PDF", key=f"pdf_toggle_pay_{paper_key}", value=True)
                        if show_pdf:
                            pdf_path = _find_project_pdf_path(project_name, paper_filename_pay)
                            if pdf_path and pdf_path.exists():
                                _render_pdf_viewer_with_highlights(
                                    pdf_path,
                                    p,
                                    paper_key,
                                    project_name,
                                    session_id_export,
                                )
                            else:
                                flash(f"PDF file not found: {paper_filename_pay}", level="warning")
